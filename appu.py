import streamlit as st
import csv
import datetime
import pandas as pd
import numpy as np
import datetime
from datetime import timedelta
import xlsxwriter
import openpyxl
from openpyxl import Workbook
from io import BytesIO
from io import StringIO
import os
import time 
import random 

def format_date_with_suffix(date):
    """Formats a date as 'Month Day[st/nd/rd/th], Year' (e.g., 'February 3rd, 2025')."""
    day = date.day
    suffix = "th" if 11 <= day <= 13 else {1: "st", 2: "nd", 3: "rd"}.get(day % 10, "th")
    return date.strftime(f"%B {day}{suffix}, %Y")
	
file_configs = {
    "HAMPDEN_NURSERY.xlsx": {"title": "HAMPDEN NURSERY","custom_text": "CUSTOM_PRINT","names": ["Folaranmi, Oluwamayoda", "Alur, Pradeep", "Nanda, Sharmilarani", "HAMPDEN_NURSERY"]},
    "SJR_HOSP.xlsx": {"title": "SJR HOSPITALIST","custom_text": "CUSTOM_PRINT","names": ["Spangola, Haley", "Gubitosi, Terry", "SJR_1", "SJR_2"]}, 
    "AAC.xlsx": {"title": "AAC","custom_text": "CUSTOM_PRINT","names": ["Vaishnavi Harding", "Abimbola Ajayi", "Shilu Joshi", "Desiree Webb", "Amy Zisa", "Abdullah Sakarcan", "Anna Karasik", "AAC_1", "AAC_2", "AAC_3"]} #LIST ALL NAMES
}

def generate_excel_file(start_date, title, custom_text, file_name, names):
    """
    Generates an Excel file where each week's structure aligns properly.

    Args:
        start_date (datetime): The starting date provided by the user.
        title (str): The text to be placed in cell A1.
        custom_text (str): The text to be placed in cell A2.
        file_name (str): The name of the output file.
        names (list): A list of names to be placed in the file.

    Returns:
        str: Path of the saved file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Place specific text in designated cells
    ws["A1"] = title
    ws["A2"] = custom_text

    # Columns where "custom_value" should be placed
    custom_value_columns = ["A", "C", "E", "G", "I", "K", "M"]
    name_columns = ["B", "D", "F", "H", "J", "L", "N"]

    # Ensure names list has at least one name
    if not names:
        names = ["Default Name ~"]

    # Days of the week to be placed across the row
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

    # Initial row where the first week starts
    start_row = 4
    num_weeks = 5  # Define the number of weeks
    week_height = 13  # Number of rows per week (date row + names + custom_value rows)

    for week in range(num_weeks):  
        current_date = start_date + datetime.timedelta(weeks=week)

        # Place day names and corresponding dates
        for i, day in enumerate(days):
            col_letter = chr(65 + (i * 2))  # Convert to Excel column letters (A, C, E, G, I, K, M)
            ws[f"{col_letter}{start_row}"] = day  
            formatted_date = (current_date + datetime.timedelta(days=i)).strftime("%B %-d, %Y")
            ws[f"{col_letter}{start_row + 1}"] = formatted_date  

        # Start placing names **immediately after the date row**
        names_start_row = start_row + 2  
        names_end_row = names_start_row + len(names)

        for i, col in enumerate(name_columns):
            custom_col = custom_value_columns[i]  # Get the column to the left
            for j, name in enumerate(names):
                row = names_start_row + j
                ws[f"{col}{row}"] = name  # Place the name
                ws[f"{custom_col}{row}"] = "custom_value"  # Place "custom_value" in the left column

        # Fill remaining rows with "custom_value" from the **date row** to the **next week's date row**
        next_week_start = start_row + week_height  # Set end range dynamically
        for i, col in enumerate(custom_value_columns):
            for row in range(start_row + 1, next_week_start):  # Fill from date row up to next week's start row
                if row >= names_end_row:  # Avoid overwriting names
                    ws[f"{col}{row}"] = "custom_value"

        # Move to the next week's section
        start_row = next_week_start  

    # Save the Excel file
    file_path = f"{file_name}"
    wb.save(file_path)

    # ✅ **Display & Download Immediately**
    st.success(f"✅ File '{file_name}' has been successfully created!")

    #df_display = pd.read_excel(file_path, dtype=str)
    #st.dataframe(df_display); #time.sleep(30); # Display file in Streamlit

    with open(file_path, "rb") as f:
        st.download_button("Download Generated Excel File", f, file_name, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"); return file_path  # Return file path for later use
	
# Initialize session state variables efficiently
session_defaults = {
    "page": "Home",
    "start_date": None,
    "uploaded_files": {},
    "uploaded_book4_file": {},
}
for key, value in session_defaults.items():
    st.session_state.setdefault(key, value)

# Function to change page and trigger rerun
def navigate_to(page):
    st.session_state.page = page
    st.rerun()

# Home Page
if st.session_state.page == "Home":
    st.title("Welcome to OPD Creator")
    st.write("Please choose what you'd like to do next.")

    if st.button("Go to Create OPD"):
        navigate_to("Create OPD")

    if st.button("Go to Create Student Schedule"):
        navigate_to("Create Student Schedule")

elif st.session_state.page == "Create OPD":
    st.title('Date Input for OPD')
    st.write('Enter start date in **m/d/yyyy format**, no leading zeros (e.g., 7/6/2021):')

    # User Input for Start Date
    date_input = st.text_input('Start Date')


    if st.button('Submit Date') and date_input:
        try:
            start_date, end_date = datetime.datetime.strptime(date_input, "%m/%d/%Y"), datetime.datetime.strptime(date_input, "%m/%d/%Y") + datetime.timedelta(days=34)
            st.session_state.start_date, st.session_state.end_date = start_date, end_date
		
            st.success(f"✅ Valid date entered: {start_date.strftime('%B %d, %Y')} | 📅 Date range: {start_date.strftime('%B %d, %Y')} ➝ {end_date.strftime('%B %d, %Y')}")

            # Generate all predefined Excel files
            generated_files = {}
            for file_name, config in file_configs.items():
                file_path = generate_excel_file(start_date, config["title"], config["custom_text"], file_name, config["names"])
                generated_files[file_name] = file_path

            # Store file paths in session state for later downloads
            st.session_state.generated_files = generated_files

            # Move to the next page: Upload Files
            st.session_state.page = "Upload Files"
            st.rerun()  # Force rerun to reflect the page change
    
        except ValueError:
            st.error('Invalid date format. Please enter the date in **m/d/yyyy** format.')

elif st.session_state.page == "Upload Files":
    st.title("File Upload Section")
    st.write("Upload the required Excel files:")

    # Define file name mappings based on content identifiers
    file_identifiers = {
        "Academic General Pediatrics": ["NYES.xlsx", "HOPE_DRIVE.xlsx", "ETOWN.xlsx", "PSHCH_NURSERY.xlsx"],
        "Pulmonary": ["WARD_P.xlsx"],
        "Hospitalists": ["WARD_A.xlsx"],
        "Cardiology": ["WARD_CARDIOLOGY.xlsx"],
        "Neph": ["WARD_NEPHRO.xlsx"],
        "PICU": ["PICU.xlsx"],
        "GI Daytime Service": ["WARD_GI.xlsx"],
        "Complex": ["COMPLEX.xlsx"],
        "Adol Med": ["ADOLMED.xlsx"], 
        "cl5rks1p": ["Book4.xlsx"]
    }
    # Required files for validation
    required_files = set(file for filenames in file_identifiers.values() for file in filenames)

    # Streamlit UI
    st.title("File Upload Section")
    st.write("Upload the following required Excel files:")

    # Ensure start_date and end_date exist in session state
    if "start_date" in st.session_state and "end_date" in st.session_state:
        start_date, end_date = st.session_state.start_date, st.session_state.end_date
        st.success(f"✅ Valid date entered: {start_date.strftime('%B %d, %Y')} | 📅 Date range: {start_date.strftime('%B %d, %Y')} ➝ {end_date.strftime('%B %d, %Y')}")
    else:
        st.error("❌ No valid date found. Please enter a start date first.")

    # File uploader
    uploaded_files = st.file_uploader("Choose your files", type="xlsx", accept_multiple_files=True)

    if uploaded_files:
        uploaded_files_dict = {}
        detected_files = set()

        for file in uploaded_files:
            try:
                # Read the first few rows of the Excel file
                df = pd.read_excel(file, dtype=str, nrows=10)  

                # Normalize text: strip spaces, handle line breaks, convert to lowercase
                df_clean = df.astype(str).apply(lambda x: x.str.strip().str.replace("\n", " ").str.lower())

                # Convert all values into a single string for better search
                full_text = " ".join(df_clean.to_string().split()).lower()

                # Assign multiple filenames for "Academic General Pediatrics"
                found_files = []
                for key, expected_filenames in file_identifiers.items():
                    if key.lower() in full_text:
                        found_files.extend(expected_filenames)

                if found_files:
                    for expected_filename in found_files:
                        uploaded_files_dict[expected_filename] = file  # Assign the same file to multiple expected filenames
                        detected_files.add(expected_filename)

                else:
                    st.warning(f"⚠️ Could not automatically detect file type for: {file.name}")

            except Exception as e:
                st.error(f"❌ Error reading {file.name}: {str(e)}")

        # Save detected files to session state
        st.session_state.uploaded_files = uploaded_files_dict

        # Check for missing files
        missing_files = required_files - detected_files

        if not missing_files:
            st.success("✅ All required files uploaded and detected successfully!")
            navigate_to("OPD Creator")
        else:
            st.error(f"❌ Missing files: {', '.join(missing_files)}. Please upload all required files.")

		
elif st.session_state.page == "OPD Creator":
	#test_date = datetime.datetime.strptime(x, "%m/%d/%Y")
	test_date = st.session_state.start_date
	uploaded_files = st.session_state.uploaded_files
	
	# initializing K
	K = 28
	 
	res = []
	 
	for day in range(K):
	    date = (test_date + datetime.timedelta(days = day)).strftime("%-m/%-d/%Y")
	    res.append(date)
	     
	#res
	
	dates = pd.DataFrame(res, columns =['dates'])
	
	dates['x'] = "y"
	
	dates['i'] = dates.index+1
	
	dates['x'] = dates['x'].astype(str)+dates['i'].astype(str)
	
	dates['x'] = dates['x'].astype(str) + "=" + "'"+dates['dates'].astype(str) + "'"
	
	dates = dates[['x']]
	
	dates.to_csv('dates.csv',index=False)
	
	import numpy as np
	
	datesdf = pd.read_csv('dates.csv')
	
	dates = datesdf['x'].astype(str)
	
	numpy_array=dates.to_numpy()
	np.savetxt("dates.py",numpy_array, fmt="%s")
	
	exec(open('dates.py').read())
	import xlsxwriter

	# Create workbook
	workbook = xlsxwriter.Workbook('OPD.xlsx')
	
	# Define worksheet names
	worksheet_names = ['HOPE_DRIVE', 'ETOWN', 'NYES', 'COMPLEX', 'W_A', 'W_C','W_P', 'PICU', 'PSHCH_NURSERY', 'HAMPDEN_NURSERY','SJR_HOSP', 'AAC', 'ER_CONS','NF',"ADOLMED"]
	
	# Create worksheets and store them in a dictionary
	worksheets = {name: workbook.add_worksheet(name) for name in worksheet_names}
	(worksheet, worksheet2, worksheet3, worksheet4, worksheet5, worksheet6, worksheet7, worksheet8, worksheet9, worksheet10, worksheet11, worksheet12, worksheet13, worksheet14,worksheet15) = worksheets.values()
	
	# Define format
	format1 = workbook.add_format({'font_size': 18, 'bold': 1, 'align': 'center','valign': 'vcenter', 'font_color': 'black','bg_color': '#FEFFCC', 'border': 1})
	
	# Define site names corresponding to worksheet names
	worksheet_sites = {worksheet: 'Hope Drive', 
			   worksheet2: 'Elizabethtown', 
			   worksheet3: 'Nyes Road', 
			   worksheet4: 'Complex Care', 
			   worksheet5: 'WARD A', 
			   worksheet6: 'WARD C', 
			   worksheet7: 'WARD P', 
			   worksheet8: 'PICU', 
			   worksheet9: 'PSHCH NURSERY', 
			   worksheet10: 'HAMPDEN NURSERY',
			   worksheet11: 'SJR HOSPITALIST', 
			   worksheet12: 'AAC', 
			   worksheet13: 'ER CONSULTS', 
			   worksheet14: 'NIGHT FLOAT', 
			   worksheet15: 'ADOLMED'}
	
	# Write "Site:" and corresponding site names in each worksheet
	for ws, site in worksheet_sites.items():
	    ws.write(0, 0, 'Site:', format1)
	    ws.write(0, 1, site, format1)
		
	#Color Coding
	format4 = workbook.add_format({'font_size':12,'bold': 1,'align': 'center','valign': 'vcenter','font_color':'black','bg_color':'#8ccf6f','border':1})
	format4a = workbook.add_format({'font_size':12,'bold': 1,'align': 'center','valign': 'vcenter','font_color':'black','bg_color':'#9fc5e8','border':1})    
	format5 = workbook.add_format({'font_size':12,'bold': 1,'align': 'center','valign': 'vcenter','font_color':'black','bg_color':'#FEFFCC','border':1})
	format5a = workbook.add_format({'font_size':12,'bold': 1,'align': 'center','valign': 'vcenter','font_color':'black','bg_color':'#d0e9ff','border':1})
	format11 = workbook.add_format({'font_size':18,'bold': 1,'align': 'center','valign': 'vcenter','font_color':'black','bg_color':'#FEFFCC','border':1})
	
	#H codes
	formate = workbook.add_format({'font_size':12,'bold': 0,'align': 'center','valign': 'vcenter','font_color':'white','border':0})
	
	# HOPE_DRIVE COLOR CODING AND IDENTIFYING ACUTE VERSUS CONTINUITY
	ranges_format1 = ['A8:H15', 'A32:H39', 'A56:H63', 'A80:H87']
	ranges_format5a = ['A18:H25', 'A42:H49', 'A66:H73', 'A90:H97']
	
	for cell_range in ranges_format1:
	    worksheet.conditional_format(cell_range, {'type': 'cell', 'criteria': '>=', 'value': 0, 'format': format1})
	
	for cell_range in ranges_format5a:
	    worksheet.conditional_format(cell_range, {'type': 'cell', 'criteria': '>=', 'value': 0, 'format': format5a})
	
	# HOPE_DRIVE CONDITIONAL FORMATTING
	ranges_format4 = ['A6:H6', 'A7:H7', 'A30:H30', 'A31:H31', 'A54:H54', 'A55:H55', 'A78:H78', 'A79:H79']
	ranges_format4a = ['A16:H16', 'A17:H17', 'A40:H40', 'A41:H41', 'A64:H64', 'A65:H65', 'A88:H88', 'A89:H89']
	
	for cell_range in ranges_format4:
	    worksheet.conditional_format(cell_range, {'type': 'cell', 'criteria': '>=', 'value': 0, 'format': format4})
	
	for cell_range in ranges_format4a:
	    worksheet.conditional_format(cell_range, {'type': 'cell', 'criteria': '>=', 'value': 0, 'format': format4a})
	
	# HOPE_DRIVE WRITING ACUTE AND CONTINUITY LABELS
	acute_format_ranges = [
	    (6, 7, 'AM - ACUTES', format4), (16, 17, 'PM - ACUTES', format4a), 
	    (30, 31, 'AM - ACUTES', format4), (40, 41, 'PM - ACUTES', format4a),
	    (54, 55, 'AM - ACUTES', format4), (64, 65, 'PM - ACUTES', format4a),
	    (78, 79, 'AM - ACUTES', format4), (88, 89, 'PM - ACUTES', format4a)
	]
	
	continuity_format_ranges = [
	    (8, 15, 'AM - Continuity', format5a), (18, 25, 'PM - Continuity', format5a),
	    (32, 39, 'AM - Continuity', format5a), (42, 49, 'PM - Continuity', format5a),
	    (56, 63, 'AM - Continuity', format5a), (66, 73, 'PM - Continuity', format5a),
	    (80, 87, 'AM - Continuity', format5a), (90, 97, 'PM - Continuity', format5a)
	]
	
	# Write Acute Labels
	for start_row, end_row, label, fmt in acute_format_ranges:
	    for row in range(start_row, end_row + 1):
	        worksheet.write(f'A{row}', label, fmt)
	
	# Write Continuity Labels
	for start_row, end_row, label, fmt in continuity_format_ranges:
	    for row in range(start_row, end_row + 1):
	        worksheet.write(f'A{row}', label, fmt)
	
	# Define the labels
	#labels = ['HX1', 'HX2', 'H2', 'H3', 'H4', 'H5', 'H6', 'H7', 'H8', 'H9', 'HXX1', 'HXX2', 'H12', 'H13', 'H14', 'H15', 'H16', 'H17', 'H18', 'H19']
	
	labels = ['H{}'.format(i) for i in range(20)]
	
	# Define the starting rows for each group
	start_rows = [6, 30, 54, 78]
	
	# Write the labels in each group
	for start_row in start_rows:
	    for i, label in enumerate(labels):
	        worksheet.write(f'I{start_row + i}', label, formate)
	
	# Simplify common formatting and label assignment for worksheets 2, 3, 4, 5
	worksheets = [worksheet2, worksheet3, worksheet4, worksheet5, worksheet6, worksheet7, worksheet8, worksheet9, worksheet10, worksheet11, worksheet12, worksheet13, worksheet14, worksheet15]
	
	ranges_format1 = ['A6:H15', 'A30:H39', 'A54:H63', 'A78:H87']
	ranges_format5a = ['A16:H25', 'A40:H49', 'A64:H73', 'A88:H97']
	specific_format_ranges = [
	    ('B6:H6', format4), ('B16:H16', format4a),
	    ('B30:H30', format4), ('B40:H40', format4a),
	    ('B54:H54', format4), ('B64:H64', format4a),
	    ('B78:H78', format4), ('B88:H88', format4a)
	]
	
	am_pm_labels = ['AM'] * 10 + ['PM'] * 10
	h_labels = ['H{}'.format(i) for i in range(20)]
	
	for worksheet in worksheets:
	    # Apply conditional formatting
	    for cell_range in ranges_format1:
	        worksheet.conditional_format(cell_range, {'type': 'cell', 'criteria': '>=', 'value': 0, 'format': format1})
	
	    for cell_range in ranges_format5a:
	        worksheet.conditional_format(cell_range, {'type': 'cell', 'criteria': '>=', 'value': 0, 'format': format5a})
	
	    for cell_range, fmt in specific_format_ranges:
	        worksheet.conditional_format(cell_range, {'type': 'cell', 'criteria': '>=', 'value': 0, 'format': fmt})
	
	    # Write AM/PM labels
	    sections = [(6, 25), (30, 49), (54, 73), (78, 97)]
	    for start_row, end_row in sections:
	        for i, label in enumerate(am_pm_labels):
	            worksheet.write(f'A{start_row + i}', label, format5a)
	
	    # Write H labels in column 'I'
	    start_rows = [6, 30, 54, 78]
	    for start_row in start_rows:
	        for i, label in enumerate(h_labels):
	            worksheet.write(f'I{start_row + i}', label, formate)
	
	# Loop through each worksheet in workbook
	for worksheet in workbook.worksheets():
	
	    # Set Zoom for all sheets
	    worksheet.set_zoom(80)
	
	    # Set Days
	    format3 = workbook.add_format({
	        'font_size': 12, 'bold': 1, 'align': 'center', 'valign': 'vcenter',
	        'font_color': 'black', 'bg_color': '#FFC7CE', 'border': 1
	    })
	    day_labels = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
	    start_rows = [2, 26, 50, 74] #[3, 27, 51, 75]
	    for start_row in start_rows:
	        for i, day in enumerate(day_labels):
	            worksheet.write(start_row, 1 + i, day, format3)  # B=1, C=2, etc.
	
	    # Set Date Formats
	    format_date = workbook.add_format({
	        'num_format': 'm/d/yyyy', 'font_size': 12, 'bold': 1, 'align': 'center', 'valign': 'vcenter',
	        'font_color': 'black', 'bg_color': '#FFC7CE', 'border': 1
	    })
	
	    format_label = workbook.add_format({
	        'font_size': 12, 'bold': 1, 'align': 'center', 'valign': 'vcenter',
	        'font_color': 'black', 'bg_color': '#FFC7CE', 'border': 1
	    })
	
	    # Set Date Formulas
	    date_rows = [3, 27, 51, 75] #[4, 28, 52, 76]
	    for i, start_row in enumerate(date_rows):
	        worksheet.write(f'A{start_row - 1}', "", format_label)
	        #worksheet.write_formula(f'A{start_row}', f'="Week of:"&" "&TEXT(B{start_row},"m/d/yy")', format_label) #If want to place Week of Date in
	        worksheet.write_formula(f'A{start_row}', f'=""', format_label)
	        worksheet.write(f'A{start_row + 1}', "", format_label)
	
	    # Set Pink Bars (Conditional Format)
	    pink_bar_rows = [5, 29, 53, 77]
	    for row in pink_bar_rows:
	        worksheet.conditional_format(f'A{row}:H{row}', {
	            'type': 'cell', 'criteria': '>=', 'value': 0, 'format': format_label
	        })
	
	    # Black Bars
	    format2 = workbook.add_format({'bg_color': 'black'})
	    black_bar_rows = [2, 26, 50, 74, 98]
	    for row in black_bar_rows:
	        worksheet.merge_range(f'A{row}:H{row}', " ", format2)
	        
	    # Write More Dates
	    date_values = [
	        [y1, y2, y3, y4, y5, y6, y7],
	        [y8, y9, y10, y11, y12, y13, y14],
	        [y15, y16, y17, y18, y19, y20, y21],
	        [y22, y23, y24, y25, y26, y27, y28]
	    ]
	    for i, start_row in enumerate(date_rows):
	        for j, value in enumerate(date_values[i]):
	            worksheet.write(start_row, 1 + j, value, format_date)  # B=1, C=2, etc.
	
	    # Set Column Widths
	    worksheet.set_column('A:A', 10)
	    worksheet.set_column('B:H', 65)
	    worksheet.set_row(0, 37.25)
	
	    # Merge Format for Text
	    merge_format = workbook.add_format({
	        'bold': 1, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True,
	        'font_color': 'red', 'bg_color': '#FEFFCC', 'border': 1
	    })
	    text1 = 'Students are to alert their preceptors when they have a Clinical Reasoning Teaching Session (CRTS).  Please allow the students to leave approximately 15 minutes prior to the start of their session so they can be prepared to actively participate.  ~ Thank you!'
	
	    # Merge and Write Important Message
	    worksheet.merge_range('C1:F1', text1, merge_format)
	    worksheet.write('G1', "", merge_format)
	    worksheet.write('H1', "", merge_format)
	
	# Close Workbook
	workbook.close()
	
	####################################################################################################################################
	import pandas as pd
	import datetime
	from datetime import timedelta
	
	# Disable chained assignment warning
	pd.options.mode.chained_assignment = None
	
	start_date = test_date.replace(hour=0, minute=0, second=0, microsecond=0)
	end_date = start_date + timedelta(days=34)
	
	# Create DataFrame with formatted dates
	xf201 = pd.DataFrame({'date': pd.date_range(start=start_date, end=end_date)})
	xf201['convert'] = xf201['date'].dt.strftime('%B %-d, %Y')
	
	# Generate 'T' column
	xf201['T'] = "T" + xf201.index.astype(str)
		
	for i in range(35):
	    exec(f"day{i} = xf201['convert'][{i}]")

	column_pairs = [(0, 1), (2, 3), (4, 5), (6, 7), (8, 9), (10, 11), (12, 13)]
	
	days = [day0, day1, day2, day3, day4, day5, day6, day7, day8, day9, day10, day11, day12, day13, day14, day15, day16, day17, day18, day19, day20, day21, day22, day23, day24, day25, day26, day27, day28, day29, day30, day31, day32, day33, day34]
	
	# Function to process each file

	def process_file(file_key, clinic_name, replacements=None, df=None):
	    """Process a file (either uploaded or generated) and return a cleaned DataFrame."""
	    
	    # 1️⃣ **Use the provided DataFrame if already passed**
	    if df is not None:
	        print(f"Processing provided DataFrame for {clinic_name}...")
	    
	    else:
	        # 2️⃣ **Check if the locally generated file exists**
	        local_file_path = f"{file_key}"
	        if os.path.exists(local_file_path):
	            print(f"Found locally generated file: {local_file_path}. Using it for {clinic_name}...")
	            df = pd.read_excel(local_file_path, dtype=str)
	        
	        # 3️⃣ **Otherwise, fall back to uploaded file**
	        elif file_key in uploaded_files:
	            print(f"Using uploaded file for {clinic_name}...")
	            df = pd.read_excel(uploaded_files[file_key], dtype=str)
	        
	        else:
	            print(f"❌ ERROR: No file found for {clinic_name} ({file_key}). Skipping...")
	            return None  # Handle missing file case
	    
	    # ✅ **Continue normal processing**
	    df.rename(columns={col: str(i) for i, col in enumerate(df.columns)}, inplace=True)
	
	    D_dict = {}
	    for i in range(28):
	        col_idx = column_pairs[i % len(column_pairs)]
	        start_day = days[i]
	        end_day = days[i + 7]
	
	        start_idx = df.loc[df[str(col_idx[0])] == start_day].index[0]
	        end_idx = df.loc[df[str(col_idx[0])] == end_day].index[0]
	
	        extracted_data = df.iloc[start_idx + 1:end_idx, list(col_idx)].copy()
	        extracted_data.columns = ['type', 'provider']
	        extracted_data.insert(0, 'date', start_day)
	        extracted_data = extracted_data[:-1]
	
	        D_dict[f"D{i}"] = extracted_data
	
	    dfx = pd.concat(D_dict.values(), ignore_index=True)
	    dfx['clinic'] = clinic_name
	
	    # ✅ **Apply replacements if provided**
	    if replacements:
	        dfx = dfx.replace(replacements, regex=True)
	
	    # ✅ **Save the cleaned file**
	    filename = f"{clinic_name.lower()}.csv"
	    dfx.to_csv(filename, index=False)
	
	    print(f"✅ Processed {clinic_name} and saved to {filename}")
	    return dfx  # Return DataFrame for further processing

	def duplicate_am_continuity(df, clinic_name, special_cases=None):
	    """
	    Duplicates "AM - Continuity" rows as "PM - Continuity" for normal cases.
	    If the clinic is in `special_cases`, only duplicates AM - Continuity and renames it to PM - Continuity.
	    """
	    if df is not None and not df.empty:
	        # Ensure special_cases is a set for fast lookup
	        special_cases = special_cases or set()
	
	        if clinic_name in special_cases:
	            # Special handling: Only copy AM - Continuity and rename it
	            am_continuity_rows = df[df.eq("AM - Continuity ").any(axis=1)].copy()
	
	            if am_continuity_rows.empty:
	                print(f"⚠️ No AM - Continuity rows found in {clinic_name}. Skipping special processing.")
	                return df  # No modifications needed
	            
	            # Replace AM -> PM
	            pm_continuity_rows = am_continuity_rows.replace("AM - Continuity ", "PM - Continuity ")
	
	            # Only append the PM version, no double duplication
	            df = pd.concat([df, pm_continuity_rows], ignore_index=True).reset_index(drop=True); #df = pd.concat([df, pm_continuity_rows], ignore_index=True).sort_values(by=["date", "provider"]).reset_index(drop=True)
	
	            print(f"✅ Special processing for {clinic_name}: Only duplicated AM - Continuity as PM - Continuity.")
	
	        else:
	            # Default behavior (original functionality)
	            am_continuity_rows = df[df.eq("AM - Continuity ").any(axis=1)].copy()
	            pm_continuity_rows = am_continuity_rows.replace("AM - Continuity ", "PM - Continuity ")
	
	            df = pd.concat([df, df, pm_continuity_rows, pm_continuity_rows], ignore_index=True).reset_index(drop=True); #df = pd.concat([df, df, pm_continuity_rows, pm_continuity_rows], ignore_index=True).sort_values(by=["date", "provider"]).reset_index(drop=True)
	
	            print(f"✅ Standard processing for {clinic_name}: Fully duplicated AM - Continuity.")
	
	        # Save the updated data
	        filename = f"{clinic_name.lower()}.csv"
	        df.to_csv(filename, index=False)
	
	    return df  # Return modified DataFrame

	def process_continuity_classes(df, clinic_name, am_csv, pm_csv):
	    if df is not None:
	        # Process AM - Continuity
	        am_df = df[df['type'] == 'AM - Continuity '].assign(count=lambda x: x.groupby(['date'])['provider'].cumcount()).assign(**{"class": lambda x: "H" + x['count'].astype(str)})[['date', 'type', 'provider', 'clinic', 'class']]
	        
	        # Process PM - Continuity
	        pm_df = df[df['type'] == 'PM - Continuity '].assign(count=lambda x: x.groupby(['date'])['provider'].cumcount()+10).assign(**{"class": lambda x: "H" + x['count'].astype(str)})[['date', 'type', 'provider', 'clinic', 'class']]
	        
	        # Save to CSV
	        am_df.to_csv(am_csv, index=False)
	        pm_df.to_csv(pm_csv, index=False)
	
	        # Display in Streamlit
	        #st.write(f"### {clinic_name} - AM Continuity Assignments")
	        #st.dataframe(am_df)  # Display AM - Continuity table
	
	        #st.write(f"### {clinic_name} - PM Continuity Assignments")
	        #st.dataframe(pm_df)  # Display PM - Continuity table
	
	        return am_df, pm_df  # Return processed DataFrames for further use if needed
		    
	def process_hope_classes(df, clinic_name):
	    """
	    Processes Hope Drive's different continuity and acute types by assigning a count and class.
	    Saves the resulting DataFrame to separate CSV files based on type.
	    """
	    if df is not None:
	        hope_files = {
	            "AM - ACUTES": "5.csv",
	            "AM - ACUTES ": "6.csv",  # Handles potential trailing space issue
	            "PM - ACUTES ": "7.csv",
	            "AM - Continuity ": "8.csv",
	            "PM - Continuity ": "9.csv"
	        }
	
	        for type_key, filename in hope_files.items():
	            if type_key in df['type'].values:
	                subset_df = df[df['type'] == type_key].copy()
	
	                # Assign custom count logic based on type
	                if "AM - ACUTES" in type_key:
	                    subset_df['count'] = subset_df.groupby(['date'])['provider'].cumcount()
	                    subset_df['class'] = subset_df['count'].apply(
	                        lambda count: "H0" if count == 0 else ("H1" if count == 1 else "H" + str(count + 2))
	                    )
	                
	                elif "PM - ACUTES" in type_key:
	                    subset_df['count'] = subset_df.groupby(['date'])['provider'].cumcount()
	                    subset_df['class'] = subset_df['count'].apply(
	                        lambda count: "H10" if count == 0 else ("H11" if count == 1 else "H" + str(count + 12))
	                    )
	
	                elif "AM - Continuity" in type_key:
	                    subset_df['count'] = subset_df.groupby(['date'])['provider'].cumcount() + 2
	                    subset_df['class'] = "H" + subset_df['count'].astype(str)
	
	                elif "PM - Continuity" in type_key:
	                    subset_df['count'] = subset_df.groupby(['date'])['provider'].cumcount() + 12
	                    subset_df['class'] = "H" + subset_df['count'].astype(str)
	
	                # Keep only relevant columns
	                subset_df = subset_df[['date', 'type', 'provider', 'clinic', 'class']]
	
	                # Save to CSV
	                subset_df.to_csv(filename, index=False)
	                print(f"{clinic_name} {type_key} saved to {filename}.")

	# Define replacement rules for each clinic
	replacement_rules = {
	    "HOPE_DRIVE.xlsx": {
	        "Hope Drive AM Continuity": "AM - Continuity",
	        "Hope Drive PM Continuity": "PM - Continuity",
	        "Hope Drive\xa0AM Acute Precept ": "AM - ACUTES",  # Handles non-breaking space (\xa0)
	        "Hope Drive PM Acute Precept": "PM - ACUTES",
	        "Hope Drive Weekend Continuity": "AM - Continuity",
	        "Hope Drive Weekend Acute 1": "AM - ACUTES",
	        "Hope Drive Weekend Acute 2": "AM - ACUTES"
	    },
	    "PICU.xlsx": {"2nd PICU Attending 7:45a-4p": "AM - Continuity", "1st PICU Attending 7:30a-5p": "AM - Continuity"},
	    "ETOWN.xlsx": {"Etown AM Continuity": "AM - Continuity", "Etown PM Continuity": "PM - Continuity"},
	    "NYES.xlsx": {"Nyes Rd AM Continuity": "AM - Continuity", "Nyes Rd PM Continuity": "PM - Continuity"},
	    "COMPLEX.xlsx": {"Hope Drive Clinic AM": "AM - Continuity", "Hope Drive Clinic PM": "PM - Continuity"},
	    "WARD_A.xlsx": {"Rounder 1 7a-7p": "AM - Continuity", "Rounder 2 7a-7p": "AM - Continuity", "Rounder 3 7a-7p": "AM - Continuity", "Night Call 9p-7a": "night_float", "AM Pager 7a-12p": "consultsa", "PM Pager 12p-4p":"consultsp", "Evening Pager 4p-9p":"consultsp", "Overnight Pager 9p-7a":"consultsp"}, #Assume Day Admitting is Consults
	    "WARD_P.xlsx": {"On-Call 8a-8a": "AM - Continuity", "On-Call": "AM - Continuity"},
	    "PSHCH_NURSERY.xlsx": {"Nursery Weekday 8a-6p": "AM - Continuity", "Nursery Weekend": "AM - Continuity"},
	    "HAMPDEN_NURSERY.xlsx": {"custom_value": "AM - Continuity "},  # Replace "custom_value" with "AM - Continuity" (must add space!)
	    "SJR_HOSP.xlsx": {"custom_value": "AM - Continuity "},  # Same format as HAMPDEN_NURSERY.xlsx
	    "AAC.xlsx": {"custom_value": "AM - Continuity "},  # Same format as HAMPDEN_NURSERY.xlsx
	    "WARD_CARDIOLOGY.xlsx": {"Wards 8a-5p": "AM - Continuity", "Wards 8a-8a": "AM - Continuity"},  
	    "WARD_GI.xlsx": {"GI Daytime Service 7:30a-5p": "AM - Continuity", "GI Daytime Service 7:30a-3p": "AM - Continuity", "GI Weekend Call 7:30a-7:30a": "AM - Continuity"},  
	    "WARD_NEPHRO.xlsx": {"Neph On Call 8a-8a": "AM - Continuity"},  
	    "ADOLMED.xlsx": {"Briarcrest Clinic AM": "AM - Continuity", "Briarcrest Clinic PM": "PM - Continuity"},  
	    "Book4.xlsx": {"": "", "": ""},  
	}	

	# Process each file
	hope_drive_df = process_file("HOPE_DRIVE.xlsx", "HOPE_DRIVE", replacement_rules.get("HOPE_DRIVE.xlsx"))
	etown_df = process_file("ETOWN.xlsx", "ETOWN", replacement_rules.get("ETOWN.xlsx"))
	nyes_df = process_file("NYES.xlsx", "NYES", replacement_rules.get("NYES.xlsx"))
	complex_df = process_file("COMPLEX.xlsx", "COMPLEX", replacement_rules.get("COMPLEX.xlsx"))
	
	warda_df = process_file("WARD_A.xlsx", "WARD_A", replacement_rules.get("WARD_A.xlsx"))
	wardp_df = process_file("WARD_P.xlsx", "WARD_P", replacement_rules.get("WARD_P.xlsx"))
	pshchnursery_df = process_file("PSHCH_NURSERY.xlsx", "PSHCH_NURSERY", replacement_rules.get("PSHCH_NURSERY.xlsx"))
	hampdennursery_df = process_file("HAMPDEN_NURSERY.xlsx", "HAMPDEN_NURSERY", replacement_rules.get("HAMPDEN_NURSERY.xlsx"))
	sjrhosp_df = process_file("SJR_HOSP.xlsx", "SJR_HOSP", replacement_rules.get("SJR_HOSP.xlsx"))
	aac_df = process_file("AAC.xlsx", "AAC", replacement_rules.get("AAC.xlsx"))
	
	nf_df = warda_df[warda_df["type"] == "night_float "].assign(type="PM - Continuity ", clinic="NF")
	
	consults_df = warda_df[warda_df["type"].isin(["consultsp ", "consultsa "])].assign(type=lambda df: df["type"].map({"consultsp ": "PM - Continuity ", "consultsa ": "AM - Continuity "}), clinic="ER_CONS")
	consults_df = consults_df.groupby(["date", "type"], as_index=False).agg({"provider": lambda x: "/".join(x) + " ~" if "PM - Continuity " in x.name else "/".join(x)})
	consults_df["clinic"] = "ER_CONS"
	
	adolmed_df = process_file("ADOLMED.xlsx", "ADOLMED", replacement_rules.get("ADOLMED.xlsx"))
	adolmed_df = adolmed_df[adolmed_df["provider"] == "Shook, Jennifer"] #Only Extract Jennifer Shook

	#Combine Ward C Together
	wcard_df = process_file("WARD_CARDIOLOGY.xlsx", "WARD_CARDIOLOGY", replacement_rules.get("WARD_CARDIOLOGY.xlsx"))
	wgi_df = process_file("WARD_GI.xlsx", "WARD_GI", replacement_rules.get("WARD_GI.xlsx"))
	wnephro_df = process_file("WARD_NEPHRO.xlsx", "WARD_NEPHRO", replacement_rules.get("WARD_NEPHRO.xlsx"))
	
	wardc_df = (pd.concat([wcard_df, wgi_df, wnephro_df], ignore_index=True).query("type == 'AM - Continuity '").assign(clinic="WARD_C").groupby(["date", "clinic"], as_index=False).agg({"type": "first", "provider": lambda x: "/".join(x)}))

	picu_df = process_file("PICU.xlsx", "PICU", replacement_rules.get("PICU.xlsx"))
	
	special_clinics = {"AAC","HAMPDEN_NURSERY","SJR_HOSP"}
	
	process_hope_classes(hope_drive_df, "HOPE_DRIVE")
	
	# Apply AM → PM Continuity Transformation... df and the name
	warda_df = duplicate_am_continuity(warda_df, "WARD_A")
	wardp_df = duplicate_am_continuity(wardp_df, "WARD_P")
	picu_df = duplicate_am_continuity(picu_df, "PICU")
	pshchnursery_df = duplicate_am_continuity(pshchnursery_df, "PSHCH_NURSERY")
	hampdennursery_df = duplicate_am_continuity(hampdennursery_df, "HAMPDEN_NURSERY", special_clinics)
	sjrhosp_df = duplicate_am_continuity(sjrhosp_df, "SJR_HOSP")
	aac_df = duplicate_am_continuity(aac_df, "AAC", special_clinics)
	nf_df = duplicate_am_continuity(nf_df, "NF")
	
	wardc_df = duplicate_am_continuity(wardc_df, "WARD_C") 

	process_continuity_classes(etown_df, "ETOWN", "1.csv", "2.csv")
	process_continuity_classes(nyes_df, "NYES", "3.csv", "4.csv")
	process_continuity_classes(complex_df, "COMPLEX", "10.csv", "11.csv")
	
	process_continuity_classes(warda_df, "WARD_A", "12.csv", "13.csv")
	process_continuity_classes(wardp_df, "WARD_P", "14.csv", "15.csv")
	process_continuity_classes(picu_df, "PICU", "16.csv", "17.csv")
	process_continuity_classes(pshchnursery_df, "PSHCH_NURSERY", "18.csv", "19.csv")
	process_continuity_classes(hampdennursery_df, "HAMPDEN_NURSERY", "20.csv", "21.csv")
	process_continuity_classes(sjrhosp_df, "SJR_HOSP", "22.csv", "23.csv")
	process_continuity_classes(aac_df, "AAC", "24.csv", "25.csv")
	process_continuity_classes(nf_df, "NF", "26.csv", "27.csv")
	process_continuity_classes(consults_df, "ER_CONS", "28.csv", "29.csv")
	process_continuity_classes(adolmed_df, "ADOLMED", "30.csv", "31.csv")
	process_continuity_classes(wardc_df, "WARD_C", "32.csv", "33.csv")

	############################################################################################################################
	tables = {f"t{i}": pd.read_csv(f"{i}.csv") for i in range(1, 34)}
	t1, t2, t3, t4, t5, t6, t7, t8, t9, t10, t11, t12, t13, t14, t15, t16, t17, t18, t19, t20, t21, t22, t23, t24, t25, t26, t27, t28, t29, t30, t31, t32, t33 = tables.values()
	
	final2 = pd.DataFrame(columns=t1.columns)
	final2 = pd.concat([final2] + list(tables.values()), ignore_index=True)
	final2.to_csv('final2.csv',index=False); #st.dataframe(final2)
	
	df=pd.read_csv('final2.csv',dtype=str) #MAP to Final2
	
	df['date'] = pd.to_datetime(df['date'])
	df['date'] = df['date'].dt.strftime('%m/%d/%Y')
	
	import csv
	
	dateMAP = xf201[['date','T']]
	
	dateMAP['date'] = pd.to_datetime(dateMAP['date'])
	dateMAP['date'] = dateMAP['date'].dt.strftime('%m/%d/%Y')
	
	dateMAP.to_csv('xxxDATEMAP.csv',index=False)
	
	mydict = {}
	with open('xxxDATEMAP.csv', mode='r')as inp:     #file is the objects you want to map. I want to map the IMP in this file to diagnosis.csv
		reader = csv.reader(inp)
		df1 = {rows[0]:rows[1] for rows in reader} 
	
	df['datecode'] = df.date.map(df1)               #'type' is the new column in the diagnosis file. 'encounter_id' is the key you are using to MAP 

	df['student'] = ""

	# Define the mapping for missing providers
	team_mapping = {'H0': 'WARD_A_Team 1', 'H1': 'WARD_A_Team 2', 'H2': 'WARD_A_Team 3','H3': 'WARD_A_Team 1', 'H4': 'WARD_A_Team 2', 'H5': 'WARD_A_Team 3','H10': 'WARD_A_Team 1', 'H11': 'WARD_A_Team 2', 'H12': 'WARD_A_Team 3','H13': 'WARD_A_Team 1', 'H14': 'WARD_A_Team 2', 'H15': 'WARD_A_Team 3'}

	all_t_values = [f'T{i}' for i in range(28)]
	all_classes = list(team_mapping.keys())
	
	# Create a complete set of WARD_A entries
	df_ward_a = pd.DataFrame([(t, h) for t in all_t_values for h in all_classes], columns=['datecode', 'class'])
	df_ward_a['clinic'] = 'WARD_A'
	
	# Merge with existing data
	df = df.merge(df_ward_a, on=['clinic', 'datecode', 'class'], how='outer')
	
	# Apply the mapping only for clinic "WARD_A" where provider is missing
	df.loc[(df['clinic'] == 'WARD_A') & (df['provider'].isna()), 'provider'] = df['class'].map(team_mapping)

	mydict = {}
	with open('xxxDATEMAP.csv', mode='r')as inp:     #file is the objects you want to map. I want to map the IMP in this file to diagnosis.csv
		reader = csv.reader(inp)
		df1 = {rows[1]:rows[0] for rows in reader} 
	
	df['date'] = df.datecode.map(df1)               #'type' is the new column in the diagnosis file. 'encounter_id' is the key you are using to MAP 

	df.to_csv('final2.csv', index=False)

	df = pd.read_csv('final2.csv',dtype=str) 
	
	list_df = pd.read_excel(uploaded_files['Book4.xlsx']); student_names = list_df["Student Name:"].dropna().astype(str).str.strip(); student_names = student_names[student_names != ""]; unique_student_names = sorted(student_names.unique()); random.shuffle(unique_student_names); st.write(", ".join(unique_student_names))
	
	# Extract the minimum date
	min_date = df['date'].min()

	# Ensure date column is in datetime format (strip timestamps)
	df['date'] = pd.to_datetime(df['date'], errors='coerce').dt.date  
	
	# Filter for WARD_A and exclude providers with class H5 and H15
	df_filtered = df[(df['clinic'] == 'WARD_A') & (~df['class'].isin(['H5', 'H15']))].copy()
	
	# Compute the Monday start of each week
	df_filtered['week_start'] = df_filtered['date'] - pd.to_timedelta(df_filtered['date'].apply(lambda x: x.weekday()), unit='D')
	unique_weeks = sorted(df_filtered['week_start'].unique())
	
	# Define class groups mapping (ordered so we fill one group at a time)
	class_groups = [('H0', 'H10'), ('H1', 'H11'),('H2', 'H12'), ('H3', 'H13'), ('H4', 'H14')]
	
	# Shuffle students before assigning (so they are not in a fixed order)
	random.shuffle(unique_student_names)
	
	# Track assigned students to ensure they are only used once (globally)
	assigned_students = set()
	total_students = len(unique_student_names)
	alert_triggered = False  # Flag to detect if no student was available
	
	# Assign students **by group first** instead of filling whole weeks at once
	for class_group in class_groups:
	    for week_start in unique_weeks:
	        # Select students for this group (only unassigned students)
	        available_students = [s for s in unique_student_names if s not in assigned_students]
	        
	        # If not enough students are left, trigger an alert and stop assigning
	        if len(available_students) < 1:
	            alert_triggered = True  # No students left to assign
	            break  # Stop assignment process
	
	        selected_student = available_students[0]  # Take one student for this group
	        assigned_students.add(selected_student)  # Mark as assigned
	
			        # Assign student to all classes in this group for that week
		for class_type in class_group:
		    # Ensure 'date' is in datetime format
		    df["date"] = pd.to_datetime(df["date"])
		
		    # Compute week start separately to avoid chained operations
		    df["week_start"] = df["date"] - pd.to_timedelta(df["date"].dt.weekday, unit="D")
		
		    # Create filter condition
		    class_filter = (
		        (df["class"] == class_type) &
		        (df["clinic"] == "WARD_A") &  # ✅ Ensure only WARD_A is assigned
		        (df["week_start"] == week_start) &  # ✅ Match the calculated week start
		        (df["date"].dt.weekday < 5)  # ✅ Exclude Saturday (5) & Sunday (6)
		    )
		
		    # Assign student where conditions match
		    df.loc[class_filter, "student"] = selected_student
		
			
	# Alert if no students were available for assignment
	if alert_triggered:
	    st.warning("⚠️ Not enough students to complete assignments! Some providers may be unassigned.")
	
	df['text'] = df['provider'] + " ~ " + df['student']

	df = df.loc[:, ('date','type','provider','student','clinic','text','class','datecode')]
	
	df.to_csv('final.csv',index=False); #st.dataframe(df)
	
	################################################################################################################################################################################################
	
	# ✅ Load dataset
	df = pd.read_csv('final.csv')
	
	# ✅ Convert date to datetime and strip timestamps
	df['date'] = pd.to_datetime(df['date'], errors='coerce').dt.date  
	
	# ✅ Ensure `week_start` exists in the main dataframe BEFORE filtering
	df['week_start'] = df['date'] - pd.to_timedelta(df['date'].apply(lambda x: x.weekday()), unit='D')
	
	# ✅ Define the weeks and class groups for each clinic
	unique_weeks = sorted(df['week_start'].unique())
	
	sjr_hosp_groups = [('H2', 'H12'), ('H3', 'H13')]
	hampden_nursery_groups = [('H3', 'H13')]
	pshch_nursery_groups = [('H0', 'H10'), ('H1', 'H11')]
	
	# ✅ Track assigned students to avoid duplicates
	assigned_students = set()
	
	### **1️⃣ Assign Students to `SJR_HOSP` First**
	for week_start in unique_weeks:
	    available_students = [s for s in unique_student_names if s not in assigned_students]
	    
	    if not available_students:
	        continue  # ✅ Skip if no students left
	
	    # ✅ Exclude students already assigned to WARD_A in the same week
	    unavailable_students = set(df[(df['clinic'] == 'WARD_A') & (df['week_start'] == week_start)]['student'].dropna())
	
	    # ✅ Keep only students not assigned in WARD_A that week
	    available_students = [s for s in available_students if s not in unavailable_students]
	
	    # ✅ Shuffle students to distribute fairly
	    random.shuffle(available_students)
	    
	    for class_group in sjr_hosp_groups:
	        if not available_students:
	            break  # ✅ Stop if no students left
	
	        selected_student = available_students.pop(0)  # Take one student
	        assigned_students.add(selected_student)  # Mark as assigned
	
	        class_filter = df['class'].isin(class_group) & \
	                       (df['clinic'] == 'SJR_HOSP') & \
	                       (df['week_start'] == week_start) & \
	                       (df['date'].apply(lambda x: x.weekday()) < 5)  # ✅ Exclude Saturday & Sunday
	
	        df.loc[class_filter, 'student'] = selected_student
	
	### **2️⃣ Assign Students to `HAMPDEN_NURSERY` (H3/H13)**
	for week_start in unique_weeks:
	    available_students = [s for s in unique_student_names if s not in assigned_students]
	    
	    if not available_students:
	        continue
	
	    for class_group in hampden_nursery_groups:
	        if not available_students:
	            break
	
	        selected_student = available_students.pop(0)
	        assigned_students.add(selected_student)
	
	        class_filter = df['class'].isin(class_group) & \
	                       (df['clinic'] == 'HAMPDEN_NURSERY') & \
	                       (df['week_start'] == week_start) & \
	                       (df['date'].apply(lambda x: x.weekday()) < 5)
	
	        df.loc[class_filter, 'student'] = selected_student
	
	### **3️⃣ Assign Remaining Students to `PSHCH_NURSERY` (H0/H10, H1/H11)**
	for week_start in unique_weeks:
	    available_students = [s for s in unique_student_names if s not in assigned_students]
	    
	    if not available_students:
	        continue
	
	    for class_group in pshch_nursery_groups:
	        if not available_students:
	            break
	
	        selected_student = available_students.pop(0)
	        assigned_students.add(selected_student)
	
	        class_filter = df['class'].isin(class_group) & \
	                       (df['clinic'] == 'PSHCH_NURSERY') & \
	                       (df['week_start'] == week_start) & \
	                       (df['date'].apply(lambda x: x.weekday()) < 5)
	
	        df.loc[class_filter, 'student'] = selected_student
	
	# ✅ Create a text column for easier viewing
	df['text'] = df['provider'].fillna("").astype(str) + " ~ " + df['student'].fillna("").astype(str)
	
	# ✅ Save and display the updated dataset
	df.to_csv('final.csv', index=False)
	
	################################################################################################################################################################################################
	df['student'] = df['student'].astype(str).str.strip()  # Convert to string & strip spaces
	df['student'].replace("nan", pd.NA, inplace=True)
	df_filtered = df.dropna(subset=['student'])
	
	#df_filtered = df[df['student'].ne("") & df['student'].ne("nan")]  # Exclude empty & 'nan'

	# ✅ Find duplicate student assignments across all clinics
	duplicate_students = df_filtered[df_filtered.duplicated(subset=['datecode', 'class', 'student'], keep=False)]
	
	if not duplicate_students.empty:
	    st.warning("⚠️ Duplicate student assignments found across different clinics!")
	
	    # ✅ Group by student, datecode, class, and clinic
	    duplicate_summary = (
	        duplicate_students.groupby(['student', 'datecode', 'class', 'clinic'])
	        .size()
	        .reset_index(name='Count')
	    )
	
	    # ✅ Show only problematic cases (students assigned to multiple clinics in the same datecode and class)
	    duplicate_summary = duplicate_summary[duplicate_summary.duplicated(subset=['student', 'datecode', 'class'], keep=False)]


	    st.write('Duplicate Check')
	    st.dataframe(duplicate_summary)
		
	else:
	    st.success("✅ No duplicate student assignments detected across clinics!")

	#####################################################################OUTPATIENT SHIFT ANALYIS#####################################################################################################################################
	clinics_of_interest = ["HOPE_DRIVE", "ETOWN", "NYES", "COMPLEX"]; types_of_interest = ["AM - Continuity ", "PM - Continuity ", "AM - ACUTES", "PM - ACUTES "]; df["date"] = pd.to_datetime(df["date"], format="%m/%d/%Y")
	
	start_date = pd.to_datetime(st.session_state.start_date); df["week_num"] = ((df["date"] - start_date).dt.days // 7) + 1; df["week_label"] = "Week " + df["week_num"].astype(str)
	
	filtered_df = df[(df["clinic"].isin(clinics_of_interest)) & (df["type"].isin(types_of_interest))]
	
	# Count shifts per provider per clinic per week
	shift_counts = (filtered_df.groupby(["week_label", "provider", "clinic"]).size().reset_index(name="shift_count"));sorted_shift_counts = shift_counts.sort_values(by=["week_label", "shift_count"], ascending=[True, False])
	
	#st.dataframe(sorted_shift_counts)
	
	df['text'] = df['provider'].fillna("").astype(str) + " ~ " + df['student'].fillna("").astype(str)
	
	df.to_excel('final.xlsx',index=False)
	
	# Select relevant columns
	table_df = df[['student', 'clinic', 'date']]
	
	# Convert 'date' column to datetime format
	table_df["date"] = pd.to_datetime(table_df["date"])
	
	# Calculate week number and create a week label
	table_df["week_num"] = ((table_df["date"] - start_date).dt.days // 7) + 1
	table_df["week_label"] = "Week " + table_df["week_num"].astype(str)
	
	# Sort students alphabetically within each week before grouping
	table_df = table_df.sort_values(by=["week_label", "student"])
	
	# Group by week_label, combining student names and assigned clinics
	grouped_df = table_df.groupby(["week_label", "student"])["clinic"].apply(lambda x: ", ".join(x.dropna().unique())).reset_index()
	
	# Pivot the table to show weeks as columns while ensuring unique column names
	pivot_df = grouped_df.pivot(index="student", columns="week_label", values="clinic")
	
	# Reset index to ensure student names appear as a column
	pivot_df = pivot_df.reset_index()
	
	# Rename columns to remove multi-level indexing issues
	pivot_df.columns.name = None  # Remove multi-level index name
	pivot_df = pivot_df.rename_axis(None, axis=1)  # Ensure a clean dataframe

	
	st.dataframe(pivot_df)

	########################################################################################################################################################################
	import openpyxl
	from openpyxl.styles import Alignment
	
	def generate_mapping(start_value):
	    """
	    Generates a mapping dictionary for H0 to H19 starting at a given start_value.
	    """
	    return {f"H{i}": start_value + i for i in range(20)}
	
	def create_t_mapping():
	    """
	    Creates the combined mapping for T0 to T27.
	    """
	    t_mappings = [
	        (0, 6),  # T0 to T6 starts at 6
	        (7, 30),  # T7 to T13 starts at 30
	        (14, 54),  # T14 to T20 starts at 54
	        (21, 78)   # T21 to T27 starts at 78
	    ]
	
	    combined_mapping = {}
	    for start_t, start_value in t_mappings:
	        common_mapping = generate_mapping(start_value)
	        combined_mapping.update({f"T{i}": common_mapping for i in range(start_t, start_t + 7)})
	
	    return combined_mapping
	
	def process_excel_mapping(location, sheet_name):
	    """
	    Processes an Excel sheet for a given location and writes data to the corresponding OPD sheet.
	    """
	    wb = openpyxl.load_workbook('final.xlsx')
	    ws = wb['Sheet1']
	    
	    wb1 = openpyxl.load_workbook('OPD.xlsx')
	    ws1 = wb1[sheet_name]
	
	    combined_t_mapping = create_t_mapping()
	
	    column_mapping = {f"T{i}": (i % 7) + 2 for i in range(28)}
	
	    for row in ws.iter_rows():
	        t_value = row[7].value  # Column H (index 7)
	        h_value = row[6].value  # Column G (index 6)
	        row_location = row[4].value  # Column E (index 4)
	
	        if row_location == location and t_value in combined_t_mapping and h_value in combined_t_mapping[t_value]:
	            target_row = combined_t_mapping[t_value][h_value]
	            target_column = column_mapping[t_value]
	
	            ws1.cell(row=target_row, column=target_column).value = row[5].value  # Column F (index 5)
	            ws1.cell(row=target_row, column=target_column).alignment = Alignment(horizontal='center')
	
	    wb1.save('OPD.xlsx')
	    print(f"Processed mapping for {location} in {sheet_name}.")

	# Process HOPE_DRIVE
	process_excel_mapping("HOPE_DRIVE", "HOPE_DRIVE")
	process_excel_mapping("ETOWN", "ETOWN")
	process_excel_mapping("NYES", "NYES")
	process_excel_mapping("COMPLEX", "COMPLEX")
	process_excel_mapping("WARD_A", "W_A")
	process_excel_mapping("WARD_P", "W_P")
	process_excel_mapping("PICU", "PICU")
	process_excel_mapping("PSHCH_NURSERY","PSHCH_NURSERY")
	process_excel_mapping("HAMPDEN_NURSERY","HAMPDEN_NURSERY")
	process_excel_mapping("SJR_HOSP","SJR_HOSP")
	process_excel_mapping("AAC","AAC")
	process_excel_mapping("NF","NF")
	process_excel_mapping("ER_CONS","ER_CONS")
	process_excel_mapping("ADOLMED","ADOLMED")
	process_excel_mapping("WARD_C","W_C")

	###############################################################################################

	# Button to trigger the download
	if st.button('Create OPD'):
	    # Path to the existing 'OPD.xlsx' workbook
	    file_path = 'OPD.xlsx'  # Replace with your file path if it's stored somewhere else
	
	    # Read the workbook into memory
	    with open(file_path, 'rb') as file:
	        file_data = file.read()
	
	    # Provide a download button for the existing OPD.xlsx file
	    st.download_button(
	        label="Download OPD.xlsx",
	        data=file_data,
	        file_name="OPD.xlsx",
	        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
	    )

elif st.session_state.page == "Create Student Schedule":
    st.title("Create Student Schedule")
    # Upload the OPD.xlsx file
    uploaded_opd_file = st.file_uploader("Upload OPD.xlsx file", type="xlsx")
    uploaded_book4_file = st.file_uploader("Upload Book4.xlsx file", type="xlsx")
    
    if uploaded_opd_file:
        try:
            # Read the uploaded OPD file into a pandas dataframe
            df_opd = pd.read_excel(uploaded_opd_file)
            st.write("File successfully uploaded and loaded.")
            
            # Store the uploaded file in session state for use later
            st.session_state.uploaded_files['OPD.xlsx'] = uploaded_opd_file
                
        except Exception as e:
            st.error(f"Error reading the uploaded file: {e}")
    else:
        st.write("Please upload the OPD.xlsx file to proceed.")

    if uploaded_book4_file:
        try:
            # Read the uploaded OPD file into a pandas dataframe
            df_opd = pd.read_excel(uploaded_book4_file)
            st.write("File successfully uploaded and loaded.")
            
            # Store the uploaded file in session state for use later
            st.session_state.uploaded_book4_file['Book4.xlsx'] = uploaded_book4_file
                
        except Exception as e:
            st.error(f"Error reading the uploaded file: {e}")
    else:
        st.write("Please upload the Book4.xlsx file to proceed.")

    # Button to go to the next page
    if st.button("Load Student Schedule"):
        st.session_state.page = "Create List"  # Update the session state to go to the next page
        st.rerun()  # Use st.rerun() instead of st.experimental_rerun() to force rerun and update the page

elif st.session_state.page == "Create List":
##############################################################################################
    def process_week(df, start_row, end_row, date_row, clinic_name, filename):
        clinictype = df.iloc[start_row:end_row, [0]]
        days, providers = df.iloc[date_row, 1:8].values, [df.iloc[start_row:end_row, i] for i in range(1, 8)]

        week = pd.concat([
            clinictype.assign(
                type=clinictype.iloc[:, 0].str.replace(r'- Continuity', '', regex=True),
                date=days[i],
                provider=providers[i],
                clinic=clinic_name
            ) for i in range(7)
        ])

        week.to_csv(filename, index=False)
        #st.dataframe(week)
        return week
##############################################################################################
    def process_hope_data(df, type_filter, start_count, filename):
        subset = df[df['type'] == type_filter].copy()  # Ensure we’re working with a copy
        subset['count'] = subset.groupby(['date'])['provider'].cumcount() + start_count
        subset['class'] = "H" + subset['count'].astype(str)
        subset = subset.loc[:, ('date', 'type', 'provider', 'clinic', 'class')]
        subset.to_csv(filename, index=False)
        return subset

    def process_clinic_schedule(sheet_name, file_prefix, uploaded_file):
        """Processes a given clinic schedule sheet and outputs relevant CSV files."""

        # Read Excel sheet and save as CSV
        df = pd.read_excel(uploaded_file, sheet_name=sheet_name)
        df.to_csv(f'{file_prefix}_road.csv', index=False, header=False)
        df = pd.read_csv(f'{file_prefix}_road.csv')

        def process_week(df, start_row, end_row, date_row, clinic_name, filename):
            """Processes one week of clinic schedules and saves to CSV."""
            clinictype = df.iloc[start_row:end_row, [0]]
            days, providers = df.iloc[date_row, 1:8].values, [df.iloc[start_row:end_row, i] for i in range(1, 8)]

            week = pd.concat([
                clinictype.assign(
                    type=clinictype.iloc[:, 0].str.replace(r'- Continuity', '', regex=True),
                    date=days[i],
                    provider=providers[i],
                    clinic=clinic_name
                ) for i in range(7)
            ])
            week.to_csv(filename, index=False)
            return week

        # Process 4 weeks of clinic data
        week1 = process_week(df, 3, 23, 1, sheet_name, f"{file_prefix}_week1.csv")
        week2 = process_week(df, 27, 47, 25, sheet_name, f"{file_prefix}_week2.csv")
        week3 = process_week(df, 51, 71, 49, sheet_name, f"{file_prefix}_week3.csv")
        week4 = process_week(df, 75, 95, 73, sheet_name, f"{file_prefix}_week4.csv")

        # Combine weeks into a single DataFrame
        hope = pd.concat([week1, week2, week3, week4])
        hope.to_csv(f'{file_prefix}.csv', index=False)

        # Handle AM and PM classifications
        def process_classification(df, type_filter, start_count, filename):
            """Processes a classification (AM, PM) and saves it to CSV."""
            subset = df[df['type'].str.strip() == type_filter].copy()  # Handle extra spaces
            subset['count'] = subset.groupby(['date'])['provider'].cumcount() + start_count
            subset['class'] = "H" + subset['count'].astype(str)
            subset = subset.loc[:, ('date', 'type', 'provider', 'clinic', 'class')]
            subset.to_csv(filename, index=False)
            return subset

        # Process AM and PM classifications
        hope['H'] = "H"
        NYEi = process_classification(hope, 'AM', 0, f"{file_prefix}_1.csv")
        NYEii = process_classification(hope, 'PM', 10, f"{file_prefix}_2.csv")

        # Combine AM and PM DataFrames
        nyess = pd.concat([NYEi, NYEii])
        nyess.to_csv(f'{file_prefix}_summary.csv', index=False)

        # Display in Streamlit
        #st.dataframe(nyess); st.write(f"Processed file: {file_prefix}_summary.csv")
        return nyess
	    
    st.title("Load Student Schedule")

    # Ensure the OPD.xlsx file exists in the session state before proceeding
    if 'OPD.xlsx' in st.session_state.uploaded_files:
        uploaded_opd_file = st.session_state.uploaded_files['OPD.xlsx']
        
        try:
            # Read the OPD file into a dataframe
            df_opd = pd.read_excel(uploaded_opd_file)
            
            # Display the first few rows of the OPD data for verification
            #st.dataframe(df_opd.head())
            
            # Save the OPD file again without the index column
            df_opd.to_excel('OPD.xlsx', index=False)
            st.write("OPD.xlsx file has been successfully saved.")
        
        except Exception as e:
            st.error(f"Error processing the OPD file: {e}")
    else:
        st.error("No OPD file found in session state.")


    if 'Book4.xlsx' in st.session_state.uploaded_book4_file:
        uploaded_book4_file = st.session_state.uploaded_book4_file['Book4.xlsx']
        
        try:
            # Read the OPD file into a dataframe
            book4 = pd.read_excel(uploaded_book4_file)
            
            # Display the first few rows of the OPD data for verification
            #st.dataframe(book4.head())
            
            # Save the OPD file again without the index column
            book4.to_excel('Book4.xlsx', index=False)
            st.write("Book4.xlsx file has been successfully saved.")
        
        except Exception as e:
            st.error(f"Error processing the Book4 file: {e}")
    else:
        st.error("No Book4 file found in session state.")
	    
    
    # Ensure the "HOPE_DRIVE" sheet exists in the uploaded Excel file
    try:
        df = pd.read_excel(uploaded_opd_file)

        # Extract the value from row 2, column 1
        test_date = df.iloc[2, 1]

        # Ensure that test_date is a valid datetime object
        # If it's a string, convert it into a datetime object using pd.to_datetime
        if isinstance(test_date, str):
            test_date = pd.to_datetime(test_date, errors='coerce')  # Handle invalid date gracefully

        # Check if the date is valid (not NaT)
        if pd.isna(test_date):
            print("Invalid date format in the cell.")
        else:
            # Format the date to mm-dd-yyyy
            formatted_date = test_date.strftime('%m-%d-%Y')

            # Calculate the start date (use test_date directly as it's already a datetime object)
            start_date = test_date

            # Calculate the end date (34 days from the start date)
            end_date = start_date + timedelta(days=34)

            # Generate the date range
            date_range = pd.date_range(start=start_date, end=end_date)

            # Create a DataFrame with the formatted dates
            xf201 = pd.DataFrame({'date': date_range})

            # Convert the date to the desired format (e.g., "November 25, 2024")
            xf201['convert'] = xf201['date'].dt.strftime('%B %-d, %Y')  # This works in Unix-like systems

            # Add additional columns
            xf201['t'] = "T"
            xf201['c'] = xf201.index + 0  # Simple index-based column
            xf201['T'] = xf201['t'].astype(str) + xf201['c'].astype(str)

        # Creating the dateMAP DataFrame
        dateMAP = xf201[['date', 'T']].copy()  # Use .copy() to avoid the SettingWithCopyWarning

        # Convert 'date' column to datetime and then format it
        dateMAP['date'] = pd.to_datetime(dateMAP['date'])
        dateMAP['date'] = dateMAP['date'].dt.strftime('%m/%d/%Y')

        dateMAP.to_csv('xxxDATEMAP.csv', index=False)

	####################################HOPE_DRIVE#############################################################################
        read_file = pd.read_excel(uploaded_opd_file, sheet_name='HOPE_DRIVE')
        read_file.to_csv ('hopedrive.csv', index = False, header=False)
        df=pd.read_csv('hopedrive.csv')
        
        week1 = process_week(df, 3, 23, 1, "HOPE_DRIVE", "week1.csv")
        week2 = process_week(df, 27, 47, 25, "HOPE_DRIVE", "week2.csv")
        week3 = process_week(df, 51, 71, 49, "HOPE_DRIVE", "week3.csv")
        week4 = process_week(df, 75, 95, 73, "HOPE_DRIVE", "week4.csv")

        hope=pd.DataFrame(columns=week1.columns)
        hope=pd.concat([hope,week1,week2,week3,week4])

        hope.to_csv('hope.csv',index=False)
	    
        hope['H'] = "H"
        hopei = process_hope_data(hope, 'AM ', 2, '5.csv')       # AM Continuity starts at H2
        hopeii = process_hope_data(hope, 'PM ', 12, '6.csv')     # PM Continuity starts at H12
        hopeiii = process_hope_data(hope, 'AM - ACUTES', 0, '7.csv')  # AM - ACUTES starts at H0
        hopeiiii = process_hope_data(hope, 'PM - ACUTES', 10, '8.csv') # PM - ACUTES starts at H10

        # Combine all the data into one DataFrame
        hopes = pd.DataFrame(columns=hopei.columns)
        hopes = pd.concat([hopei, hopeii, hopeiii, hopeiiii])

        # Save the combined DataFrame to CSV
        hopes.to_csv('hopes.csv', index=False); #st.dataframe(hopes)
        #################################################################################################################
        # List of sheet names to process
        sheet_names = ['ETOWN', 'NYES', 'COMPLEX', 'W_A', 'W_C', 'W_P', 'PICU', 'PSHCH_NURSERY', 'HAMPDEN_NURSERY', 'SJR_HOSP', 'AAC', 'ER_CONS', 'NF', 'ADOLMED']  # Add more as needed

        # Process all sheets dynamically
        for sheet in sheet_names:
            process_clinic_schedule(sheet, sheet.lower(), uploaded_opd_file)	
	############################################################################
        summary_files = [f"{sheet.lower()}_summary.csv" for sheet in sheet_names]
        dfx = pd.concat([pd.read_csv(file) for file in summary_files] + [pd.read_csv('hopes.csv')], ignore_index=True)

        dfx['providers'] = dfx['provider'].str.split('~').str[0]
        dfx['student'] = dfx['provider'].str.split('~').str[1]
        dfx1 = dfx[['date', 'type', 'providers', 'student', 'clinic', 'provider', 'class']]

        dfx1['date'] = pd.to_datetime(dfx1['date'])
        dfx1['date'] = dfx1['date'].dt.strftime('%m/%d/%Y')

        mydict = {}
        with open('xxxDATEMAP.csv', mode='r')as inp:     #file is the objects you want to map. I want to map the IMP in this file to diagnosis.csv
            reader = csv.reader(inp)
            df1 = {rows[0]:rows[1] for rows in reader} 
        dfx1['datecode'] = dfx1.date.map(df1)               #'type' is the new column in the diagnosis file. 'encounter_id' is the key you are using to MAP 

        dfx1.to_excel('STUDENTLIST.xlsx',index=False)
        dfx1.to_csv('STUDENTLIST.csv',index=False)

        dfx1['type'] = dfx1['type'].str.lstrip()
        dfx1['type'] = dfx1['type'].str.rstrip()

        dfx1['providers'] = dfx1['providers'].str.lstrip()
        dfx1['providers'] = dfx1['providers'].str.rstrip()

        dfx1['student'] = dfx1['student'].str.lstrip()
        dfx1['student'] = dfx1['student'].str.rstrip()

        dfs = [["AM","S1"],["PM","S2"],["AM - ACUTES","S11"],["PM - ACUTES","S12"]]
        dftype = pd.DataFrame(dfs, columns = ['type', 'datecode2'])
        dftype.to_csv('dftype.csv',index=False)

        mydict = {}
        with open('dftype.csv', mode='r')as inp:     #file is the objects you want to map. I want to map the IMP in this file to diagnosis.csv
            reader = csv.reader(inp)
            df1 = {rows[0]:rows[1] for rows in reader} 
        dfx1['datecode2'] = dfx1.type.map(df1)               #'type' is the new column in the diagnosis file. 'encounter_id' is the key you are using to MAP 

        dfx1.to_excel('PALIST.xlsx',index=False)

        dfx1.to_csv('PALIST.csv',index=False)

        dfx1 = pd.read_csv('PALIST.csv')

        new_row = pd.DataFrame({'date':0, 'type':0, 'providers':0,'student':0, 'clinic':0, 'provider':0,'class':0, 'datecode':0, 'datecode2':0}, index =[0])
        # simply concatenate both dataframes
        df = pd.concat([new_row, dfx1]).reset_index(drop = True)

        df['clinic'] = df['clinic'].replace({"ETOWN": "ETOWN", "NYES": "NYES", "COMPLEX": "COMPLEX", "W_A": "WARD A", "W_C": "WARD C", "W_P": "WARD P", "PICU": "PICU", "PSHCH_NURSERY": "PSHCH NURSERY", "HAMPDEN_NURSERY": "HAMPDEN NURSERY", "SJR_HOSP": "SJR HOSP", "AAC": "AAC", "ER_CONS": "ER CONSULTS", "NF": "NIGHT FLOAT", "ADOLMED": "ADOLMED", "HOPE DRIVE": "HOPE_DRIVE"}); df.to_csv('PALIST.csv',index=False)

        df = pd.read_excel('Book4.xlsx')

        # Keep the first NaN in the first row as it is for the first column
        df.iloc[0, 0] = np.nan  # Ensure the first cell is NaN (or leave it as it is)

        # Initialize a counter for BLANK replacements
        blank_counter = 1

        # Replace NaN values in the first column with BLANK1, BLANK2, etc. starting from the second row
        for i in range(1, len(df)):
            if pd.isna(df.iloc[i, 0]):
                df.iloc[i, 0] = f'BLANK{blank_counter}'
                blank_counter += 1

        x = (df.loc[0, 'Week 1'])

        x = x.strftime("%m/%d/%Y")

        test_date = datetime.datetime.strptime(x, "%m/%d/%Y")

        # initializing K
        K = 28

        res = []

        for day in range(K):
            date = (test_date + datetime.timedelta(days = day)).strftime("%-m/%-d/%Y")
            res.append(date)

        #res

        dates = pd.DataFrame(res, columns =['dates'])

        dates['x'] = "y"

        dates['i'] = dates.index+1

        dates['i2']= dates.index+0

        dates['T'] = "T" + dates['i2'].astype(str)

        dates['x'] = dates['x'].astype(str)+dates['i'].astype(str)

        dates['x'] = dates['x'].astype(str) + "=" + "'"+dates['dates'].astype(str) + "'"

        dateT = dates[['dates','T']]

        dates = dates[['x']]

        dates.to_csv('dates.csv',index=False)

        dateT.to_csv('datesT.csv',index=False)

        import numpy as np

        datesdf = pd.read_csv('dates.csv')

        dates = datesdf['x'].astype(str)

        numpy_array=dates.to_numpy()
        np.savetxt("dates.py",numpy_array, fmt="%s")

        exec(open('dates.py').read())

        #################################STUDENT NAME and CLINICAL EXPERIENCES INPUT#################################
        column = "Student Name:"
        L=df[(column)].unique().tolist() #Must Use Unique
        xf200 = pd.DataFrame({'col':L})
        xf200['i'] = xf200.index

        column = "Week 1"
        L=df[(column)].tolist() #No Unique Required 
        xf201 = pd.DataFrame({'col':L})
        xf201['i'] = xf201.index

        column = "Week 2"
        L=df[(column)].tolist()
        xf202 = pd.DataFrame({'col':L})
        xf202['i'] = xf202.index

        column = "Week 3"
        L=df[(column)].tolist()
        xf203 = pd.DataFrame({'col':L})
        xf203['i'] = xf203.index

        column = "Week 4"
        L=df[(column)].tolist()
        xf204 = pd.DataFrame({'col':L})
        xf204['i'] = xf204.index

        ######################################################################################################################
        # Create a workbook and add the main worksheet
        # Create a workbook and add the main worksheet
        workbook = xlsxwriter.Workbook('Main_Schedule_MS.xlsx')

        # Define the formats
        format1 = workbook.add_format({'font_size': 14,'bold': 1,'align': 'center','valign': 'vcenter','font_color': 'black','text_wrap': True,'bg_color': '#FEFFCC','border': 1})
        format2 = workbook.add_format({'font_size': 10,'bold': 1,'align': 'center','valign': 'vcenter','font_color': 'yellow','bg_color': 'black','border': 1,'text_wrap': True})
        format3 = workbook.add_format({'font_size':12,'bold': 1,'align': 'center','valign': 'vcenter','font_color':'black','bg_color':'#FFC7CE','border':1})
        format4 = workbook.add_format({'num_format':'mm/dd/yyyy','font_size':12,'bold': 1,'align': 'center','valign': 'vcenter','font_color':'black','bg_color':'#F4F6F7','border':1})
        format5 = workbook.add_format({'font_size':12,'bold': 1,'align': 'center','valign': 'vcenter','font_color':'black','bg_color':'#F4F6F7','border':1})
        format6 = workbook.add_format({'bg_color':'black','border':1})
        format7 = workbook.add_format({'font_size':12,'bold': 1,'align': 'center','valign': 'vcenter','font_color':'black','bg_color':'#90EE90','border':1})
        format8 = workbook.add_format({'font_size':12,'bold': 1,'align': 'center','valign':'vcenter','font_color':'black','bg_color':'#89CFF0','border':1})
	    
        # Initialize a list to store worksheets
        worksheets = []

        # Add a worksheet for each unique name in the DataFrame
        for name in xf200['col']:
            # Ensure that the name is a valid string (i.e., no NaN or floats)
            if isinstance(name, str):  # Check if it's a string
                name = name[:31]  # Truncate the name if it's longer than 31 characters
            else:
                continue  # Skip invalid entries (like NaN or numbers)

            # Ensure the name is not an empty string
            if len(name) > 0:
                worksheet = workbook.add_worksheet(name)
                worksheets.append(worksheet)  # Store the worksheet in the list

                # Merge range for the student name column
                worksheet.merge_range('A1:A2', 'Student Name:', format1)
                worksheet.merge_range('B1:B2', name, format1)

                # Merge range for the note
                note = '*Note** Asynchronous time is for coursework only. During this time period, we expect students to do coursework, be available for any additional educational activities, and any extra clinical time that may be available. If the student is not available during this time period and has not made an absence request, the student will be cited for unprofessionalism and will risk failing the course.'
                worksheet.merge_range('C1:H2', note, format2)

                # Set column widths (you can set them all at once or in a loop)
                worksheet.set_column('A:A', 20)
                worksheet.set_column('B:B', 30)
                worksheet.set_column('C:G', 40)
                worksheet.set_column('H:H', 155)
		    
                # Set row height for header row
                worksheet.set_row(0, 37.25)

                days_of_week = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']

                # Start row positions for each week
                start_rows = [3, 11, 19, 27]

                # Loop through each week's starting row and write the days
                for start_row in start_rows:
                    for col, day in enumerate(days_of_week, start=1):  
                        worksheet.write(chr(65 + col) + str(start_row), day, format3)

                # List of weeks
                weeks = ['Week 1', 'Week 2', 'Week 3', 'Week 4']

                for week_index, week in enumerate(weeks):
                    worksheet.write(f'A{4 + (week_index * 8)}', week, format3)

                for week_index in range(4):  # For each week (4 weeks)
                    row = 6 + (week_index * 8)  # Calculate row for AM (6, 14, 22, 30)
                    worksheet.write(f'A{row}', 'AM', format3)

                for week_index in range(4):  # For each week (4 weeks)
                    row = 7 + (week_index * 8)  # Calculate row for PM (7, 15, 23, 31)
                    worksheet.write(f'A{row}', 'PM', format3)

                column_idx = 1  # Starting at column B (Excel columns are 1-indexed)
                date_idx = 0   # Start from the first date in `res`

                # Loop through the weeks and the corresponding dates
                for week_row in start_rows:
                    for col in range(7):  # Loop through Monday to Sunday (7 days)
                        if date_idx < len(res):
                            date = res[date_idx]
                            worksheet.write(week_row, column_idx + col, date, format4)
                            date_idx += 1  # Move to the next date

                worksheet.set_zoom(70)
                worksheet.merge_range('A10:H10','',format6)
                worksheet.merge_range('A18:H18','',format6)
                worksheet.merge_range('A26:H26','',format6)
                worksheet.merge_range('A34:H34','',format6)

                worksheet.write('A9',' ', format7)
                worksheet.write('B9',' ', format7)
                worksheet.write('C9',' ', format7)
                worksheet.write('D9',' ', format7)
                worksheet.write('E9',' ', format7)
                worksheet.write('F9',' ', format7)
                worksheet.write('G9',' ', format7)
                worksheet.write('H9',' ', format7)

                worksheet.write('A17',' ', format7)
                worksheet.write('B17',' ', format7)
                worksheet.write('C17',' ', format7)
                worksheet.write('D17',' ', format7)
                worksheet.write('E17',' ', format7)
                worksheet.write('F17',' ', format7)
                worksheet.write('G17',' ', format7)
                worksheet.write('H17',' ', format7)

                worksheet.write('A25',' ', format7)
                worksheet.write('B25',' ', format7)
                worksheet.write('C25',' ', format7)
                worksheet.write('D25',' ', format7)
                worksheet.write('E25',' ', format7)
                worksheet.write('F25',' ', format7)
                worksheet.write('G25',' ', format7)
                worksheet.write('H25',' ', format7)

                worksheet.write('A33',' ', format7)
                worksheet.write('B33',' ', format7)
                worksheet.write('C33',' ', format7)
                worksheet.write('D33',' ', format7)
                worksheet.write('E33',' ', format7)
                worksheet.write('F33',' ', format7)
                worksheet.write('G33',' ', format7)
                worksheet.write('H33',' ', format7)

                # Writing to row 8
                worksheet.write('A8', 'ASSIGNMENT DUE:', format8)
                worksheet.write('B8', ' ', format8)
                worksheet.write('C8', ' ', format8)
                worksheet.write('D8', ' ', format8)
                worksheet.write('E8', ' ', format8)
                worksheet.write('F8', 'Ask for Feedback!', format8)
                worksheet.write('G8', ' ', format8)
                worksheet.write('H8', 'Quiz 1 Due', format8)

                # Writing to row 16
                worksheet.write('A16', 'ASSIGNMENT DUE:', format8)
                worksheet.write('B16', ' ', format8)
                worksheet.write('C16', ' ', format8)
                worksheet.write('D16', ' ', format8)
                worksheet.write('E16', ' ', format8)
                worksheet.write('F16', 'Ask for Feedback!', format8)
                worksheet.write('G16', ' ', format8)
                worksheet.write('H16', 'Quiz 2, Pediatric Documentation #1, 1 Clinical Encounter Log Due', format8)

                # Writing to row 24
                worksheet.write('A24', 'ASSIGNMENT DUE:', format8)
                worksheet.write('B24', ' ', format8)
                worksheet.write('C24', ' ', format8)
                worksheet.write('D24', ' ', format8)
                worksheet.write('E24', ' ', format8)
                worksheet.write('F24', 'Ask for Feedback!', format8)
                worksheet.write('G24', ' ', format8)
                worksheet.write('H24', 'Quiz 3 Due', format8)

                # Writing to row 32
                worksheet.write('A32', 'ASSIGNMENT DUE:', format8)
                worksheet.write('B32', ' ', format8)
                worksheet.write('C32', ' ', format8)
                worksheet.write('D32', ' ', format8)
                worksheet.write('E32', ' ', format8)
                worksheet.write('F32', ' ', format8)
                worksheet.write('G32', ' ', format8)
                worksheet.write('H32', 'Quiz 4, Pediatric Documentation #2, Social Drivers of Health Assessment Form, Developmental Assessment of Pediatric Patient Form, All Clinical Encounter Logs are Due!', format8)

        # Now, we'll write the location data
        locations = xf201['col'].tolist()[1:]  # Assuming 'col' is a column in xf201 dataframe

        # Ensure locations match the number of worksheets
        if len(locations) != len(worksheets):
            raise ValueError(f"Number of locations ({len(locations)}) does not match number of worksheets ({len(worksheets)})")

        # Iterate over the names and locations and write the data
        for i, (name, location) in enumerate(zip(xf200['col'], locations)):
            worksheet = worksheets[i]  # Get the corresponding worksheet for this index

            # Write the location data to rows 6 and 7 for each worksheet
            for row in range(6, 8):  # Rows 6 and 7
                worksheet.write(f'B{row}', location, format5)
                worksheet.write(f'C{row}', location, format5)
                worksheet.write(f'D{row}', location, format5)
                worksheet.write(f'E{row}', location, format5)
                worksheet.write(f'F{row}', location, format5)
                worksheet.write(f'G{row}', "OFF", format5)
                worksheet.write(f'H{row}', "OFF", format5)

        # Now, we'll write the location data
        locations = xf202['col'].tolist()[1:]  # Assuming 'col' is a column in xf201 dataframe

        # Ensure locations match the number of worksheets
        if len(locations) != len(worksheets):
            raise ValueError(f"Number of locations ({len(locations)}) does not match number of worksheets ({len(worksheets)})")

        # Iterate over the names and locations and write the data
        for i, (name, location) in enumerate(zip(xf200['col'], locations)):
            worksheet = worksheets[i]  # Get the corresponding worksheet for this index

            for row in range(14, 16): 
                worksheet.write(f'B{row}', location, format5)
                worksheet.write(f'C{row}', location, format5)
                worksheet.write(f'D{row}', location, format5)
                worksheet.write(f'E{row}', location, format5)
                worksheet.write(f'F{row}', location, format5)
                worksheet.write(f'G{row}', "OFF", format5)
                worksheet.write(f'H{row}', "OFF", format5)

        # Now, we'll write the location data
        locations = xf203['col'].tolist()[1:]  # Assuming 'col' is a column in xf201 dataframe

        # Ensure locations match the number of worksheets
        if len(locations) != len(worksheets):
            raise ValueError(f"Number of locations ({len(locations)}) does not match number of worksheets ({len(worksheets)})")

        # Iterate over the names and locations and write the data
        for i, (name, location) in enumerate(zip(xf200['col'], locations)):
            worksheet = worksheets[i]  # Get the corresponding worksheet for this index

            for row in range(22, 24): 
                worksheet.write(f'B{row}', location, format5)
                worksheet.write(f'C{row}', location, format5)
                worksheet.write(f'D{row}', location, format5)
                worksheet.write(f'E{row}', location, format5)
                worksheet.write(f'F{row}', location, format5)
                worksheet.write(f'G{row}', "OFF", format5)
                worksheet.write(f'H{row}', "OFF", format5)

        # Now, we'll write the location data
        locations = xf204['col'].tolist()[1:]  # Assuming 'col' is a column in xf201 dataframe

        # Ensure locations match the number of worksheets
        if len(locations) != len(worksheets):
            raise ValueError(f"Number of locations ({len(locations)}) does not match number of worksheets ({len(worksheets)})")

        # Iterate over the names and locations and write the data
        for i, (name, location) in enumerate(zip(xf200['col'], locations)):
            worksheet = worksheets[i]  # Get the corresponding worksheet for this index

            for row in range(30, 32): 
                worksheet.write(f'B{row}', location, format5)
                worksheet.write(f'C{row}', location, format5)
                worksheet.write(f'D{row}', location, format5)
                worksheet.write(f'E{row}', location, format5)
                worksheet.write(f'F{row}', location, format5)
                worksheet.write(f'G{row}', "OFF", format5)
                worksheet.write(f'H{row}', "OFF", format5)


        # Close the workbook
        workbook.close()
        
        df = pd.read_csv('datesT.csv')
        dx = df

        df["dates"] = pd.to_datetime(df["dates"])
        df['dates'] = df['dates'].dt.strftime('%m/%d/%Y')

        df.to_csv('datesT.csv',index=False)

        df=pd.read_csv('PALIST.csv',dtype=str)
	    
        df['text'] = df['providers'] + " - " + "[" + df['clinic'] + "]"
        df = df[['datecode','type','student','text','date','clinic']]

        mydict = {}
        with open('datesT.csv', mode='r')as inp:     #file is the objects you want to map. I want to map the IMP in this file to diagnosis.csv
            reader = csv.reader(inp)
            df1 = {rows[0]:rows[1] for rows in reader} 
        df['datecode'] = df.date.map(df1)

        df = df[~df['student'].isnull()]

        df = df.loc[df['student'] != "0"]

        df.to_excel('Source1.xlsx', index=False); st.dataframe(df)

        import openpyxl
        import numpy as np 

        column = "student"
        L=df[(column)].tolist() #Must Use Unique
        xf200 = pd.DataFrame({'col':L})
        xf200['i'] = xf200.index
        xf200['blank'] = ''
        
        wb = openpyxl.load_workbook('Source1.xlsx')
        ws = wb['Sheet1']
        wb1 = openpyxl.load_workbook('Main_Schedule_MS.xlsx')

        # Assuming xf200 is a pandas DataFrame loaded from somewhere, for example:
        # xf200 = pd.read_csv('your_dataframe.csv')  # Or load your DataFrame as needed

        # Iterate over the names in the xf200 DataFrame
        for name in xf200['col']:  # Assuming 'names' is the column with the names you want to look up
            # Check if the sheet with the name exists in the workbook
            if name in wb1.sheetnames:
                ws1 = wb1[name]  # Access the sheet directly since it exists
            else:
                print(f"Sheet for {name} not found, skipping.")
                continue  # Skip to the next name if the sheet does not exist

            for row in ws.iter_rows():
                if row[2].value == name:  # Check if the name matches (assuming name is in column 3)
                    # Handle T0 to T6
                    if row[0].value == "T0":
                        if row[1].value == "AM":
                            ws1.cell(row=6, column=2).value = row[3].value  # AM, T0
                        elif row[1].value == "PM":
                            ws1.cell(row=7, column=2).value = row[3].value  # PM, T0
                    elif row[0].value == "T1":
                        if row[1].value == "AM":
                            ws1.cell(row=6, column=3).value = row[3].value  # AM, T1
                        elif row[1].value == "PM":
                            ws1.cell(row=7, column=3).value = row[3].value  # PM, T1
                    elif row[0].value == "T2":
                        if row[1].value == "AM":
                            ws1.cell(row=6, column=4).value = row[3].value  # AM, T2
                        elif row[1].value == "PM":
                            ws1.cell(row=7, column=4).value = row[3].value  # PM, T2
                    elif row[0].value == "T3":
                        if row[1].value == "AM":
                            ws1.cell(row=6, column=5).value = row[3].value  # AM, T3
                        elif row[1].value == "PM":
                            ws1.cell(row=7, column=5).value = row[3].value  # PM, T3
                    elif row[0].value == "T4":
                        if row[1].value == "AM":
                            ws1.cell(row=6, column=6).value = row[3].value  # AM, T4
                        elif row[1].value == "PM":
                            ws1.cell(row=7, column=6).value = row[3].value  # PM, T4
                    elif row[0].value == "T5":
                        if row[1].value == "AM":
                            ws1.cell(row=6, column=7).value = row[3].value  # AM, T5
                        elif row[1].value == "PM":
                            ws1.cell(row=7, column=7).value = row[3].value  # PM, T5
                    elif row[0].value == "T6":
                        if row[1].value == "AM":
                            ws1.cell(row=6, column=8).value = row[3].value  # AM, T6
                        elif row[1].value == "PM":
                            ws1.cell(row=7, column=8).value = row[3].value  # PM, T6

                    # Handle T7 to T13
                    elif row[0].value == "T7":
                        if row[1].value == "AM":
                            ws1.cell(row=14, column=2).value = row[3].value  # AM, T7
                        elif row[1].value == "PM":
                            ws1.cell(row=15, column=2).value = row[3].value  # PM, T7
                    elif row[0].value == "T8":
                        if row[1].value == "AM":
                            ws1.cell(row=14, column=3).value = row[3].value  # AM, T8
                        elif row[1].value == "PM":
                            ws1.cell(row=15, column=3).value = row[3].value  # PM, T8
                    elif row[0].value == "T9":
                        if row[1].value == "AM":
                            ws1.cell(row=14, column=4).value = row[3].value  # AM, T9
                        elif row[1].value == "PM":
                            ws1.cell(row=15, column=4).value = row[3].value  # PM, T9
                    elif row[0].value == "T10":
                        if row[1].value == "AM":
                            ws1.cell(row=14, column=5).value = row[3].value  # AM, T10
                        elif row[1].value == "PM":
                            ws1.cell(row=15, column=5).value = row[3].value  # PM, T10
                    elif row[0].value == "T11":
                        if row[1].value == "AM":
                            ws1.cell(row=14, column=6).value = row[3].value  # AM, T11
                        elif row[1].value == "PM":
                            ws1.cell(row=15, column=6).value = row[3].value  # PM, T11
                    elif row[0].value == "T12":
                        if row[1].value == "AM":
                            ws1.cell(row=14, column=7).value = row[3].value  # AM, T12
                        elif row[1].value == "PM":
                            ws1.cell(row=15, column=7).value = row[3].value  # PM, T12
                    elif row[0].value == "T13":
                        if row[1].value == "AM":
                            ws1.cell(row=14, column=8).value = row[3].value  # AM, T13
                        elif row[1].value == "PM":
                            ws1.cell(row=15, column=8).value = row[3].value  # PM, T13

                    # Handle T14 to T20
                    elif row[0].value == "T14":
                        if row[1].value == "AM":
                            ws1.cell(row=22, column=2).value = row[3].value  # AM, T14
                        elif row[1].value == "PM":
                            ws1.cell(row=23, column=2).value = row[3].value  # PM, T14
                    elif row[0].value == "T15":
                        if row[1].value == "AM":
                            ws1.cell(row=22, column=3).value = row[3].value  # AM, T15
                        elif row[1].value == "PM":
                            ws1.cell(row=23, column=3).value = row[3].value  # PM, T15
                    elif row[0].value == "T16":
                        if row[1].value == "AM":
                            ws1.cell(row=22, column=4).value = row[3].value  # AM, T16
                        elif row[1].value == "PM":
                            ws1.cell(row=23, column=4).value = row[3].value  # PM, T16
                    elif row[0].value == "T17":
                        if row[1].value == "AM":
                            ws1.cell(row=22, column=5).value = row[3].value  # AM, T17
                        elif row[1].value == "PM":
                            ws1.cell(row=23, column=5).value = row[3].value  # PM, T17
                    elif row[0].value == "T18":
                        if row[1].value == "AM":
                            ws1.cell(row=22, column=6).value = row[3].value  # AM, T18
                        elif row[1].value == "PM":
                            ws1.cell(row=23, column=6).value = row[3].value  # PM, T18
                    elif row[0].value == "T19":
                        if row[1].value == "AM":
                            ws1.cell(row=22, column=7).value = row[3].value  # AM, T19
                        elif row[1].value == "PM":
                            ws1.cell(row=23, column=7).value = row[3].value  # PM, T19
                    elif row[0].value == "T20":
                        if row[1].value == "AM":
                            ws1.cell(row=22, column=8).value = row[3].value  # AM, T20
                        elif row[1].value == "PM":
                            ws1.cell(row=23, column=8).value = row[3].value  # PM, T20

                    # Handle T21 to T27
                    elif row[0].value == "T21":
                        if row[1].value == "AM":
                            ws1.cell(row=30, column=2).value = row[3].value  # AM, T21
                        elif row[1].value == "PM":
                            ws1.cell(row=31, column=2).value = row[3].value  # PM, T21
                    elif row[0].value == "T22":
                        if row[1].value == "AM":
                            ws1.cell(row=30, column=3).value = row[3].value  # AM, T22
                        elif row[1].value == "PM":
                            ws1.cell(row=31, column=3).value = row[3].value  # PM, T22
                    elif row[0].value == "T23":
                        if row[1].value == "AM":
                            ws1.cell(row=30, column=4).value = row[3].value  # AM, T23
                        elif row[1].value == "PM":
                            ws1.cell(row=31, column=4).value = row[3].value  # PM, T23
                    elif row[0].value == "T24":
                        if row[1].value == "AM":
                            ws1.cell(row=30, column=5).value = row[3].value  # AM, T24
                        elif row[1].value == "PM":
                            ws1.cell(row=31, column=5).value = row[3].value  # PM, T24
                    elif row[0].value == "T25":
                        if row[1].value == "AM":
                            ws1.cell(row=30, column=6).value = row[3].value  # AM, T25
                        elif row[1].value == "PM":
                            ws1.cell(row=31, column=6).value = row[3].value  # PM, T25
                    elif row[0].value == "T26":
                        if row[1].value == "AM":
                            ws1.cell(row=30, column=7).value = row[3].value  # AM, T26
                        elif row[1].value == "PM":
                            ws1.cell(row=31, column=7).value = row[3].value  # PM, T26
                    elif row[0].value == "T27":
                        if row[1].value == "AM":
                            ws1.cell(row=30, column=8).value = row[3].value  # AM, T27
                        elif row[1].value == "PM":
                            ws1.cell(row=31, column=8).value = row[3].value  # PM, T27

            for row in ws.iter_rows():
                if row[2].value == name:  # Check if the name matches (assuming name is in column 3)

                    # Handle T0 to T6
                    if row[0].value == "T0":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=6, column=2).value = row[3].value  # AM - ACUTES, T0
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=7, column=2).value = row[3].value  # PM - ACUTES, T0
                    elif row[0].value == "T1":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=6, column=3).value = row[3].value  # AM - ACUTES, T1
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=7, column=3).value = row[3].value  # PM - ACUTES, T1
                    elif row[0].value == "T2":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=6, column=4).value = row[3].value  # AM - ACUTES, T2
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=7, column=4).value = row[3].value  # PM - ACUTES, T2
                    elif row[0].value == "T3":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=6, column=5).value = row[3].value  # AM - ACUTES, T3
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=7, column=5).value = row[3].value  # PM - ACUTES, T3
                    elif row[0].value == "T4":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=6, column=6).value = row[3].value  # AM - ACUTES, T4
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=7, column=6).value = row[3].value  # PM - ACUTES, T4
                    elif row[0].value == "T5":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=6, column=7).value = row[3].value  # AM - ACUTES, T5
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=7, column=7).value = row[3].value  # PM - ACUTES, T5
                    elif row[0].value == "T6":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=6, column=8).value = row[3].value  # AM - ACUTES, T6
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=7, column=8).value = row[3].value  # PM - ACUTES, T6

                    # Handle T7 to T13
                    elif row[0].value == "T7":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=14, column=2).value = row[3].value  # AM - ACUTES, T7
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=15, column=2).value = row[3].value  # PM - ACUTES, T7
                    elif row[0].value == "T8":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=14, column=3).value = row[3].value  # AM - ACUTES, T8
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=15, column=3).value = row[3].value  # PM - ACUTES, T8
                    elif row[0].value == "T9":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=14, column=4).value = row[3].value  # AM - ACUTES, T9
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=15, column=4).value = row[3].value  # PM - ACUTES, T9
                    elif row[0].value == "T10":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=14, column=5).value = row[3].value  # AM - ACUTES, T10
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=15, column=5).value = row[3].value  # PM - ACUTES, T10
                    elif row[0].value == "T11":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=14, column=6).value = row[3].value  # AM - ACUTES, T11
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=15, column=6).value = row[3].value  # PM - ACUTES, T11
                    elif row[0].value == "T12":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=14, column=7).value = row[3].value  # AM - ACUTES, T12
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=15, column=7).value = row[3].value  # PM - ACUTES, T12
                    elif row[0].value == "T13":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=14, column=8).value = row[3].value  # AM - ACUTES, T13
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=15, column=8).value = row[3].value  # PM - ACUTES, T13

                    # Handle T14 to T20
                    elif row[0].value == "T14":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=22, column=2).value = row[3].value  # AM - ACUTES, T14
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=23, column=2).value = row[3].value  # PM - ACUTES, T14
                    elif row[0].value == "T15":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=22, column=3).value = row[3].value  # AM - ACUTES, T15
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=23, column=3).value = row[3].value  # PM - ACUTES, T15
                    elif row[0].value == "T16":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=22, column=4).value = row[3].value  # AM - ACUTES, T16
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=23, column=4).value = row[3].value  # PM - ACUTES, T16
                    elif row[0].value == "T17":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=22, column=5).value = row[3].value  # AM - ACUTES, T17
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=23, column=5).value = row[3].value  # PM - ACUTES, T17
                    elif row[0].value == "T18":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=22, column=6).value = row[3].value  # AM - ACUTES, T18
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=23, column=6).value = row[3].value  # PM - ACUTES, T18
                    elif row[0].value == "T19":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=22, column=7).value = row[3].value  # AM - ACUTES, T19
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=23, column=7).value = row[3].value  # PM - ACUTES, T19
                    elif row[0].value == "T20":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=22, column=8).value = row[3].value  # AM - ACUTES, T20
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=23, column=8).value = row[3].value  # PM - ACUTES, T20

                    # Handle T21 to T27
                    elif row[0].value == "T21":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=30, column=2).value = row[3].value  # AM - ACUTES, T21
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=31, column=2).value = row[3].value  # PM - ACUTES, T21
                    elif row[0].value == "T22":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=30, column=3).value = row[3].value  # AM - ACUTES, T22
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=31, column=3).value = row[3].value  # PM - ACUTES, T22
                    elif row[0].value == "T23":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=30, column=4).value = row[3].value  # AM - ACUTES, T23
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=31, column=4).value = row[3].value  # PM - ACUTES, T23
                    elif row[0].value == "T24":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=30, column=5).value = row[3].value  # AM - ACUTES, T24
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=31, column=5).value = row[3].value  # PM - ACUTES, T24
                    elif row[0].value == "T25":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=30, column=6).value = row[3].value  # AM - ACUTES, T25
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=31, column=6).value = row[3].value  # PM - ACUTES, T25
                    elif row[0].value == "T26":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=30, column=7).value = row[3].value  # AM - ACUTES, T26
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=31, column=7).value = row[3].value  # PM - ACUTES, T26
                    elif row[0].value == "T27":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=30, column=8).value = row[3].value  # AM - ACUTES, T27
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=31, column=8).value = row[3].value  # PM - ACUTES, T27

        # Save the modified workbook
        wb1.save('Main_Schedule_MS.xlsx')


            # Function to save the workbook to a BytesIO object
        def save_to_bytes_wb(wb):
                output = BytesIO()
                wb.save(output)
                output.seek(0)  # Rewind the file pointer to the start
                return output
		
        def save_to_bytes_csv(df):
                output = StringIO()
                df.to_csv(output, index=False) 
                output.seek(0)  # Rewind the file pointer to the start
                return output.getvalue()  
	
        # Prepare the workbook for download
        wb_bytes = save_to_bytes_wb(wb1)
        st.download_button(label="Download Medical Student Schedule",data=wb_bytes,file_name="Main_Schedule_MS.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        df = pd.read_csv('PALIST.csv', dtype=str)
        mapping_df = st.secrets["dataset"]["data"]
        st.dataframe(mapping_df)
        mapping_df = pd.DataFrame(mapping_df)
        mapping_df.to_csv('mapping_df.csv',index=False)

        records_df = st.secrets["dataset_record"]["data_record"]
        records_df = pd.DataFrame(records_df)
        records_df.to_csv('record_df.csv',index=False)
	    
        # Normalize 'type' for HOPE_DRIVE clinic
        df['type_adj'] = df['type']
        df.loc[(df['clinic'] == 'HOPE_DRIVE') & (df['type'].str.startswith('AM')), 'type_adj'] = 'AM'
        df.loc[(df['clinic'] == 'HOPE_DRIVE') & (df['type'].str.startswith('PM')), 'type_adj'] = 'PM'
    
        # Ensure 'student' column is clean (replace NaN with empty string)
        df['student'] = df['student'].fillna("").str.strip()
    
        # Identify duplicate student assignments (excluding HOPE_DRIVE)
        df['duplicate_flag'] = df.duplicated(subset=['date', 'type', 'student'], keep=False) & (df['clinic'] != 'HOPE_DRIVE') & (df['student'] != "")
    
        # Identify duplicate assignments within HOPE_DRIVE using adjusted type
        df['duplicate_flag'] |= df.duplicated(subset=['date', 'type_adj', 'student'], keep=False) & (df['clinic'] == 'HOPE_DRIVE') & (df['student'] != "")
    
        # Filter only flagged duplicates
        df_duplicates = df[df['duplicate_flag']]
    
        # Display only flagged duplicate records
        st.write("Duplicate Check:"); st.dataframe(df_duplicates);                

        provider_df = (df[df['student'].notna() & (df['student'].str.strip() != "")].assign(date=pd.to_datetime(df['date'], errors='coerce')).groupby(['student', 'providers'], as_index=False)['date'].max().assign(eval_due_date=lambda x: x['date'] + pd.Timedelta(days=14)))
        provider_df['providers'] = provider_df['providers'].str.split('/'); provider_df = provider_df.explode('providers').reset_index(drop=True); st.write("Evaluation Due Dates:");st.dataframe(provider_df); provider_df.to_csv('provider_df.csv',index=False)

        provider_df = pd.read_csv("provider_df.csv")
        mapping_df = pd.read_csv("mapping_df.csv")  # The file with 'name' and 'Formatted Name'

        # Normalize columns for matching
        provider_df["providers"] = provider_df["providers"].str.strip().str.lower()
        mapping_df["name"] = mapping_df["name"].str.strip().str.lower()

        # Convert mapping dataframe to dictionary
        mapping_dict = dict(zip(mapping_df["name"], mapping_df["Formatted Name"]))

        # Map names
        provider_df["formatted_name"] = provider_df["providers"].map(mapping_dict).fillna(provider_df["providers"])

        # Print unmatched values
        unmatched = provider_df[provider_df["formatted_name"].isna()]["providers"].unique()
        if unmatched.size > 0:
            st.write("Warning: Unmatched names found:", unmatched)

        # Convert mapping dataframe to dictionary
        mapping_dict = dict(zip(records_df["legal_name"], records_df["record_id"]))

        # Map names
        provider_df["record_id"] = provider_df["student"].map(mapping_dict)
        
        provider_df = provider_df[['record_id','formatted_name','date', 'eval_due_date']]  
	
        csv_bytes = save_to_bytes_csv(provider_df); st.dataframe(provider_df); st.download_button(label="Download Evaluation Due Dates",data=csv_bytes,file_name="PALIST.csv",mime="text/csv")
	    
    except Exception as e:
        st.error(f"Error processing the HOPE_DRIVE sheet: {e}")

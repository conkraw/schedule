import streamlit as st
import csv
import datetime
import pandas as pd
import numpy as np
from datetime import timedelta
import xlsxwriter
import openpyxl
from openpyxl import Workbook
import io
from io import BytesIO, StringIO
import os
import time 
import random
from openpyxl.styles import Font, Alignment

# Configure page
st.set_page_config(
    page_title="OPD Creator",      # shown in browser tab
    layout="wide",                 # full‑width
    initial_sidebar_state="expanded"
)

# Main title
st.title("Outpatient Department (OPD) Schedule Creator")

# Sidebar
st.sidebar.title("Navigation")
page = st.sidebar.radio("Go to:", [
    "Home",
    "Create OPD",
    "Upload Files",
    "Generate Schedule",
    "Download OPD"
])

file_configs = {
    "HAMPDEN_NURSERY.xlsx": {"title": "HAMPDEN NURSERY","custom_text": "CUSTOM_PRINT","names": ["Folaranmi, Oluwamayoda", "Alur, Pradeep", "Nanda, Sharmilarani", "HAMPDEN_NURSERY"]},
    "SJR_HOSP.xlsx": {"title": "SJR HOSPITALIST","custom_text": "CUSTOM_PRINT","names": ["Spangola, Haley", "Gubitosi, Terry", "SJR_1", "SJR_2"]}, 
    "AAC.xlsx": {"title": "AAC","custom_text": "CUSTOM_PRINT","names": ["Vaishnavi Harding", "Abimbola Ajayi", "Shilu Joshi", "Desiree Webb", "Amy Zisa", "Abdullah Sakarcan", "Anna Karasik", "AAC_1", "AAC_2", "AAC_3"]},
    "AL.xlsx": {"title": "AL","custom_text": "CUSTOM_PRINT","names": ["Aholoukpe, Mahoussi"]}, 
}

def generate_excel_file(start_date, title, custom_text, file_name, names):
    """
    Generates an Excel file where each week's structure aligns properly.

    Args:
        start_date (datetime): The starting date provided by the user.
        title (str): The text to be placed in cell A1.
        custom_text (str): The text to be placed in cell A2.
        file_name (str): The name of the output file.
        names (list): A list of names to be placed in the file.

    Returns:
        str: Path of the saved file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Place specific text in designated cells
    ws["A1"] = title
    ws["A2"] = custom_text

    # Columns where "custom_value" should be placed
    custom_value_columns = ["A", "C", "E", "G", "I", "K", "M"]
    name_columns = ["B", "D", "F", "H", "J", "L", "N"]

    # Ensure names list has at least one name
    if not names:
        names = ["Default Name ~"]

    # Days of the week to be placed across the row
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

    # Initial row where the first week starts
    start_row = 4
    num_weeks = 5  # Define the number of weeks
    week_height = 13  # Number of rows per week (date row + names + custom_value rows)

    for week in range(num_weeks):  
        current_date = start_date + datetime.timedelta(weeks=week)

        # Place day names and corresponding dates
        for i, day in enumerate(days):
            col_letter = chr(65 + (i * 2))  # Convert to Excel column letters (A, C, E, G, I, K, M)
            ws[f"{col_letter}{start_row}"] = day  
            formatted_date = (current_date + datetime.timedelta(days=i)).strftime("%B %-d, %Y")
            ws[f"{col_letter}{start_row + 1}"] = formatted_date  

        # Start placing names **immediately after the date row**
        names_start_row = start_row + 2  
        names_end_row = names_start_row + len(names)

        for i, col in enumerate(name_columns):
            custom_col = custom_value_columns[i]  # Get the column to the left
            for j, name in enumerate(names):
                row = names_start_row + j
                ws[f"{col}{row}"] = name  # Place the name
                ws[f"{custom_col}{row}"] = "custom_value"  # Place "custom_value" in the left column

        # Fill remaining rows with "custom_value" from the **date row** to the **next week's date row**
        next_week_start = start_row + week_height  # Set end range dynamically
        for i, col in enumerate(custom_value_columns):
            for row in range(start_row + 1, next_week_start):  # Fill from date row up to next week's start row
                if row >= names_end_row:  # Avoid overwriting names
                    ws[f"{col}{row}"] = "custom_value"

        # Move to the next week's section
        start_row = next_week_start  

    # Save the Excel file
    file_path = f"{file_name}"
    wb.save(file_path)
    
# Example of reacting to the sidebar choice
if page == "Home":
    st.write("👋 Welcome! Use the sidebar to navigate through the app.")
# … after your sidebar radio and imports …
st.set_page_config(page_title="OPD Creator", layout="wide")

# 1️⃣ Ensure a default page in session_state
if "page" not in st.session_state:
    st.session_state.page = "Home"

# 2️⃣ Bind the sidebar radio to that same session_state key
page = st.sidebar.radio(
    "Go to:",
    ["Home", "Create OPD", "Upload Files", "Generate Schedule", "Download OPD"],
    key="page"
)

elif page == "Create OPD":
    st.header("Date Input for OPD")
    st.write("Enter start date in **m/d/yyyy** (no leading zeros, e.g. 7/6/2021):")

    date_input = st.text_input("Start Date")

    if st.button("Submit Date") and date_input:
        try:
            start_date = datetime.datetime.strptime(date_input, "%m/%d/%Y")
            end_date   = start_date + datetime.timedelta(days=34)
            st.session_state.start_date = start_date
            st.session_state.end_date   = end_date

            st.success(
                f"✅ Valid date: {start_date:%B %d, %Y}  |  "
                f"Range: {start_date:%B %d, %Y} ➝ {end_date:%B %d, %Y}"
            )

            # generate all clinic workbooks
            generated = {}
            for fname, cfg in file_configs.items():
                path = generate_excel_file(
                    start_date,
                    cfg["title"],
                    cfg["custom_text"],
                    fname,
                    cfg["names"]
                )
                generated[fname] = path
            st.session_state.generated_files = generated

            # ➡️ Navigate to the next page
            st.session_state.page = "Upload Files"
            st.rerun()

        except ValueError:
            st.error("❌ Invalid format. Please use m/d/yyyy.")


elif page == "Upload Files":
    st.write("⬆️ Upload your Excel/CSV files to be processed.")
elif page == "Generate Schedule":
    st.write("⚙️ Processing files and assigning students...")
elif page == "Download OPD":
    st.write("👇 Finally, download your completed OPD.xlsx here.")

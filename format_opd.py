# format_opd.py
import io
import xlsxwriter
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def generate_opd_workbook(full_df: pd.DataFrame) -> bytes:
    import io
    import xlsxwriter

    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})

    # ─── Formats ─────────────────────────────────────────────────────────────────
    format1     = workbook.add_format({
        'font_size':18,'bold':1,'align':'center','valign':'vcenter',
        'font_color':'black','bg_color':'#FEFFCC','border':1
    })
    format4     = workbook.add_format({
        'font_size':12,'bold':1,'align':'center','valign':'vcenter',
        'font_color':'black','bg_color':'#8ccf6f','border':1
    })
    format4a    = workbook.add_format({
        'font_size':12,'bold':1,'align':'center','valign':'vcenter',
        'font_color':'black','bg_color':'#9fc5e8','border':1
    })
    format5     = workbook.add_format({
        'font_size':12,'bold':1,'align':'center','valign':'vcenter',
        'font_color':'black','bg_color':'#FEFFCC','border':1
    })
    format5a    = workbook.add_format({
        'font_size':12,'bold':1,'align':'center','valign':'vcenter',
        'font_color':'black','bg_color':'#d0e9ff','border':1
    })
    format11    = workbook.add_format({
        'font_size':18,'bold':1,'align':'center','valign':'vcenter',
        'font_color':'black','bg_color':'#FEFFCC','border':1
    })
    formate     = workbook.add_format({
        'font_size':12,'bold':0,'align':'center','valign':'vcenter',
        'font_color':'white','border':0
    })
    format3     = workbook.add_format({
        'font_size':12,'bold':1,'align':'center','valign':'vcenter',
        'font_color':'black','bg_color':'#FFC7CE','border':1
    })
    format2     = workbook.add_format({'bg_color':'black'})
    format_date = workbook.add_format({
        'num_format':'m/d/yyyy','font_size':12,'bold':1,
        'align':'center','valign':'vcenter',
        'font_color':'black','bg_color':'#FFC7CE','border':1
    })
    format_label= workbook.add_format({
        'font_size':12,'bold':1,'align':'center','valign':'vcenter',
        'font_color':'black','bg_color':'#FFC7CE','border':1
    })
    merge_format= workbook.add_format({
        'bold':1,'align':'center','valign':'vcenter','text_wrap':True,
        'font_color':'red','bg_color':'#FEFFCC','border':1
    })

    # ─── Worksheets ─────────────────────────────────────────────────────────────
    worksheet_names = [
        'HOPE_DRIVE','ETOWN','NYES','COMPLEX',
        'W_A','PSHCH_NURSERY','HAMPDEN_NURSERY','SJR_HOSP','AAC','AHOLOUKPE','ADOLMED'
    ]
    sheets = {name: workbook.add_worksheet(name) for name in worksheet_names}

    # ─── Site headers ────────────────────────────────────────────────────────────
    site_list = [
        'Hope Drive','Elizabethtown','Nyes Road','Complex Care',
        'WARD A','PSHCH NURSERY','HAMPDEN NURSERY','SJR HOSPITALIST','AAC','AHOLOUKPE','ADOLMED'
    ]
    for ws, site in zip(sheets.values(), site_list):
        ws.write(0, 0, 'Site:', format1)
        ws.write(0, 1, site,   format1)

    # ─── HOPE_DRIVE specific ────────────────────────────────────────────────────
    hd = sheets['HOPE_DRIVE']
    for cr in ['A8:H15','A32:H39','A56:H63','A80:H87']:
        hd.conditional_format(cr, {'type':'cell','criteria':'>=','value':0,'format':format1})
    for cr in ['A18:H25','A42:H49','A66:H73','A90:H97']:
        hd.conditional_format(cr, {'type':'cell','criteria':'>=','value':0,'format':format5a})
    for cr in ['A6:H6','A7:H7','A30:H30','A31:H31','A54:H54','A55:H55','A78:H78','A79:H79']:
        hd.conditional_format(cr, {'type':'cell','criteria':'>=','value':0,'format':format4})
    for cr in ['A16:H16','A17:H17','A40:H40','A41:H41','A64:H64','A65:H65','A88:H88','A89:H89']:
        hd.conditional_format(cr, {'type':'cell','criteria':'>=','value':0,'format':format4a})
    acute_ranges = [(6,7),(16,17),(30,31),(40,41),(54,55),(64,65),(78,79),(88,89)]
    for r1, r2 in acute_ranges:
        fmt   = format4 if r1 % 2 == 0 else format4a
        label = 'AM - ACUTES' if r1 % 2 == 0 else 'PM - ACUTES'
        for r in range(r1, r2+1):
            hd.write(f'A{r}', label, fmt)
    cont_ranges = [(8,15),(18,25),(32,39),(42,49),(56,63),(66,73),(80,87),(90,97)]
    for r1, r2 in cont_ranges:
        for r in range(r1, r2+1):
            hd.write(
                f'A{r}',
                'AM - Continuity' if r1 % 2 == 0 else 'PM - Continuity',
                format5a
            )
    labels = [f'H{i}' for i in range(20)]
    for start in [6, 30, 54, 78]:
        for i, lab in enumerate(labels):
            hd.write(f'I{start+i}', lab, formate)

    # ─── GENERIC SHEETS ─────────────────────────────────────────────────────────
    others       = [ws for name, ws in sheets.items() if name != 'HOPE_DRIVE']
    AM_COUNT     = 10
    PM_COUNT     = 10
    BLOCK_STARTS = [6, 30, 54, 78]

    for ws in others:
        # 1) conditional formats
        for cr in ['A6:H15','A30:H39','A54:H63','A78:H87']:
            ws.conditional_format(cr, {
                'type':'cell','criteria':'>=','value':0,'format':format1
            })
        for cr in ['A16:H25','A40:H49','A64:H73','A88:H97']:
            ws.conditional_format(cr, {
                'type':'cell','criteria':'>=','value':0,'format':format5a
            })
        for cr, fmt in [
            ('B6:H6',   format4), ('B16:H16', format4a),
            ('B30:H30', format4), ('B40:H40', format4a),
            ('B54:H54', format4), ('B64:H64', format4a),
            ('B78:H78', format4), ('B88:H88', format4a)
        ]:
            ws.conditional_format(cr, {
                'type':'cell','criteria':'>=','value':0,'format':fmt
            })

        # 2) Write exactly 10 AM then 10 PM in column A
        for start in BLOCK_STARTS:
            zero_row = start - 1
            for i in range(AM_COUNT):
                ws.write(zero_row + i, 0, 'AM', format5a)
            for i in range(PM_COUNT):
                ws.write(zero_row + AM_COUNT + i, 0, 'PM', format5a)
            for i, lab in enumerate(labels):
                ws.write(start + i, 8, lab, formate)

        # 3) Write H0…H19 in column I
        for start in BLOCK_STARTS:
            for i, lab in enumerate(labels):
                ws.write(f'I{start + i}', lab, formate)


    # ─── Universal formatting & dates ────────────────────────────────────────────
    date_cols = [f"hd_day_date{i}" for i in range(1,29)]
    dates     = pd.to_datetime(full_df[date_cols].iloc[0]).tolist()
    weeks     = [dates[i*7:(i+1)*7] for i in range(4)]
    days      = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']

    for ws in workbook.worksheets():
        ws.set_zoom(80)
        ws.set_column('A:A', 10)
        ws.set_column('B:H', 65)
        ws.set_row(0, 37.25)

        for idx, start in enumerate([2,26,50,74]):
                    # day names
                    for c, d in enumerate(days):
                        ws.write(start, 1+c, d, format3)
                    # dates
                    for c, val in enumerate(weeks[idx]):
                        ws.write(start+1, 1+c, val, format_date)
                        
                    # padding formula bars
                    ws.write_formula(f'A{start}',   '""', format_label)
                    
                    ws.conditional_format(
                        f'A{start+3}:H{start+3}',
                        {'type':'cell','criteria':'>=','value':0,'format':format_label}
                    )


        # black bars every 24 rows
        step = 24
        for row in range(2, 98, step):
            ws.merge_range(f'A{row}:H{row}', ' ', format2)

        # merge CRTS message on every sheet
        text1 = (
            'Students are to alert their preceptors when they have a Clinical '
            'Reasoning Teaching Session (CRTS).  Please allow the students to '
            'leave approximately 15 minutes prior to the start of their session '
            'so they can be prepared to actively participate.  ~ Thank you!'
        )
        ws.merge_range('C1:F1', text1, merge_format)
        ws.write('G1', '', merge_format)
        ws.write('H1', '', merge_format)

        #PAINT Empty White Spaces
        ws.write('A3', '', format_date)
        ws.write('A4', '', format_date)
        ws.write('A27', '', format_date)
        ws.write('A28', '', format_date)
        ws.write('A51', '', format_date)
        ws.write('A52', '', format_date)
        ws.write('A75', '', format_date)
        ws.write('A76', '', format_date)


    workbook.close()
    output.seek(0)
    return output.read()

def update_excel_from_csv(excel_template_bytes: bytes, csv_data_bytes: bytes, mappings: list) -> bytes | None:
    import io
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    """
    Updates an Excel file (from bytes) with values from a CSV (from bytes)
    based on provided mappings, and returns the updated Excel as bytes.

    Args:
        excel_template_bytes (bytes): The bytes content of the Excel file to be updated.
        csv_data_bytes (bytes): The bytes content of the CSV file containing the data.
        mappings (list of dict): A list of dictionaries, where each dictionary
                                 defines a mapping:
                                 {'csv_column': 'name_of_csv_column',
                                  'excel_sheet': 'Sheet Name',
                                  'excel_cell': 'Cell Address (e.g., B8)'}
    Returns:
        bytes: The bytes content of the updated Excel workbook, or None if an error occurs.
    """
    try:
        # Load the CSV data from bytes using BytesIO
        df_csv = pd.read_csv(io.BytesIO(csv_data_bytes))

        if df_csv.empty:
            # Assuming st is available in the Streamlit environment
            # If running outside Streamlit, you might use print() or logging
            # st.error("Error: The CSV data is empty. Cannot update Excel.")
            return None

        # Load the Excel workbook from bytes using BytesIO
        wb = load_workbook(io.BytesIO(excel_template_bytes))

        # Iterate through the mappings and update the Excel file
        for mapping in mappings:
            csv_column = mapping['csv_column']
            excel_sheet_name = mapping['excel_sheet']
            excel_cell = mapping['excel_cell']

            if csv_column not in df_csv.columns:
                # st.warning(f"Warning: CSV column '{csv_column}' not found in CSV data. Skipping this mapping.")
                continue

            # Get the value from the first row of the specified CSV column
            value_to_transfer = df_csv.loc[0, csv_column]

            if excel_sheet_name not in wb.sheetnames:
                # st.warning(f"Warning: Excel sheet '{excel_sheet_name}' not found in the Excel template. Skipping this mapping.")
                continue

            ws = wb[excel_sheet_name]

            # --- APPLY THE REQUESTED FORMATTING HERE ---
            # 1. Convert to string and append " ~ "
            # This ensures that even if value_to_transfer is a number, it can be concatenated
            formatted_value = str(value_to_transfer) + ' ~ '

            # 2. Write the formatted value to the cell
            ws[excel_cell] = formatted_value

            # 3. Set the alignment for the cell using openpyxl's Alignment
            cell = ws[excel_cell] # Get the cell object
            cell.alignment = Alignment(horizontal='center', vertical='center') # Set horizontal and vertical centering

            # st.info(f"Successfully wrote '{formatted_value}' from CSV column '{csv_column}' to '{excel_sheet_name}'!{excel_cell}")

        # Save the modified Excel workbook to a BytesIO object
        output_excel_bytes_io = io.BytesIO()
        wb.save(output_excel_bytes_io)
        output_excel_bytes_io.seek(0) # Rewind the buffer to the beginning

        return output_excel_bytes_io.getvalue()

    except Exception as e:
        # st.error(f"An error occurred during Excel update: {e}")
        return None

# --- Configuration for update_excel_from_csv (your mappings) ---
data_mappings        = []
excel_column_letters = ['B','C','D','E','F','G','H']
num_weeks            = 4

# HOPE_DRIVE acute + continuity row offsets
hd_row_defs = {
    'AM': {'acute_start': 6,  'cont_start': 8},
    'PM': {'acute_start': 16, 'cont_start': 18},
}

# Other sheets only need continuity (rows 6–13 for AM, 16–23 for PM)
cont_row_defs = {
    'AM':  6,
    'PM': 16,
}

# your prefix map
base_map = {
    "hope drive am continuity":      "hd_am_",
    "hope drive pm continuity":      "hd_pm_",
    "hope drive am acute precept":   "hd_am_acute_",
    "hope drive pm acute precept":   "hd_pm_acute_",
    "hope drive weekend acute 1":    "hd_wknd_acute_1_",
    "hope drive weekend acute 2":    "hd_wknd_acute_2_",
    "hope drive weekend continuity": "hd_wknd_am_",
    
    "etown am continuity":           "etown_am_",
    "etown pm continuity":           "etown_pm_",
    
    "nyes rd am continuity":         "nyes_am_",
    "nyes rd pm continuity":         "nyes_pm_",
    
    "nursery weekday 8a-6p":         ["nursery_am_","nursery_pm_"],
    
    "rounder 1 7a-7p":               ["ward_a_am_","ward_a_pm_"],
    "rounder 2 7a-7p":               ["ward_a_am_","ward_a_pm_"],
    "rounder 3 7a-7p":               ["ward_a_am_","ward_a_pm_"],
    
    "hope drive clinic am":          "complex_am_",
    "hope drive clinic pm":          "complex_pm_",
    
    "briarcrest clinic am":          "adol_med_am_",
    "briarcrest clinic pm":          "adol_med_pm_",

    'hampden_nursery_print':    'custom_print_hampden_nursery_',
    'sjr_hospitalist_print':    'custom_print_sjr_hospitalist_',
    'aac_print':                'custom_print_aac_',

    'mahoussi_aholoukpe_print': 'custom_print_mahoussi_aholoukpe_',
    
}

# which keys from base_map for each sheet
sheet_map = {
    'ETOWN':           ('etown am continuity','etown pm continuity'),
    'NYES':            ('nyes rd am continuity','nyes rd pm continuity'),
    'COMPLEX':         ('hope drive clinic am','hope drive clinic pm'),
    'W_A':             ('rounder 1 7a-7p','rounder 2 7a-7p','rounder 3 7a-7p'),
    'PSHCH_NURSERY':    ("nursery weekday 8a-6p","nursery weekday 8a-6p"),
    
    'HAMPDEN_NURSERY': ('hampden_nursery_print',),
    'SJR_HOSP':        ('sjr_hospitalist_print',),
    'AAC':             ('aac_print',),
    'AHOLOUKPE':        ('mahoussi_aholoukpe_print',),
    
    'ADOLMED':             ('briarcrest clinic am','briarcrest clinic pm'),
}

worksheet_names = ['HOPE_DRIVE','ETOWN','NYES','COMPLEX','W_A','PSHCH_NURSERY','HAMPDEN_NURSERY','SJR_HOSP','AAC','AHOLOUKPE','ADOLMED']



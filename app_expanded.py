import streamlit as st
import csv
import datetime
import pandas as pd
import numpy as np
import datetime
from datetime import timedelta
import xlsxwriter
import openpyxl
from openpyxl import Workbook
from io import BytesIO
from io import StringIO
import os
import time 
import random 

st.set_page_config(layout="wide")

def format_date_with_suffix(date):
    """Formats a date as 'Month Day[st/nd/rd/th], Year' (e.g., 'February 3rd, 2025')."""
    day = date.day
    suffix = "th" if 11 <= day <= 13 else {1: "st", 2: "nd", 3: "rd"}.get(day % 10, "th")
    return date.strftime(f"%B {day}{suffix}, %Y")
	
file_configs = {
    "HAMPDEN_NURSERY.xlsx": {"title": "HAMPDEN NURSERY","custom_text": "CUSTOM_PRINT","names": ["Folaranmi, Oluwamayoda", "Alur, Pradeep", "Nanda, Sharmilarani", "HAMPDEN_NURSERY"]},
    "SJR_HOSP.xlsx": {"title": "SJR HOSPITALIST","custom_text": "CUSTOM_PRINT","names": ["Spangola, Haley", "Gubitosi, Terry", "SJR_1", "SJR_2"]}, 
    "AAC.xlsx": {"title": "AAC","custom_text": "CUSTOM_PRINT","names": ["Vaishnavi Harding", "Abimbola Ajayi", "Shilu Joshi", "Desiree Webb", "Amy Zisa", "Abdullah Sakarcan", "Anna Karasik", "AAC_1", "AAC_2", "AAC_3"]} #LIST ALL NAMES
}

def generate_excel_file(start_date, title, custom_text, file_name, names):
    """
    Generates an Excel file where each week's structure aligns properly.

    Args:
        start_date (datetime): The starting date provided by the user.
        title (str): The text to be placed in cell A1.
        custom_text (str): The text to be placed in cell A2.
        file_name (str): The name of the output file.
        names (list): A list of names to be placed in the file.

    Returns:
        str: Path of the saved file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Place specific text in designated cells
    ws["A1"] = title
    ws["A2"] = custom_text

    # Columns where "custom_value" should be placed
    custom_value_columns = ["A", "C", "E", "G", "I", "K", "M"]
    name_columns = ["B", "D", "F", "H", "J", "L", "N"]

    # Ensure names list has at least one name
    if not names:
        names = ["Default Name ~"]

    # Days of the week to be placed across the row
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

    # Initial row where the first week starts
    start_row = 4
    num_weeks = 5  # Define the number of weeks
    week_height = 13  # Number of rows per week (date row + names + custom_value rows)

    for week in range(num_weeks):  
        current_date = start_date + datetime.timedelta(weeks=week)

        # Place day names and corresponding dates
        for i, day in enumerate(days):
            col_letter = chr(65 + (i * 2))  # Convert to Excel column letters (A, C, E, G, I, K, M)
            ws[f"{col_letter}{start_row}"] = day  
            formatted_date = (current_date + datetime.timedelta(days=i)).strftime("%B %-d, %Y")
            ws[f"{col_letter}{start_row + 1}"] = formatted_date  

        # Start placing names **immediately after the date row**
        names_start_row = start_row + 2  
        names_end_row = names_start_row + len(names)

        for i, col in enumerate(name_columns):
            custom_col = custom_value_columns[i]  # Get the column to the left
            for j, name in enumerate(names):
                row = names_start_row + j
                ws[f"{col}{row}"] = name  # Place the name
                ws[f"{custom_col}{row}"] = "custom_value"  # Place "custom_value" in the left column

        # Fill remaining rows with "custom_value" from the **date row** to the **next week's date row**
        next_week_start = start_row + week_height  # Set end range dynamically
        for i, col in enumerate(custom_value_columns):
            for row in range(start_row + 1, next_week_start):  # Fill from date row up to next week's start row
                if row >= names_end_row:  # Avoid overwriting names
                    ws[f"{col}{row}"] = "custom_value"

        # Move to the next week's section
        start_row = next_week_start  

    # Save the Excel file
    file_path = f"{file_name}"
    wb.save(file_path)

    # ✅ **Display & Download Immediately**
    st.success(f"✅ File '{file_name}' has been successfully created!")

    #df_display = pd.read_excel(file_path, dtype=str)
    #st.dataframe(df_display); #time.sleep(30); # Display file in Streamlit

    with open(file_path, "rb") as f:
        st.download_button("Download Generated Excel File", f, file_name, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"); return file_path  # Return file path for later use
	
# Initialize session state variables efficiently
session_defaults = {
    "page": "Home",
    "start_date": None,
    "uploaded_files": {},
    "uploaded_book4_file": {},
}
for key, value in session_defaults.items():
    st.session_state.setdefault(key, value)

# Function to change page and trigger rerun
def navigate_to(page):
    st.session_state.page = page
    st.rerun()

# Home Page
if st.session_state.page == "Home":
    st.title("Welcome to OPD Creator")
    st.write("Please choose what you'd like to do next.")

    if st.button("Go to Create OPD"):
        navigate_to("Create OPD")

    if st.button("Go to Create Student Schedule"):
        navigate_to("Create Student Schedule")

elif st.session_state.page == "Create OPD":
    st.title('Date Input for OPD')
    st.write('Enter start date in **m/d/yyyy format**, no leading zeros (e.g., 7/6/2021):')

    # User Input for Start Date
    date_input = st.text_input('Start Date')


    if st.button('Submit Date') and date_input:
        try:
            start_date, end_date = datetime.datetime.strptime(date_input, "%m/%d/%Y"), datetime.datetime.strptime(date_input, "%m/%d/%Y") + datetime.timedelta(days=34)
            st.session_state.start_date, st.session_state.end_date = start_date, end_date
		
            st.success(f"✅ Valid date entered: {start_date.strftime('%B %d, %Y')} | 📅 Date range: {start_date.strftime('%B %d, %Y')} ➝ {end_date.strftime('%B %d, %Y')}")

            # Generate all predefined Excel files
            generated_files = {}
            for file_name, config in file_configs.items():
                file_path = generate_excel_file(start_date, config["title"], config["custom_text"], file_name, config["names"])
                generated_files[file_name] = file_path

            # Store file paths in session state for later downloads
            st.session_state.generated_files = generated_files

            # Move to the next page: Upload Files
            st.session_state.page = "Upload Files"
            st.rerun()  # Force rerun to reflect the page change
    
        except ValueError:
            st.error('Invalid date format. Please enter the date in **m/d/yyyy** format.')

elif st.session_state.page == "Upload Files":
    st.title("File Upload Section")
    st.write("Upload the required Excel files:")

    # Define file name mappings based on content identifiers
    file_identifiers = {
        "Academic General Pediatrics": ["NYES.xlsx", "HOPE_DRIVE.xlsx", "ETOWN.xlsx", "PSHCH_NURSERY.xlsx"],
        "Pulmonary": ["WARD_P.xlsx"],
        "Hospitalists": ["WARD_A.xlsx"],
        "Cardiology": ["WARD_CARDIOLOGY.xlsx"],
        "Neph": ["WARD_NEPHRO.xlsx"],
        "PICU": ["PICU.xlsx"],
        "GI Daytime Service": ["WARD_GI.xlsx"],
        "Complex": ["COMPLEX.xlsx"],
        "Adol Med": ["ADOLMED.xlsx"], 
        "cl5rks1p": ["Book4.xlsx"], 
        "Rotation": ["RESIDENT.xlsx"], 
    }
    # Required files for validation
    required_files = set(file for filenames in file_identifiers.values() for file in filenames)

    # Streamlit UI
    st.title("File Upload Section")
    st.write("Upload the following required Excel files:")

    # Ensure start_date and end_date exist in session state
    if "start_date" in st.session_state and "end_date" in st.session_state:
        start_date, end_date = st.session_state.start_date, st.session_state.end_date
        st.success(f"✅ Valid date entered: {start_date.strftime('%B %d, %Y')} | 📅 Date range: {start_date.strftime('%B %d, %Y')} ➝ {end_date.strftime('%B %d, %Y')}")
    else:
        st.error("❌ No valid date found. Please enter a start date first.")

    # File uploader
    uploaded_files = st.file_uploader("Choose your files", type="xlsx", accept_multiple_files=True)

    if uploaded_files:
        uploaded_files_dict = {}
        detected_files = set()

        for file in uploaded_files:
            try:
                # Read the first few rows of the Excel file
                df = pd.read_excel(file, dtype=str, nrows=10)  

                # Normalize text: strip spaces, handle line breaks, convert to lowercase
                df_clean = df.astype(str).apply(lambda x: x.str.strip().str.replace("\n", " ").str.lower())

                # Convert all values into a single string for better search
                full_text = " ".join(df_clean.to_string().split()).lower()

                # Assign multiple filenames for "Academic General Pediatrics"
                found_files = []
                for key, expected_filenames in file_identifiers.items():
                    if key.lower() in full_text:
                        found_files.extend(expected_filenames)

                if found_files:
                    for expected_filename in found_files:
                        uploaded_files_dict[expected_filename] = file  # Assign the same file to multiple expected filenames
                        detected_files.add(expected_filename)

                else:
                    st.warning(f"⚠️ Could not automatically detect file type for: {file.name}")

            except Exception as e:
                st.error(f"❌ Error reading {file.name}: {str(e)}")

        # Save detected files to session state
        st.session_state.uploaded_files = uploaded_files_dict

        # Check for missing files
        missing_files = required_files - detected_files

        if not missing_files:
            st.success("✅ All required files uploaded and detected successfully!")
            navigate_to("OPD Creator")
        else:
            st.error(f"❌ Missing files: {', '.join(missing_files)}. Please upload all required files.")

		
elif st.session_state.page == "OPD Creator":
	#test_date = datetime.datetime.strptime(x, "%m/%d/%Y")
	test_date = st.session_state.start_date
	uploaded_files = st.session_state.uploaded_files
	
	# initializing K
	K = 28
	 
	res = []
	 
	for day in range(K):
	    date = (test_date + datetime.timedelta(days = day)).strftime("%-m/%-d/%Y")
	    res.append(date)
	     
	#res
	
	dates = pd.DataFrame(res, columns =['dates'])
	
	dates['x'] = "y"
	
	dates['i'] = dates.index+1
	
	dates['x'] = dates['x'].astype(str)+dates['i'].astype(str)
	
	dates['x'] = dates['x'].astype(str) + "=" + "'"+dates['dates'].astype(str) + "'"
	
	dates = dates[['x']]
	
	dates.to_csv('dates.csv',index=False)
	
	import numpy as np
	
	datesdf = pd.read_csv('dates.csv')
	
	dates = datesdf['x'].astype(str)
	
	numpy_array=dates.to_numpy()
	np.savetxt("dates.py",numpy_array, fmt="%s")
	
	exec(open('dates.py').read())
	import xlsxwriter

	# Create workbook
	workbook = xlsxwriter.Workbook('OPD.xlsx')
	
	# Define worksheet names
	worksheet_names = ['HOPE_DRIVE', 'ETOWN', 'NYES', 'COMPLEX', 'W_A', 'W_C','W_P', 'PICU', 'PSHCH_NURSERY', 'HAMPDEN_NURSERY','SJR_HOSP', 'AAC', 'ER_CONS','NF',"ADOLMED","RESIDENT"]
	
	# Create worksheets and store them in a dictionary
	worksheets = {name: workbook.add_worksheet(name) for name in worksheet_names}
	(worksheet, worksheet2, worksheet3, worksheet4, worksheet5, worksheet6, worksheet7, worksheet8, worksheet9, worksheet10, worksheet11, worksheet12, worksheet13, worksheet14,worksheet15, worksheet16) = worksheets.values()
	
	# Define format
	format1 = workbook.add_format({'font_size': 18, 'bold': 1, 'align': 'center','valign': 'vcenter', 'font_color': 'black','bg_color': '#FEFFCC', 'border': 1})
	
	# Define site names corresponding to worksheet names
	worksheet_sites = {worksheet: 'Hope Drive', 
			   worksheet2: 'Elizabethtown', 
			   worksheet3: 'Nyes Road', 
			   worksheet4: 'Complex Care', 
			   worksheet5: 'WARD A', 
			   worksheet6: 'WARD C', 
			   worksheet7: 'WARD P', 
			   worksheet8: 'PICU', 
			   worksheet9: 'PSHCH NURSERY', 
			   worksheet10: 'HAMPDEN NURSERY',
			   worksheet11: 'SJR HOSPITALIST', 
			   worksheet12: 'AAC', 
			   worksheet13: 'ER CONSULTS', 
			   worksheet14: 'NIGHT FLOAT', 
			   worksheet15: 'ADOLMED', worksheet16: 'RESIDENT'}
	
	# Write "Site:" and corresponding site names in each worksheet
	for ws, site in worksheet_sites.items():
	    ws.write(0, 0, 'Site:', format1)
	    ws.write(0, 1, site, format1)
		
	#Color Coding
	format4 = workbook.add_format({'font_size':12,'bold': 1,'align': 'center','valign': 'vcenter','font_color':'black','bg_color':'#8ccf6f','border':1})
	format4a = workbook.add_format({'font_size':12,'bold': 1,'align': 'center','valign': 'vcenter','font_color':'black','bg_color':'#9fc5e8','border':1})    
	format5 = workbook.add_format({'font_size':12,'bold': 1,'align': 'center','valign': 'vcenter','font_color':'black','bg_color':'#FEFFCC','border':1})
	format5a = workbook.add_format({'font_size':12,'bold': 1,'align': 'center','valign': 'vcenter','font_color':'black','bg_color':'#d0e9ff','border':1})
	format11 = workbook.add_format({'font_size':18,'bold': 1,'align': 'center','valign': 'vcenter','font_color':'black','bg_color':'#FEFFCC','border':1})
	
	#H codes
	formate = workbook.add_format({'font_size':12,'bold': 0,'align': 'center','valign': 'vcenter','font_color':'white','border':0})
	
	# HOPE_DRIVE COLOR CODING AND IDENTIFYING ACUTE VERSUS CONTINUITY
	ranges_format1 = ['A8:H15', 'A32:H39', 'A56:H63', 'A80:H87']
	ranges_format5a = ['A18:H25', 'A42:H49', 'A66:H73', 'A90:H97']
	
	for cell_range in ranges_format1:
	    worksheet.conditional_format(cell_range, {'type': 'cell', 'criteria': '>=', 'value': 0, 'format': format1})
	
	for cell_range in ranges_format5a:
	    worksheet.conditional_format(cell_range, {'type': 'cell', 'criteria': '>=', 'value': 0, 'format': format5a})
	
	# HOPE_DRIVE CONDITIONAL FORMATTING
	ranges_format4 = ['A6:H6', 'A7:H7', 'A30:H30', 'A31:H31', 'A54:H54', 'A55:H55', 'A78:H78', 'A79:H79']
	ranges_format4a = ['A16:H16', 'A17:H17', 'A40:H40', 'A41:H41', 'A64:H64', 'A65:H65', 'A88:H88', 'A89:H89']
	
	for cell_range in ranges_format4:
	    worksheet.conditional_format(cell_range, {'type': 'cell', 'criteria': '>=', 'value': 0, 'format': format4})
	
	for cell_range in ranges_format4a:
	    worksheet.conditional_format(cell_range, {'type': 'cell', 'criteria': '>=', 'value': 0, 'format': format4a})
	
	# HOPE_DRIVE WRITING ACUTE AND CONTINUITY LABELS
	acute_format_ranges = [
	    (6, 7, 'AM - ACUTES', format4), (16, 17, 'PM - ACUTES', format4a), 
	    (30, 31, 'AM - ACUTES', format4), (40, 41, 'PM - ACUTES', format4a),
	    (54, 55, 'AM - ACUTES', format4), (64, 65, 'PM - ACUTES', format4a),
	    (78, 79, 'AM - ACUTES', format4), (88, 89, 'PM - ACUTES', format4a)
	]
	
	continuity_format_ranges = [
	    (8, 15, 'AM - Continuity', format5a), (18, 25, 'PM - Continuity', format5a),
	    (32, 39, 'AM - Continuity', format5a), (42, 49, 'PM - Continuity', format5a),
	    (56, 63, 'AM - Continuity', format5a), (66, 73, 'PM - Continuity', format5a),
	    (80, 87, 'AM - Continuity', format5a), (90, 97, 'PM - Continuity', format5a)
	]
	
	# Write Acute Labels
	for start_row, end_row, label, fmt in acute_format_ranges:
	    for row in range(start_row, end_row + 1):
	        worksheet.write(f'A{row}', label, fmt)
	
	# Write Continuity Labels
	for start_row, end_row, label, fmt in continuity_format_ranges:
	    for row in range(start_row, end_row + 1):
	        worksheet.write(f'A{row}', label, fmt)
	
	# Define the labels
	#labels = ['HX1', 'HX2', 'H2', 'H3', 'H4', 'H5', 'H6', 'H7', 'H8', 'H9', 'HXX1', 'HXX2', 'H12', 'H13', 'H14', 'H15', 'H16', 'H17', 'H18', 'H19']
	
	labels = ['H{}'.format(i) for i in range(20)]
	
	# Define the starting rows for each group
	start_rows = [6, 30, 54, 78]
	
	# Write the labels in each group
	for start_row in start_rows:
	    for i, label in enumerate(labels):
	        worksheet.write(f'I{start_row + i}', label, formate)
	
	# Simplify common formatting and label assignment for worksheets 2, 3, 4, 5
	worksheets = [worksheet2, worksheet3, worksheet4, worksheet5, worksheet6, worksheet7, worksheet8, worksheet9, worksheet10, worksheet11, worksheet12, worksheet13, worksheet14, worksheet15,worksheet16]
	
	ranges_format1 = ['A6:H15', 'A30:H39', 'A54:H63', 'A78:H87']
	ranges_format5a = ['A16:H25', 'A40:H49', 'A64:H73', 'A88:H97']
	specific_format_ranges = [
	    ('B6:H6', format4), ('B16:H16', format4a),
	    ('B30:H30', format4), ('B40:H40', format4a),
	    ('B54:H54', format4), ('B64:H64', format4a),
	    ('B78:H78', format4), ('B88:H88', format4a)
	]
	
	am_pm_labels = ['AM'] * 10 + ['PM'] * 10
	h_labels = ['H{}'.format(i) for i in range(20)]
	
	for worksheet in worksheets:
	    # Apply conditional formatting
	    for cell_range in ranges_format1:
	        worksheet.conditional_format(cell_range, {'type': 'cell', 'criteria': '>=', 'value': 0, 'format': format1})
	
	    for cell_range in ranges_format5a:
	        worksheet.conditional_format(cell_range, {'type': 'cell', 'criteria': '>=', 'value': 0, 'format': format5a})
	
	    for cell_range, fmt in specific_format_ranges:
	        worksheet.conditional_format(cell_range, {'type': 'cell', 'criteria': '>=', 'value': 0, 'format': fmt})
	
	    # Write AM/PM labels
	    sections = [(6, 25), (30, 49), (54, 73), (78, 97)]
	    for start_row, end_row in sections:
	        for i, label in enumerate(am_pm_labels):
	            worksheet.write(f'A{start_row + i}', label, format5a)
	
	    # Write H labels in column 'I'
	    start_rows = [6, 30, 54, 78]
	    for start_row in start_rows:
	        for i, label in enumerate(h_labels):
	            worksheet.write(f'I{start_row + i}', label, formate)
	
	# Loop through each worksheet in workbook
	for worksheet in workbook.worksheets():
	
	    # Set Zoom for all sheets
	    worksheet.set_zoom(80)
	
	    # Set Days
	    format3 = workbook.add_format({
	        'font_size': 12, 'bold': 1, 'align': 'center', 'valign': 'vcenter',
	        'font_color': 'black', 'bg_color': '#FFC7CE', 'border': 1
	    })
	    day_labels = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
	    start_rows = [2, 26, 50, 74] #[3, 27, 51, 75]
	    for start_row in start_rows:
	        for i, day in enumerate(day_labels):
	            worksheet.write(start_row, 1 + i, day, format3)  # B=1, C=2, etc.
	
	    # Set Date Formats
	    format_date = workbook.add_format({
	        'num_format': 'm/d/yyyy', 'font_size': 12, 'bold': 1, 'align': 'center', 'valign': 'vcenter',
	        'font_color': 'black', 'bg_color': '#FFC7CE', 'border': 1
	    })
	
	    format_label = workbook.add_format({
	        'font_size': 12, 'bold': 1, 'align': 'center', 'valign': 'vcenter',
	        'font_color': 'black', 'bg_color': '#FFC7CE', 'border': 1
	    })
	
	    # Set Date Formulas
	    date_rows = [3, 27, 51, 75] #[4, 28, 52, 76]
	    for i, start_row in enumerate(date_rows):
	        worksheet.write(f'A{start_row - 1}', "", format_label)
	        #worksheet.write_formula(f'A{start_row}', f'="Week of:"&" "&TEXT(B{start_row},"m/d/yy")', format_label) #If want to place Week of Date in
	        worksheet.write_formula(f'A{start_row}', f'=""', format_label)
	        worksheet.write(f'A{start_row + 1}', "", format_label)
	
	    # Set Pink Bars (Conditional Format)
	    pink_bar_rows = [5, 29, 53, 77]
	    for row in pink_bar_rows:
	        worksheet.conditional_format(f'A{row}:H{row}', {
	            'type': 'cell', 'criteria': '>=', 'value': 0, 'format': format_label
	        })
	
	    # Black Bars
	    format2 = workbook.add_format({'bg_color': 'black'})
	    black_bar_rows = [2, 26, 50, 74, 98]
	    for row in black_bar_rows:
	        worksheet.merge_range(f'A{row}:H{row}', " ", format2)
	        
	    # Write More Dates
	    date_values = [
	        [y1, y2, y3, y4, y5, y6, y7],
	        [y8, y9, y10, y11, y12, y13, y14],
	        [y15, y16, y17, y18, y19, y20, y21],
	        [y22, y23, y24, y25, y26, y27, y28]
	    ]
	    for i, start_row in enumerate(date_rows):
	        for j, value in enumerate(date_values[i]):
	            worksheet.write(start_row, 1 + j, value, format_date)  # B=1, C=2, etc.
	
	    # Set Column Widths
	    worksheet.set_column('A:A', 10)
	    worksheet.set_column('B:H', 65)
	    worksheet.set_row(0, 37.25)
	
	    # Merge Format for Text
	    merge_format = workbook.add_format({
	        'bold': 1, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True,
	        'font_color': 'red', 'bg_color': '#FEFFCC', 'border': 1
	    })
	    text1 = 'Students are to alert their preceptors when they have a Clinical Reasoning Teaching Session (CRTS).  Please allow the students to leave approximately 15 minutes prior to the start of their session so they can be prepared to actively participate.  ~ Thank you!'
	
	    # Merge and Write Important Message
	    worksheet.merge_range('C1:F1', text1, merge_format)
	    worksheet.write('G1', "", merge_format)
	    worksheet.write('H1', "", merge_format)
	
	    # Close Workbook
	    workbook.close()
	
	    import openpyxl
	    from io import BytesIO
	    import streamlit as st
	
	    if st.button('Create OPD'):
	        file_path = 'OPD.xlsx'  # Path to your workbook
	
	        # Load the workbook using openpyxl
	        wb = openpyxl.load_workbook(file_path)
	
	        # Iterate over all sheets and cells to replace '<NA>' with an empty string
	        for sheet in wb.worksheets:
	            for row in sheet.iter_rows():
	                for cell in row:
	                    if cell.value is not None and str(cell.value) == "<NA>":
	                        cell.value = ""
	
	        # Save the modified workbook to a BytesIO buffer
	        buffer = BytesIO()
	        wb.save(buffer)
	        buffer.seek(0)  # Reset buffer position to the beginning
	
	        file_data = buffer.read()
	
	        # Provide a download button for the modified OPD.xlsx file
	        st.download_button(
	            label="Download OPD.xlsx",
	            data=file_data,
	            file_name="OPD.xlsx",
	            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
	        )
	

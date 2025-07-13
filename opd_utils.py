import streamlit as st
import csv
import datetime
import pandas as pd
import numpy as np
import datetime
from datetime import timedelta
import xlsxwriter
import openpyxl
from openpyxl import Workbook
import io
from io import BytesIO
from io import StringIO
import os
import time 
import random
from openpyxl.styles import Font, Alignment

def generate_excel_file(start_date, title, custom_text, file_name, names):
    """
    Generates an Excel file where each week's structure aligns properly.

    Args:
        start_date (datetime): The starting date provided by the user.
        title (str): The text to be placed in cell A1.
        custom_text (str): The text to be placed in cell A2.
        file_name (str): The name of the output file.
        names (list): A list of names to be placed in the file.

    Returns:
        str: Path of the saved file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Place specific text in designated cells
    ws["A1"] = title
    ws["A2"] = custom_text

    # Columns where "custom_value" should be placed
    custom_value_columns = ["A", "C", "E", "G", "I", "K", "M"]
    name_columns = ["B", "D", "F", "H", "J", "L", "N"]

    # Ensure names list has at least one name
    if not names:
        names = ["Default Name ~"]

    # Days of the week to be placed across the row
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

    # Initial row where the first week starts
    start_row = 4
    num_weeks = 5  # Define the number of weeks
    week_height = 13  # Number of rows per week (date row + names + custom_value rows)

    for week in range(num_weeks):  
        current_date = start_date + datetime.timedelta(weeks=week)

        # Place day names and corresponding dates
        for i, day in enumerate(days):
            col_letter = chr(65 + (i * 2))  # Convert to Excel column letters (A, C, E, G, I, K, M)
            ws[f"{col_letter}{start_row}"] = day  
            formatted_date = (current_date + datetime.timedelta(days=i)).strftime("%B %-d, %Y")
            ws[f"{col_letter}{start_row + 1}"] = formatted_date  

        # Start placing names **immediately after the date row**
        names_start_row = start_row + 2  
        names_end_row = names_start_row + len(names)

        for i, col in enumerate(name_columns):
            custom_col = custom_value_columns[i]  # Get the column to the left
            for j, name in enumerate(names):
                row = names_start_row + j
                ws[f"{col}{row}"] = name  # Place the name
                ws[f"{custom_col}{row}"] = "custom_value"  # Place "custom_value" in the left column

        # Fill remaining rows with "custom_value" from the **date row** to the **next week's date row**
        next_week_start = start_row + week_height  # Set end range dynamically
        for i, col in enumerate(custom_value_columns):
            for row in range(start_row + 1, next_week_start):  # Fill from date row up to next week's start row
                if row >= names_end_row:  # Avoid overwriting names
                    ws[f"{col}{row}"] = "custom_value"

        # Move to the next week's section
        start_row = next_week_start  

    # Save the Excel file
    file_path = f"{file_name}"
    wb.save(file_path)

# Function to change page and trigger rerun
def navigate_to(page):
    st.session_state.page = page
    st.rerun()

def process_file(file_key, clinic_name, replacements=None, df=None):
    """Process a file (either uploaded or generated) and return a cleaned DataFrame."""
    
    # 1️⃣ **Use the provided DataFrame if already passed**
    if df is not None:
        print(f"Processing provided DataFrame for {clinic_name}...")
    
    else:
        # 2️⃣ **Check if the locally generated file exists**
        local_file_path = f"{file_key}"
        if os.path.exists(local_file_path):
            print(f"Found locally generated file: {local_file_path}. Using it for {clinic_name}...")
            df = pd.read_excel(local_file_path, dtype=str)
        
        # 3️⃣ **Otherwise, fall back to uploaded file**
        elif file_key in uploaded_files:
            print(f"Using uploaded file for {clinic_name}...")
            df = pd.read_excel(uploaded_files[file_key], dtype=str)
        
        else:
            print(f"❌ ERROR: No file found for {clinic_name} ({file_key}). Skipping...")
            return None  # Handle missing file case
    
    # ✅ **Continue normal processing**
    df.rename(columns={col: str(i) for i, col in enumerate(df.columns)}, inplace=True)

    D_dict = {}
    for i in range(28):
        col_idx = column_pairs[i % len(column_pairs)]
        start_day = days[i]
        end_day = days[i + 7]

        start_idx = df.loc[df[str(col_idx[0])] == start_day].index[0]
        end_idx = df.loc[df[str(col_idx[0])] == end_day].index[0]

        extracted_data = df.iloc[start_idx + 1:end_idx, list(col_idx)].copy()
        extracted_data.columns = ['type', 'provider']
        extracted_data.insert(0, 'date', start_day)
        extracted_data = extracted_data[:-1]

        D_dict[f"D{i}"] = extracted_data

    dfx = pd.concat(D_dict.values(), ignore_index=True)
    dfx['clinic'] = clinic_name

    # ✅ **Apply replacements if provided**
    if replacements:
        dfx = dfx.replace(replacements, regex=True)

    # ✅ **Save the cleaned file**
    filename = f"{clinic_name.lower()}.csv"
    dfx.to_csv(filename, index=False)

    print(f"✅ Processed {clinic_name} and saved to {filename}")
    return dfx  # Return DataFrame for further processing

import streamlit as st
import csv
import datetime
import pandas as pd
from datetime import timedelta

if 'page' not in st.session_state:
    st.session_state.page = "Home"  # Starting page
if 'start_date' not in st.session_state:
    st.session_state.start_date = None
if 'uploaded_files' not in st.session_state:
    st.session_state.uploaded_files = {}
if "uploaded_book4_file" not in st.session_state:
    st.session_state.uploaded_book4_file = {}

if st.session_state.page == "Home":
    st.title("Welcome to OPD Creator")
    st.write("Please choose what you'd like to do next.")
    
    # Button to navigate to 'Create OPD' page
    if st.button("Go to Create OPD"):
        st.session_state.page = "Create OPD"  # Set the page to 'Create OPD'
        st.rerun()  # Force a rerun to update the page

    # Button to navigate to 'Create Student Schedule' page
    if st.button("Go to Create Student Schedule"):
        st.session_state.page = "Create Student Schedule"  # Set the page to 'Create Student Schedule'
        st.rerun()  # Force a rerun to update the page

# Date input page
elif st.session_state.page == "Create OPD":
    st.title('Date Input for OPD')

    # Display instructions to the user
    st.write('Enter start date in m/d/yyyy format, no zeros in month or date (e.g., 7/6/2021):')

    # Create a text input field for the date
    date_input = st.text_input('Start Date')

    # Add a button to trigger the date parsing
    if st.button('Submit Date'):
        if date_input:  # Check if a date was entered
            try:
                # Try to parse the date entered by the user
                test_date = datetime.datetime.strptime(date_input, "%m/%d/%Y")
                st.session_state.start_date = test_date   # Save date in session state
                st.write(f"Valid date entered: {test_date.strftime('%m/%d/%Y')}")
                # After valid date input, move to the next page (Upload Files)
                st.session_state.page = "Upload Files"
                st.rerun()  # Force a rerun to reflect the page change
            except ValueError:
                st.error('Invalid date format. Please enter the date in m/d/yyyy format.')
        else:
            st.error('Please enter a date.')

# File upload page
elif st.session_state.page == "Upload Files":
    st.title("File Upload Section")
    st.write('Upload the following Excel files:')
    
    # Allow multiple file uploads at once
    uploaded_files = st.file_uploader("Choose your files", type="xlsx", accept_multiple_files=True)
    
    # Initialize the dictionary for storing files in session state
    uploaded_files_dict = {}
    
    if uploaded_files:
        # Iterate through the uploaded files and assign them based on their names
        for file in uploaded_files:
            if 'HOPE_DRIVE.xlsx' in file.name:
                uploaded_files_dict['HOPE_DRIVE.xlsx'] = file
            elif 'ETOWN.xlsx' in file.name:
                uploaded_files_dict['ETOWN.xlsx'] = file
            elif 'NYES.xlsx' in file.name:
                uploaded_files_dict['NYES.xlsx'] = file
        
        # Store the uploaded files in session state
        st.session_state.uploaded_files = uploaded_files_dict

    # Check if all files are uploaded
    if all(key in uploaded_files_dict for key in ['HOPE_DRIVE.xlsx', 'ETOWN.xlsx', 'NYES.xlsx']):
        st.write("All files uploaded successfully!")
        st.session_state.page = "OPD Creator"  # Move to next page after uploading
        st.rerun()  # Force a rerun to reflect the page change
    else:
        st.write("Please upload all required files.")

elif st.session_state.page == "OPD Creator":
	#test_date = datetime.datetime.strptime(x, "%m/%d/%Y")
	test_date = st.session_state.start_date
	uploaded_files = st.session_state.uploaded_files
	
	# initializing K
	K = 28
	 
	res = []
	 
	for day in range(K):
	    date = (test_date + datetime.timedelta(days = day)).strftime("%-m/%-d/%Y")
	    res.append(date)
	     
	#res
	
	dates = pd.DataFrame(res, columns =['dates'])
	
	dates['x'] = "y"
	
	dates['i'] = dates.index+1
	
	dates['x'] = dates['x'].astype(str)+dates['i'].astype(str)
	
	dates['x'] = dates['x'].astype(str) + "=" + "'"+dates['dates'].astype(str) + "'"
	
	dates = dates[['x']]
	
	dates.to_csv('dates.csv',index=False)
	
	import numpy as np
	
	datesdf = pd.read_csv('dates.csv')
	
	dates = datesdf['x'].astype(str)
	
	numpy_array=dates.to_numpy()
	np.savetxt("dates.py",numpy_array, fmt="%s")
	
	exec(open('dates.py').read())
	
	import xlsxwriter
	
	workbook = xlsxwriter.Workbook('OPD.xlsx')
	worksheet = workbook.add_worksheet('HOPE_DRIVE')
	worksheet2 = workbook.add_worksheet('ETOWN')
	worksheet3 = workbook.add_worksheet('NYES')
	worksheet4 = workbook.add_worksheet('EXTRA')
	worksheet5 = workbook.add_worksheet('MHS')
	
	format1 = workbook.add_format({'font_size':18,'bold': 1,'align': 'center','valign': 'vcenter','color':'black','bg_color':'#FEFFCC','border':1})
	    
	worksheet.write(0, 0, 'Site:',format1)
	worksheet.write(0, 1, 'Hope Drive',format1)
	                
	worksheet2.write(0, 0, 'Site:',format1)
	worksheet2.write(0, 1, 'Elizabethtown',format1)
	
	worksheet3.write(0, 0, 'Site:',format1)
	worksheet3.write(0, 1, 'Nyes Road',format1)
	
	worksheet4.write(0, 0, 'Site:',format1)
	worksheet4.write(0, 1, 'EXTRA',format1)
	
	worksheet5.write(0, 0, 'Site:',format1)
	worksheet5.write(0, 1, 'MHS',format1)
	
	#Color Coding
	format4 = workbook.add_format({'font_size':12,'bold': 1,'align': 'center','valign': 'vcenter','color':'black','bg_color':'#8ccf6f','border':1})
	format4a = workbook.add_format({'font_size':12,'bold': 1,'align': 'center','valign': 'vcenter','color':'black','bg_color':'#9fc5e8','border':1})    
	format5 = workbook.add_format({'font_size':12,'bold': 1,'align': 'center','valign': 'vcenter','color':'black','bg_color':'#FEFFCC','border':1})
	format5a = workbook.add_format({'font_size':12,'bold': 1,'align': 'center','valign': 'vcenter','color':'black','bg_color':'#d0e9ff','border':1})
	format11 = workbook.add_format({'font_size':18,'bold': 1,'align': 'center','valign': 'vcenter','color':'black','bg_color':'#FEFFCC','border':1})
	#H codes
	formate = workbook.add_format({'font_size':12,'bold': 0,'align': 'center','valign': 'vcenter','color':'white','border':0})
	
	# HOPE_DRIVE COLOR CODING AND IDENTIFYING ACUTE VERSUS CONTINUITY
	ranges_format1 = ['A8:H15', 'A32:H39', 'A56:H63', 'A80:H87']
	ranges_format5a = ['A18:H25', 'A42:H49', 'A66:H73', 'A90:H97']
	
	for cell_range in ranges_format1:
	    worksheet.conditional_format(cell_range, {'type': 'cell', 'criteria': '>=', 'value': 0, 'format': format1})
	
	for cell_range in ranges_format5a:
	    worksheet.conditional_format(cell_range, {'type': 'cell', 'criteria': '>=', 'value': 0, 'format': format5a})
	
	# HOPE_DRIVE CONDITIONAL FORMATTING
	ranges_format4 = ['A6:H6', 'A7:H7', 'A30:H30', 'A31:H31', 'A54:H54', 'A55:H55', 'A78:H78', 'A79:H79']
	ranges_format4a = ['A16:H16', 'A17:H17', 'A40:H40', 'A41:H41', 'A64:H64', 'A65:H65', 'A88:H88', 'A89:H89']
	
	for cell_range in ranges_format4:
	    worksheet.conditional_format(cell_range, {'type': 'cell', 'criteria': '>=', 'value': 0, 'format': format4})
	
	for cell_range in ranges_format4a:
	    worksheet.conditional_format(cell_range, {'type': 'cell', 'criteria': '>=', 'value': 0, 'format': format4a})
	
	# HOPE_DRIVE WRITING ACUTE AND CONTINUITY LABELS
	acute_format_ranges = [
	    (6, 7, 'AM - ACUTES', format4), (16, 17, 'PM - ACUTES', format4a), 
	    (30, 31, 'AM - ACUTES', format4), (40, 41, 'PM - ACUTES', format4a),
	    (54, 55, 'AM - ACUTES', format4), (64, 65, 'PM - ACUTES', format4a),
	    (78, 79, 'AM - ACUTES', format4), (88, 89, 'PM - ACUTES', format4a)
	]
	
	continuity_format_ranges = [
	    (8, 15, 'AM - Continuity', format5a), (18, 25, 'PM - Continuity', format5a),
	    (32, 39, 'AM - Continuity', format5a), (42, 49, 'PM - Continuity', format5a),
	    (56, 63, 'AM - Continuity', format5a), (66, 73, 'PM - Continuity', format5a),
	    (80, 87, 'AM - Continuity', format5a), (90, 97, 'PM - Continuity', format5a)
	]
	
	# Write Acute Labels
	for start_row, end_row, label, fmt in acute_format_ranges:
	    for row in range(start_row, end_row + 1):
	        worksheet.write(f'A{row}', label, fmt)
	
	# Write Continuity Labels
	for start_row, end_row, label, fmt in continuity_format_ranges:
	    for row in range(start_row, end_row + 1):
	        worksheet.write(f'A{row}', label, fmt)
	
	# Define the labels
	#labels = ['HX1', 'HX2', 'H2', 'H3', 'H4', 'H5', 'H6', 'H7', 'H8', 'H9', 'HXX1', 'HXX2', 'H12', 'H13', 'H14', 'H15', 'H16', 'H17', 'H18', 'H19']
	
	labels = ['H{}'.format(i) for i in range(20)]
	
	# Define the starting rows for each group
	start_rows = [6, 30, 54, 78]
	
	# Write the labels in each group
	for start_row in start_rows:
	    for i, label in enumerate(labels):
	        worksheet.write(f'I{start_row + i}', label, formate)
	
	# Simplify common formatting and label assignment for worksheets 2, 3, 4, 5
	worksheets = [worksheet2, worksheet3, worksheet4, worksheet5]
	
	ranges_format1 = ['A6:H15', 'A30:H39', 'A54:H63', 'A78:H87']
	ranges_format5a = ['A16:H25', 'A40:H49', 'A64:H73', 'A88:H97']
	specific_format_ranges = [
	    ('B6:H6', format4), ('B16:H16', format4a),
	    ('B30:H30', format4), ('B40:H40', format4a),
	    ('B54:H54', format4), ('B64:H64', format4a),
	    ('B78:H78', format4), ('B88:H88', format4a)
	]
	
	am_pm_labels = ['AM'] * 10 + ['PM'] * 10
	h_labels = ['H{}'.format(i) for i in range(20)]
	
	for worksheet in worksheets:
	    # Apply conditional formatting
	    for cell_range in ranges_format1:
	        worksheet.conditional_format(cell_range, {'type': 'cell', 'criteria': '>=', 'value': 0, 'format': format1})
	
	    for cell_range in ranges_format5a:
	        worksheet.conditional_format(cell_range, {'type': 'cell', 'criteria': '>=', 'value': 0, 'format': format5a})
	
	    for cell_range, fmt in specific_format_ranges:
	        worksheet.conditional_format(cell_range, {'type': 'cell', 'criteria': '>=', 'value': 0, 'format': fmt})
	
	    # Write AM/PM labels
	    sections = [(6, 25), (30, 49), (54, 73), (78, 97)]
	    for start_row, end_row in sections:
	        for i, label in enumerate(am_pm_labels):
	            worksheet.write(f'A{start_row + i}', label, format5a)
	
	    # Write H labels in column 'I'
	    start_rows = [6, 30, 54, 78]
	    for start_row in start_rows:
	        for i, label in enumerate(h_labels):
	            worksheet.write(f'I{start_row + i}', label, formate)
	
	# Loop through each worksheet in workbook
	for worksheet in workbook.worksheets():
	
	    # Set Zoom for all sheets
	    worksheet.set_zoom(80)
	
	    # Set Days
	    format3 = workbook.add_format({
	        'font_size': 12, 'bold': 1, 'align': 'center', 'valign': 'vcenter',
	        'color': 'black', 'bg_color': '#FFC7CE', 'border': 1
	    })
	    day_labels = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
	    start_rows = [2, 26, 50, 74] #[3, 27, 51, 75]
	    for start_row in start_rows:
	        for i, day in enumerate(day_labels):
	            worksheet.write(start_row, 1 + i, day, format3)  # B=1, C=2, etc.
	
	    # Set Date Formats
	    format_date = workbook.add_format({
	        'num_format': 'm/d/yyyy', 'font_size': 12, 'bold': 1, 'align': 'center', 'valign': 'vcenter',
	        'color': 'black', 'bg_color': '#FFC7CE', 'border': 1
	    })
	
	    format_label = workbook.add_format({
	        'font_size': 12, 'bold': 1, 'align': 'center', 'valign': 'vcenter',
	        'color': 'black', 'bg_color': '#FFC7CE', 'border': 1
	    })
	
	    # Set Date Formulas
	    date_rows = [3, 27, 51, 75] #[4, 28, 52, 76]
	    for i, start_row in enumerate(date_rows):
	        worksheet.write(f'A{start_row - 1}', "", format_label)
	        worksheet.write_formula(f'A{start_row}', f'="Week of:"&" "&TEXT(B{start_row},"m/d/yy")', format_label)
	        worksheet.write(f'A{start_row + 1}', "", format_label)
	
	    # Set Pink Bars (Conditional Format)
	    pink_bar_rows = [5, 29, 53, 77]
	    for row in pink_bar_rows:
	        worksheet.conditional_format(f'A{row}:H{row}', {
	            'type': 'cell', 'criteria': '>=', 'value': 0, 'format': format_label
	        })
	
	    # Black Bars
	    format2 = workbook.add_format({'bg_color': 'black'})
	    black_bar_rows = [2, 26, 50, 74, 98]
	    for row in black_bar_rows:
	        worksheet.merge_range(f'A{row}:H{row}', " ", format2)
	        
	    # Write More Dates
	    date_values = [
	        [y1, y2, y3, y4, y5, y6, y7],
	        [y8, y9, y10, y11, y12, y13, y14],
	        [y15, y16, y17, y18, y19, y20, y21],
	        [y22, y23, y24, y25, y26, y27, y28]
	    ]
	    for i, start_row in enumerate(date_rows):
	        for j, value in enumerate(date_values[i]):
	            worksheet.write(start_row, 1 + j, value, format_date)  # B=1, C=2, etc.
	
	    # Set Column Widths
	    worksheet.set_column('A:A', 22)
	    worksheet.set_column('B:H', 40)
	    worksheet.set_row(0, 37.25)
	
	    # Merge Format for Text
	    merge_format = workbook.add_format({
	        'bold': 1, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True,
	        'color': 'red', 'bg_color': '#FEFFCC', 'border': 1
	    })
	    text1 = 'Students are to alert their preceptors when they have a Clinical Reasoning Teaching Session (CRTS).  Please allow the students to leave approximately 15 minutes prior to the start of their session so they can be prepared to actively participate.  ~ Thank you!'
	
	    # Merge and Write Important Message
	    worksheet.merge_range('C1:F1', text1, merge_format)
	    worksheet.write('G1', "", merge_format)
	    worksheet.write('H1', "", merge_format)
	
	# Close Workbook
	workbook.close()
	
	####################################################################################################################################
	import pandas as pd
	import datetime
	from datetime import timedelta
	
	# Disable chained assignment warning
	pd.options.mode.chained_assignment = None
	
	# Parse the start date (assuming `test_date` is already defined)
	formatted_date = test_date.strftime('%m-%d-%Y')
	start_date = datetime.datetime.strptime(formatted_date, '%m-%d-%Y')
	
	# Calculate the end date (34 days from the start date)
	end_date = start_date + timedelta(days=34)
	
	# Generate the date range
	date_range = pd.date_range(start=start_date, end=end_date)
	
	# Create a DataFrame with the formatted dates
	xf201 = pd.DataFrame({'date': date_range})
	xf201['convert'] = xf201['date'].dt.strftime('%B %-d, %Y')
	
	xf201['t'] = "T"
	xf201['c'] = xf201.index+0
	xf201['T'] = xf201['t'].astype(str) + xf201 ['c'].astype(str)
	
	for i in range(35):
	    exec(f"day{i} = xf201['convert'][{i}]")
	    
	####################################hopedrive###################################################################
	if uploaded_files['HOPE_DRIVE.xlsx']:
	    df = pd.read_excel(uploaded_files['HOPE_DRIVE.xlsx'], dtype=str)
	    #st.write("HOPE_DRIVE Data:")
	    #st.dataframe(df)  # Show the dataframe
	  
	#df=pd.read_excel('HOPE_DRIVE.xlsx',dtype=str)
	
	df.rename(columns={ df.columns[0]: "0" }, inplace = True)
	df.rename(columns={ df.columns[1]: "1" }, inplace = True)
	df.rename(columns={ df.columns[2]: "2" }, inplace = True)
	df.rename(columns={ df.columns[3]: "3" }, inplace = True)
	df.rename(columns={ df.columns[4]: "4" }, inplace = True)
	df.rename(columns={ df.columns[5]: "5" }, inplace = True)
	df.rename(columns={ df.columns[6]: "6" }, inplace = True)
	df.rename(columns={ df.columns[7]: "7" }, inplace = True)
	df.rename(columns={ df.columns[8]: "8" }, inplace = True)
	df.rename(columns={ df.columns[9]: "9" }, inplace = True)
	df.rename(columns={ df.columns[10]: "10" }, inplace = True)
	df.rename(columns={ df.columns[11]: "11" }, inplace = True)
	df.rename(columns={ df.columns[12]: "12" }, inplace = True)
	df.rename(columns={ df.columns[13]: "13" }, inplace = True)
	
	xf300 = pd.DataFrame({'no':['0','2','4','6','8','10','12','0','2','4','6','8','10','12','0','2','4','6','8','10','12','0','2','4','6','8','10','12']})
	
	xf300['no1'] = ['1','3','5','7','9','11','13','1','3','5','7','9','11','13','1','3','5','7','9','11','13','1','3','5','7','9','11','13']
	
	xf300['start']=['0','1','2','3','4','5','6','7','8','9','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24','25','26','27']
	
	xf300['end'] = ['7','8','9','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24','25','26','27','28','29','30','31','32','33','34']
	
	a = df.loc[df['0'] == day0].index[0]
	b = df.loc[df['0'] == day7].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z0=c.loc[:, ('0','1')]
	z0['date']=day0
	z0.rename(columns={ df.columns[0]:"type"}, inplace = True)
	z0.rename(columns={ df.columns[1]:"provider"}, inplace = True)
	D0=z0[['date','type','provider']]
	D0=D0[:-1]
	
	a = df.loc[df['2'] == day1].index[0]
	b = df.loc[df['2'] == day8].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z1=c.loc[:, ('2','3')]
	z1['date']=day1
	z1.rename(columns={ df.columns[2]:"type"}, inplace = True)
	z1.rename(columns={ df.columns[3]:"provider"}, inplace = True)
	D1=z1[['date','type','provider']]
	D1=D1[:-1]
	
	a = df.loc[df['4'] == day2].index[0]
	b = df.loc[df['4'] == day9].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z2=c.loc[:, ('4','5')]
	z2['date']=day2
	z2.rename(columns={ df.columns[4]:"type"}, inplace = True)
	z2.rename(columns={ df.columns[5]:"provider"}, inplace = True)
	D2=z2[['date','type','provider']]
	D2=D2[:-1]
	
	a = df.loc[df['6'] == day3].index[0]
	b = df.loc[df['6'] == day10].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z3=c.loc[:, ('6','7')]
	z3['date']=day3
	z3.rename(columns={ df.columns[6]:"type"}, inplace = True)
	z3.rename(columns={ df.columns[7]:"provider"}, inplace = True)
	D3=z3[['date','type','provider']]
	D3=D3[:-1]
	
	a = df.loc[df['8'] == day4].index[0]
	b = df.loc[df['8'] == day11].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z4=c.loc[:, ('8','9')]
	z4['date']=day4
	z4.rename(columns={ df.columns[8]:"type"}, inplace = True)
	z4.rename(columns={ df.columns[9]:"provider"}, inplace = True)
	D4=z4[['date','type','provider']]
	D4=D4[:-1]
	
	a = df.loc[df['10'] == day5].index[0]
	b = df.loc[df['10'] == day12].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z5=c.loc[:, ('10','11')]
	z5['date']=day5
	z5.rename(columns={ df.columns[10]:"type"}, inplace = True)
	z5.rename(columns={ df.columns[11]:"provider"}, inplace = True)
	D5=z5[['date','type','provider']]
	D5=D5[:-1]
	
	a = df.loc[df['12'] == day6].index[0]
	b = df.loc[df['12'] == day13].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z6=c.loc[:, ('12','13')]
	z6['date']=day6
	z6.rename(columns={ df.columns[12]:"type"}, inplace = True)
	z6.rename(columns={ df.columns[13]:"provider"}, inplace = True)
	D6=z6[['date','type','provider']]
	D6=D6[:-1]
	
	a = df.loc[df['0'] == day7].index[0]
	b = df.loc[df['0'] == day14].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z7=c.loc[:, ('0','1')]
	z7['date']=day7
	z7.rename(columns={ df.columns[0]:"type"}, inplace = True)
	z7.rename(columns={ df.columns[1]:"provider"}, inplace = True)
	D7=z7[['date','type','provider']]
	D7=D7[:-1]
	
	a = df.loc[df['2'] == day8].index[0]
	b = df.loc[df['2'] == day15].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z8=c.loc[:, ('2','3')]
	z8['date']=day8
	z8.rename(columns={ df.columns[2]:"type"}, inplace = True)
	z8.rename(columns={ df.columns[3]:"provider"}, inplace = True)
	D8=z8[['date','type','provider']]
	D8=D8[:-1]
	
	a = df.loc[df['4'] == day9].index[0]
	b = df.loc[df['4'] == day16].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z9=c.loc[:, ('4','5')]
	z9['date']=day9
	z9.rename(columns={ df.columns[4]:"type"}, inplace = True)
	z9.rename(columns={ df.columns[5]:"provider"}, inplace = True)
	D9=z9[['date','type','provider']]
	D9=D9[:-1]
	
	a = df.loc[df['6'] == day10].index[0]
	b = df.loc[df['6'] == day17].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z10=c.loc[:, ('6','7')]
	z10['date']=day10
	z10.rename(columns={ df.columns[6]:"type"}, inplace = True)
	z10.rename(columns={ df.columns[7]:"provider"}, inplace = True)
	D10=z10[['date','type','provider']]
	D10=D10[:-1]
	
	a = df.loc[df['8'] == day11].index[0]
	b = df.loc[df['8'] == day18].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z11=c.loc[:, ('8','9')]
	z11['date']=day11
	z11.rename(columns={ df.columns[8]:"type"}, inplace = True)
	z11.rename(columns={ df.columns[9]:"provider"}, inplace = True)
	D11=z11[['date','type','provider']]
	D11=D11[:-1]
	
	a = df.loc[df['10'] == day12].index[0]
	b = df.loc[df['10'] == day19].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z12=c.loc[:, ('10','11')]
	z12['date']=day12
	z12.rename(columns={ df.columns[10]:"type"}, inplace = True)
	z12.rename(columns={ df.columns[11]:"provider"}, inplace = True)
	D12=z12[['date','type','provider']]
	D12=D12[:-1]
	
	a = df.loc[df['12'] == day13].index[0]
	b = df.loc[df['12'] == day20].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z13=c.loc[:, ('12','13')]
	z13['date']=day13
	z13.rename(columns={ df.columns[12]:"type"}, inplace = True)
	z13.rename(columns={ df.columns[13]:"provider"}, inplace = True)
	D13=z13[['date','type','provider']]
	D13=D13[:-1]
	
	a = df.loc[df['0'] == day14].index[0]
	b = df.loc[df['0'] == day21].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z14=c.loc[:, ('0','1')]
	z14['date']=day14
	z14.rename(columns={ df.columns[0]:"type"}, inplace = True)
	z14.rename(columns={ df.columns[1]:"provider"}, inplace = True)
	D14=z14[['date','type','provider']]
	D14=D14[:-1]
	
	a = df.loc[df['2'] == day15].index[0]
	b = df.loc[df['2'] == day22].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z15=c.loc[:, ('2','3')]
	z15['date']=day15
	z15.rename(columns={ df.columns[2]:"type"}, inplace = True)
	z15.rename(columns={ df.columns[3]:"provider"}, inplace = True)
	D15=z15[['date','type','provider']]
	D15=D15[:-1]
	
	a = df.loc[df['4'] == day16].index[0]
	b = df.loc[df['4'] == day23].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z16=c.loc[:, ('4','5')]
	z16['date']=day16
	z16.rename(columns={ df.columns[4]:"type"}, inplace = True)
	z16.rename(columns={ df.columns[5]:"provider"}, inplace = True)
	D16=z16[['date','type','provider']]
	D16=D16[:-1]
	
	a = df.loc[df['6'] == day17].index[0]
	b = df.loc[df['6'] == day24].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z17=c.loc[:, ('6','7')]
	z17['date']=day17
	z17.rename(columns={ df.columns[6]:"type"}, inplace = True)
	z17.rename(columns={ df.columns[7]:"provider"}, inplace = True)
	D17=z17[['date','type','provider']]
	D17=D17[:-1]
	
	a = df.loc[df['8'] == day18].index[0]
	b = df.loc[df['8'] == day25].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z18=c.loc[:, ('8','9')]
	z18['date']=day18
	z18.rename(columns={ df.columns[8]:"type"}, inplace = True)
	z18.rename(columns={ df.columns[9]:"provider"}, inplace = True)
	D18=z18[['date','type','provider']]
	D18=D18[:-1]
	
	a = df.loc[df['10'] == day19].index[0]
	b = df.loc[df['10'] == day26].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z19=c.loc[:, ('10','11')]
	z19['date']=day19
	z19.rename(columns={ df.columns[10]:"type"}, inplace = True)
	z19.rename(columns={ df.columns[11]:"provider"}, inplace = True)
	D19=z19[['date','type','provider']]
	D19=D19[:-1]
	
	a = df.loc[df['12'] == day20].index[0]
	b = df.loc[df['12'] == day27].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z20=c.loc[:, ('12','13')]
	z20['date']=day20
	z20.rename(columns={ df.columns[12]:"type"}, inplace = True)
	z20.rename(columns={ df.columns[13]:"provider"}, inplace = True)
	D20=z20[['date','type','provider']]
	D20=D20[:-1]
	
	a = df.loc[df['0'] == day21].index[0]
	b = df.loc[df['0'] == day28].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z21=c.loc[:, ('0','1')]
	z21['date']=day21
	z21.rename(columns={ df.columns[0]:"type"}, inplace = True)
	z21.rename(columns={ df.columns[1]:"provider"}, inplace = True)
	D21=z21[['date','type','provider']]
	D21=D21[:-1]
	
	a = df.loc[df['2'] == day22].index[0]
	b = df.loc[df['2'] == day29].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z22=c.loc[:, ('2','3')]
	z22['date']=day22
	z22.rename(columns={ df.columns[2]:"type"}, inplace = True)
	z22.rename(columns={ df.columns[3]:"provider"}, inplace = True)
	D22=z22[['date','type','provider']]
	D22=D22[:-1]
	
	a = df.loc[df['4'] == day23].index[0]
	b = df.loc[df['4'] == day30].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z23=c.loc[:, ('4','5')]
	z23['date']=day23
	z23.rename(columns={ df.columns[4]:"type"}, inplace = True)
	z23.rename(columns={ df.columns[5]:"provider"}, inplace = True)
	D23=z23[['date','type','provider']]
	D23=D23[:-1]
	
	a = df.loc[df['6'] == day24].index[0]
	b = df.loc[df['6'] == day31].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z24=c.loc[:, ('6','7')]
	z24['date']=day24
	z24.rename(columns={ df.columns[6]:"type"}, inplace = True)
	z24.rename(columns={ df.columns[7]:"provider"}, inplace = True)
	D24=z24[['date','type','provider']]
	D24=D24[:-1]
	
	a = df.loc[df['8'] == day25].index[0]
	b = df.loc[df['8'] == day32].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z25=c.loc[:, ('8','9')]
	z25['date']=day25
	z25.rename(columns={ df.columns[8]:"type"}, inplace = True)
	z25.rename(columns={ df.columns[9]:"provider"}, inplace = True)
	D25=z25[['date','type','provider']]
	D25=D25[:-1]
	
	a = df.loc[df['10'] == day26].index[0]
	b = df.loc[df['10'] == day33].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z26=c.loc[:, ('10','11')]
	z26['date']=day26
	z26.rename(columns={ df.columns[10]:"type"}, inplace = True)
	z26.rename(columns={ df.columns[11]:"provider"}, inplace = True)
	D26=z26[['date','type','provider']]
	D26=D26[:-1]
	
	a = df.loc[df['12'] == day27].index[0]
	b = df.loc[df['12'] == day34].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z27=c.loc[:, ('12','13')]
	z27['date']=day27
	z27.rename(columns={ df.columns[12]:"type"}, inplace = True)
	z27.rename(columns={ df.columns[13]:"provider"}, inplace = True)
	D27=z27[['date','type','provider']]
	D27=D27[:-1]
	
	dfx=pd.DataFrame(columns=D0.columns)
	
	dfx=pd.concat([dfx,D0, D1, D2, D3, D4, D5, D6, D7, D8, D9, D10, D11, D12, D13, D14, D15, D16, D17, D18, D19, D20, D21, D22, D23, D24, D25, D26, D27])
	
	dfx['clinic'] = "HOPE_DRIVE"
	
	dfx.to_csv('hope.csv',index=False)
	hope=dfx.replace("Hope Drive AM Continuity", "AM - Continuity", regex=True)
	hope=hope.replace("Hope Drive PM Continuity", "PM - Continuity", regex=True)
	hope=hope.replace("Hope Drive\xa0AM Acute Precept ", "AM - ACUTES", regex=True)
	hope=hope.replace("Hope Drive PM Acute Precept", "PM - ACUTES", regex=True)
	hope=hope.replace("Hope Drive Weekend Continuity", "AM - Continuity", regex=True)
	hope=hope.replace("Hope Drive Weekend Acute 1", "AM - ACUTES", regex=True)
	hope=hope.replace("Hope Drive Weekend Acute 2", "AM - ACUTES", regex=True)
	
	hope.to_csv('hope.csv',index=False)
	
	####################################################ETOWN#################################################
	if uploaded_files['ETOWN.xlsx']:
	    df = pd.read_excel(uploaded_files['ETOWN.xlsx'], dtype=str)
	    #st.write("ETOWN Data:")
	    #st.dataframe(df_etown)  # Show the dataframe
	  
	#df=pd.read_excel('ETOWN.xlsx',dtype=str)
	df.rename(columns={ df.columns[0]: "0" }, inplace = True)
	df.rename(columns={ df.columns[1]: "1" }, inplace = True)
	df.rename(columns={ df.columns[2]: "2" }, inplace = True)
	df.rename(columns={ df.columns[3]: "3" }, inplace = True)
	df.rename(columns={ df.columns[4]: "4" }, inplace = True)
	df.rename(columns={ df.columns[5]: "5" }, inplace = True)
	df.rename(columns={ df.columns[6]: "6" }, inplace = True)
	df.rename(columns={ df.columns[7]: "7" }, inplace = True)
	df.rename(columns={ df.columns[8]: "8" }, inplace = True)
	df.rename(columns={ df.columns[9]: "9" }, inplace = True)
	df.rename(columns={ df.columns[10]: "10" }, inplace = True)
	df.rename(columns={ df.columns[11]: "11" }, inplace = True)
	df.rename(columns={ df.columns[12]: "12" }, inplace = True)
	df.rename(columns={ df.columns[13]: "13" }, inplace = True)
	
	xf300 = pd.DataFrame({'no':['0','2','4','6','8','10','12','0','2','4','6','8','10','12','0','2','4','6','8','10','12','0','2','4','6','8','10','12']})
	
	xf300['no1'] = ['1','3','5','7','9','11','13','1','3','5','7','9','11','13','1','3','5','7','9','11','13','1','3','5','7','9','11','13']
	
	xf300['start']=['0','1','2','3','4','5','6','7','8','9','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24','25','26','27']
	
	xf300['end'] = ['7','8','9','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24','25','26','27','28','29','30','31','32','33','34']
	
	a = df.loc[df['0'] == day0].index[0]
	b = df.loc[df['0'] == day7].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z0=c.loc[:, ('0','1')]
	z0['date']=day0
	z0.rename(columns={ df.columns[0]:"type"}, inplace = True)
	z0.rename(columns={ df.columns[1]:"provider"}, inplace = True)
	D0=z0[['date','type','provider']]
	D0=D0[:-1]
	
	a = df.loc[df['2'] == day1].index[0]
	b = df.loc[df['2'] == day8].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z1=c.loc[:, ('2','3')]
	z1['date']=day1
	z1.rename(columns={ df.columns[2]:"type"}, inplace = True)
	z1.rename(columns={ df.columns[3]:"provider"}, inplace = True)
	D1=z1[['date','type','provider']]
	D1=D1[:-1]
	
	a = df.loc[df['4'] == day2].index[0]
	b = df.loc[df['4'] == day9].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z2=c.loc[:, ('4','5')]
	z2['date']=day2
	z2.rename(columns={ df.columns[4]:"type"}, inplace = True)
	z2.rename(columns={ df.columns[5]:"provider"}, inplace = True)
	D2=z2[['date','type','provider']]
	D2=D2[:-1]
	
	a = df.loc[df['6'] == day3].index[0]
	b = df.loc[df['6'] == day10].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z3=c.loc[:, ('6','7')]
	z3['date']=day3
	z3.rename(columns={ df.columns[6]:"type"}, inplace = True)
	z3.rename(columns={ df.columns[7]:"provider"}, inplace = True)
	D3=z3[['date','type','provider']]
	D3=D3[:-1]
	
	a = df.loc[df['8'] == day4].index[0]
	b = df.loc[df['8'] == day11].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z4=c.loc[:, ('8','9')]
	z4['date']=day4
	z4.rename(columns={ df.columns[8]:"type"}, inplace = True)
	z4.rename(columns={ df.columns[9]:"provider"}, inplace = True)
	D4=z4[['date','type','provider']]
	D4=D4[:-1]
	
	a = df.loc[df['10'] == day5].index[0]
	b = df.loc[df['10'] == day12].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z5=c.loc[:, ('10','11')]
	z5['date']=day5
	z5.rename(columns={ df.columns[10]:"type"}, inplace = True)
	z5.rename(columns={ df.columns[11]:"provider"}, inplace = True)
	D5=z5[['date','type','provider']]
	D5=D5[:-1]
	
	a = df.loc[df['12'] == day6].index[0]
	b = df.loc[df['12'] == day13].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z6=c.loc[:, ('12','13')]
	z6['date']=day6
	z6.rename(columns={ df.columns[12]:"type"}, inplace = True)
	z6.rename(columns={ df.columns[13]:"provider"}, inplace = True)
	D6=z6[['date','type','provider']]
	D6=D6[:-1]
	
	a = df.loc[df['0'] == day7].index[0]
	b = df.loc[df['0'] == day14].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z7=c.loc[:, ('0','1')]
	z7['date']=day7
	z7.rename(columns={ df.columns[0]:"type"}, inplace = True)
	z7.rename(columns={ df.columns[1]:"provider"}, inplace = True)
	D7=z7[['date','type','provider']]
	D7=D7[:-1]
	
	a = df.loc[df['2'] == day8].index[0]
	b = df.loc[df['2'] == day15].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z8=c.loc[:, ('2','3')]
	z8['date']=day8
	z8.rename(columns={ df.columns[2]:"type"}, inplace = True)
	z8.rename(columns={ df.columns[3]:"provider"}, inplace = True)
	D8=z8[['date','type','provider']]
	D8=D8[:-1]
	
	a = df.loc[df['4'] == day9].index[0]
	b = df.loc[df['4'] == day16].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z9=c.loc[:, ('4','5')]
	z9['date']=day9
	z9.rename(columns={ df.columns[4]:"type"}, inplace = True)
	z9.rename(columns={ df.columns[5]:"provider"}, inplace = True)
	D9=z9[['date','type','provider']]
	D9=D9[:-1]
	
	a = df.loc[df['6'] == day10].index[0]
	b = df.loc[df['6'] == day17].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z10=c.loc[:, ('6','7')]
	z10['date']=day10
	z10.rename(columns={ df.columns[6]:"type"}, inplace = True)
	z10.rename(columns={ df.columns[7]:"provider"}, inplace = True)
	D10=z10[['date','type','provider']]
	D10=D10[:-1]
	
	a = df.loc[df['8'] == day11].index[0]
	b = df.loc[df['8'] == day18].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z11=c.loc[:, ('8','9')]
	z11['date']=day11
	z11.rename(columns={ df.columns[8]:"type"}, inplace = True)
	z11.rename(columns={ df.columns[9]:"provider"}, inplace = True)
	D11=z11[['date','type','provider']]
	D11=D11[:-1]
	
	a = df.loc[df['10'] == day12].index[0]
	b = df.loc[df['10'] == day19].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z12=c.loc[:, ('10','11')]
	z12['date']=day12
	z12.rename(columns={ df.columns[10]:"type"}, inplace = True)
	z12.rename(columns={ df.columns[11]:"provider"}, inplace = True)
	D12=z12[['date','type','provider']]
	D12=D12[:-1]
	
	a = df.loc[df['12'] == day13].index[0]
	b = df.loc[df['12'] == day20].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z13=c.loc[:, ('12','13')]
	z13['date']=day13
	z13.rename(columns={ df.columns[12]:"type"}, inplace = True)
	z13.rename(columns={ df.columns[13]:"provider"}, inplace = True)
	D13=z13[['date','type','provider']]
	D13=D13[:-1]
	
	a = df.loc[df['0'] == day14].index[0]
	b = df.loc[df['0'] == day21].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z14=c.loc[:, ('0','1')]
	z14['date']=day14
	z14.rename(columns={ df.columns[0]:"type"}, inplace = True)
	z14.rename(columns={ df.columns[1]:"provider"}, inplace = True)
	D14=z14[['date','type','provider']]
	D14=D14[:-1]
	
	a = df.loc[df['2'] == day15].index[0]
	b = df.loc[df['2'] == day22].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z15=c.loc[:, ('2','3')]
	z15['date']=day15
	z15.rename(columns={ df.columns[2]:"type"}, inplace = True)
	z15.rename(columns={ df.columns[3]:"provider"}, inplace = True)
	D15=z15[['date','type','provider']]
	D15=D15[:-1]
	
	a = df.loc[df['4'] == day16].index[0]
	b = df.loc[df['4'] == day23].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z16=c.loc[:, ('4','5')]
	z16['date']=day16
	z16.rename(columns={ df.columns[4]:"type"}, inplace = True)
	z16.rename(columns={ df.columns[5]:"provider"}, inplace = True)
	D16=z16[['date','type','provider']]
	D16=D16[:-1]
	
	a = df.loc[df['6'] == day17].index[0]
	b = df.loc[df['6'] == day24].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z17=c.loc[:, ('6','7')]
	z17['date']=day17
	z17.rename(columns={ df.columns[6]:"type"}, inplace = True)
	z17.rename(columns={ df.columns[7]:"provider"}, inplace = True)
	D17=z17[['date','type','provider']]
	D17=D17[:-1]
	
	a = df.loc[df['8'] == day18].index[0]
	b = df.loc[df['8'] == day25].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z18=c.loc[:, ('8','9')]
	z18['date']=day18
	z18.rename(columns={ df.columns[8]:"type"}, inplace = True)
	z18.rename(columns={ df.columns[9]:"provider"}, inplace = True)
	D18=z18[['date','type','provider']]
	D18=D18[:-1]
	
	a = df.loc[df['10'] == day19].index[0]
	b = df.loc[df['10'] == day26].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z19=c.loc[:, ('10','11')]
	z19['date']=day19
	z19.rename(columns={ df.columns[10]:"type"}, inplace = True)
	z19.rename(columns={ df.columns[11]:"provider"}, inplace = True)
	D19=z19[['date','type','provider']]
	D19=D19[:-1]
	
	a = df.loc[df['12'] == day20].index[0]
	b = df.loc[df['12'] == day27].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z20=c.loc[:, ('12','13')]
	z20['date']=day20
	z20.rename(columns={ df.columns[12]:"type"}, inplace = True)
	z20.rename(columns={ df.columns[13]:"provider"}, inplace = True)
	D20=z20[['date','type','provider']]
	D20=D20[:-1]
	
	a = df.loc[df['0'] == day21].index[0]
	b = df.loc[df['0'] == day28].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z21=c.loc[:, ('0','1')]
	z21['date']=day21
	z21.rename(columns={ df.columns[0]:"type"}, inplace = True)
	z21.rename(columns={ df.columns[1]:"provider"}, inplace = True)
	D21=z21[['date','type','provider']]
	D21=D21[:-1]
	
	a = df.loc[df['2'] == day22].index[0]
	b = df.loc[df['2'] == day29].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z22=c.loc[:, ('2','3')]
	z22['date']=day22
	z22.rename(columns={ df.columns[2]:"type"}, inplace = True)
	z22.rename(columns={ df.columns[3]:"provider"}, inplace = True)
	D22=z22[['date','type','provider']]
	D22=D22[:-1]
	
	a = df.loc[df['4'] == day23].index[0]
	b = df.loc[df['4'] == day30].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z23=c.loc[:, ('4','5')]
	z23['date']=day23
	z23.rename(columns={ df.columns[4]:"type"}, inplace = True)
	z23.rename(columns={ df.columns[5]:"provider"}, inplace = True)
	D23=z23[['date','type','provider']]
	D23=D23[:-1]
	
	a = df.loc[df['6'] == day24].index[0]
	b = df.loc[df['6'] == day31].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z24=c.loc[:, ('6','7')]
	z24['date']=day24
	z24.rename(columns={ df.columns[6]:"type"}, inplace = True)
	z24.rename(columns={ df.columns[7]:"provider"}, inplace = True)
	D24=z24[['date','type','provider']]
	D24=D24[:-1]
	
	a = df.loc[df['8'] == day25].index[0]
	b = df.loc[df['8'] == day32].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z25=c.loc[:, ('8','9')]
	z25['date']=day25
	z25.rename(columns={ df.columns[8]:"type"}, inplace = True)
	z25.rename(columns={ df.columns[9]:"provider"}, inplace = True)
	D25=z25[['date','type','provider']]
	D25=D25[:-1]
	
	a = df.loc[df['10'] == day26].index[0]
	b = df.loc[df['10'] == day33].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z26=c.loc[:, ('10','11')]
	z26['date']=day26
	z26.rename(columns={ df.columns[10]:"type"}, inplace = True)
	z26.rename(columns={ df.columns[11]:"provider"}, inplace = True)
	D26=z26[['date','type','provider']]
	D26=D26[:-1]
	
	a = df.loc[df['12'] == day27].index[0]
	b = df.loc[df['12'] == day34].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z27=c.loc[:, ('12','13')]
	z27['date']=day27
	z27.rename(columns={ df.columns[12]:"type"}, inplace = True)
	z27.rename(columns={ df.columns[13]:"provider"}, inplace = True)
	D27=z27[['date','type','provider']]
	D27=D27[:-1]
	
	
	dfx=pd.DataFrame(columns=D0.columns)
	
	dfx=pd.concat([dfx,D0, D1, D2, D3, D4, D5, D6, D7, D8, D9, D10, D11, D12, D13, D14, D15, D16, D17, D18, D19, D20, D21, D22, D23, D24, D25, D26, D27])
	
	dfx['clinic'] = "ETOWN"
	
	
	dfx.to_csv('etown.csv',index=False)
	ETOWN=dfx.replace("Etown AM Continuity", "AM - Continuity", regex=True)
	ETOWN=ETOWN.replace("Etown PM Continuity", "PM - Continuity", regex=True)
	ETOWN.to_csv('etown.csv',index=False)
	
	#############################################################NYES################################################
	if uploaded_files['NYES.xlsx']:
	    df = pd.read_excel(uploaded_files['NYES.xlsx'], dtype=str)
	    #st.write("NYES Data:")
	    #st.dataframe(df_nyes)
	#df=pd.read_excel('NYES.xlsx',dtype=str)
	
	df.rename(columns={ df.columns[0]: "0" }, inplace = True)
	df.rename(columns={ df.columns[1]: "1" }, inplace = True)
	df.rename(columns={ df.columns[2]: "2" }, inplace = True)
	df.rename(columns={ df.columns[3]: "3" }, inplace = True)
	df.rename(columns={ df.columns[4]: "4" }, inplace = True)
	df.rename(columns={ df.columns[5]: "5" }, inplace = True)
	df.rename(columns={ df.columns[6]: "6" }, inplace = True)
	df.rename(columns={ df.columns[7]: "7" }, inplace = True)
	df.rename(columns={ df.columns[8]: "8" }, inplace = True)
	df.rename(columns={ df.columns[9]: "9" }, inplace = True)
	df.rename(columns={ df.columns[10]: "10" }, inplace = True)
	df.rename(columns={ df.columns[11]: "11" }, inplace = True)
	df.rename(columns={ df.columns[12]: "12" }, inplace = True)
	df.rename(columns={ df.columns[13]: "13" }, inplace = True)
	
	xf300 = pd.DataFrame({'no':['0','2','4','6','8','10','12','0','2','4','6','8','10','12','0','2','4','6','8','10','12','0','2','4','6','8','10','12']})
	
	xf300['no1'] = ['1','3','5','7','9','11','13','1','3','5','7','9','11','13','1','3','5','7','9','11','13','1','3','5','7','9','11','13']
	
	xf300['start']=['0','1','2','3','4','5','6','7','8','9','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24','25','26','27']
	
	xf300['end'] = ['7','8','9','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24','25','26','27','28','29','30','31','32','33','34']
	
	a = df.loc[df['0'] == day0].index[0]
	b = df.loc[df['0'] == day7].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z0=c.loc[:, ('0','1')]
	z0['date']=day0
	z0.rename(columns={ df.columns[0]:"type"}, inplace = True)
	z0.rename(columns={ df.columns[1]:"provider"}, inplace = True)
	D0=z0[['date','type','provider']]
	D0=D0[:-1]
	
	a = df.loc[df['2'] == day1].index[0]
	b = df.loc[df['2'] == day8].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z1=c.loc[:, ('2','3')]
	z1['date']=day1
	z1.rename(columns={ df.columns[2]:"type"}, inplace = True)
	z1.rename(columns={ df.columns[3]:"provider"}, inplace = True)
	D1=z1[['date','type','provider']]
	D1=D1[:-1]
	
	a = df.loc[df['4'] == day2].index[0]
	b = df.loc[df['4'] == day9].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z2=c.loc[:, ('4','5')]
	z2['date']=day2
	z2.rename(columns={ df.columns[4]:"type"}, inplace = True)
	z2.rename(columns={ df.columns[5]:"provider"}, inplace = True)
	D2=z2[['date','type','provider']]
	D2=D2[:-1]
	
	a = df.loc[df['6'] == day3].index[0]
	b = df.loc[df['6'] == day10].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z3=c.loc[:, ('6','7')]
	z3['date']=day3
	z3.rename(columns={ df.columns[6]:"type"}, inplace = True)
	z3.rename(columns={ df.columns[7]:"provider"}, inplace = True)
	D3=z3[['date','type','provider']]
	D3=D3[:-1]
	
	a = df.loc[df['8'] == day4].index[0]
	b = df.loc[df['8'] == day11].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z4=c.loc[:, ('8','9')]
	z4['date']=day4
	z4.rename(columns={ df.columns[8]:"type"}, inplace = True)
	z4.rename(columns={ df.columns[9]:"provider"}, inplace = True)
	D4=z4[['date','type','provider']]
	D4=D4[:-1]
	
	a = df.loc[df['10'] == day5].index[0]
	b = df.loc[df['10'] == day12].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z5=c.loc[:, ('10','11')]
	z5['date']=day5
	z5.rename(columns={ df.columns[10]:"type"}, inplace = True)
	z5.rename(columns={ df.columns[11]:"provider"}, inplace = True)
	D5=z5[['date','type','provider']]
	D5=D5[:-1]
	
	a = df.loc[df['12'] == day6].index[0]
	b = df.loc[df['12'] == day13].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z6=c.loc[:, ('12','13')]
	z6['date']=day6
	z6.rename(columns={ df.columns[12]:"type"}, inplace = True)
	z6.rename(columns={ df.columns[13]:"provider"}, inplace = True)
	D6=z6[['date','type','provider']]
	D6=D6[:-1]
	
	a = df.loc[df['0'] == day7].index[0]
	b = df.loc[df['0'] == day14].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z7=c.loc[:, ('0','1')]
	z7['date']=day7
	z7.rename(columns={ df.columns[0]:"type"}, inplace = True)
	z7.rename(columns={ df.columns[1]:"provider"}, inplace = True)
	D7=z7[['date','type','provider']]
	D7=D7[:-1]
	
	a = df.loc[df['2'] == day8].index[0]
	b = df.loc[df['2'] == day15].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z8=c.loc[:, ('2','3')]
	z8['date']=day8
	z8.rename(columns={ df.columns[2]:"type"}, inplace = True)
	z8.rename(columns={ df.columns[3]:"provider"}, inplace = True)
	D8=z8[['date','type','provider']]
	D8=D8[:-1]
	
	a = df.loc[df['4'] == day9].index[0]
	b = df.loc[df['4'] == day16].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z9=c.loc[:, ('4','5')]
	z9['date']=day9
	z9.rename(columns={ df.columns[4]:"type"}, inplace = True)
	z9.rename(columns={ df.columns[5]:"provider"}, inplace = True)
	D9=z9[['date','type','provider']]
	D9=D9[:-1]
	
	a = df.loc[df['6'] == day10].index[0]
	b = df.loc[df['6'] == day17].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z10=c.loc[:, ('6','7')]
	z10['date']=day10
	z10.rename(columns={ df.columns[6]:"type"}, inplace = True)
	z10.rename(columns={ df.columns[7]:"provider"}, inplace = True)
	D10=z10[['date','type','provider']]
	D10=D10[:-1]
	
	a = df.loc[df['8'] == day11].index[0]
	b = df.loc[df['8'] == day18].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z11=c.loc[:, ('8','9')]
	z11['date']=day11
	z11.rename(columns={ df.columns[8]:"type"}, inplace = True)
	z11.rename(columns={ df.columns[9]:"provider"}, inplace = True)
	D11=z11[['date','type','provider']]
	D11=D11[:-1]
	
	a = df.loc[df['10'] == day12].index[0]
	b = df.loc[df['10'] == day19].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z12=c.loc[:, ('10','11')]
	z12['date']=day12
	z12.rename(columns={ df.columns[10]:"type"}, inplace = True)
	z12.rename(columns={ df.columns[11]:"provider"}, inplace = True)
	D12=z12[['date','type','provider']]
	D12=D12[:-1]
	
	a = df.loc[df['12'] == day13].index[0]
	b = df.loc[df['12'] == day20].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z13=c.loc[:, ('12','13')]
	z13['date']=day13
	z13.rename(columns={ df.columns[12]:"type"}, inplace = True)
	z13.rename(columns={ df.columns[13]:"provider"}, inplace = True)
	D13=z13[['date','type','provider']]
	D13=D13[:-1]
	
	a = df.loc[df['0'] == day14].index[0]
	b = df.loc[df['0'] == day21].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z14=c.loc[:, ('0','1')]
	z14['date']=day14
	z14.rename(columns={ df.columns[0]:"type"}, inplace = True)
	z14.rename(columns={ df.columns[1]:"provider"}, inplace = True)
	D14=z14[['date','type','provider']]
	D14=D14[:-1]
	
	a = df.loc[df['2'] == day15].index[0]
	b = df.loc[df['2'] == day22].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z15=c.loc[:, ('2','3')]
	z15['date']=day15
	z15.rename(columns={ df.columns[2]:"type"}, inplace = True)
	z15.rename(columns={ df.columns[3]:"provider"}, inplace = True)
	D15=z15[['date','type','provider']]
	D15=D15[:-1]
	
	a = df.loc[df['4'] == day16].index[0]
	b = df.loc[df['4'] == day23].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z16=c.loc[:, ('4','5')]
	z16['date']=day16
	z16.rename(columns={ df.columns[4]:"type"}, inplace = True)
	z16.rename(columns={ df.columns[5]:"provider"}, inplace = True)
	D16=z16[['date','type','provider']]
	D16=D16[:-1]
	
	a = df.loc[df['6'] == day17].index[0]
	b = df.loc[df['6'] == day24].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z17=c.loc[:, ('6','7')]
	z17['date']=day17
	z17.rename(columns={ df.columns[6]:"type"}, inplace = True)
	z17.rename(columns={ df.columns[7]:"provider"}, inplace = True)
	D17=z17[['date','type','provider']]
	D17=D17[:-1]
	
	a = df.loc[df['8'] == day18].index[0]
	b = df.loc[df['8'] == day25].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z18=c.loc[:, ('8','9')]
	z18['date']=day18
	z18.rename(columns={ df.columns[8]:"type"}, inplace = True)
	z18.rename(columns={ df.columns[9]:"provider"}, inplace = True)
	D18=z18[['date','type','provider']]
	D18=D18[:-1]
	
	a = df.loc[df['10'] == day19].index[0]
	b = df.loc[df['10'] == day26].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z19=c.loc[:, ('10','11')]
	z19['date']=day19
	z19.rename(columns={ df.columns[10]:"type"}, inplace = True)
	z19.rename(columns={ df.columns[11]:"provider"}, inplace = True)
	D19=z19[['date','type','provider']]
	D19=D19[:-1]
	
	a = df.loc[df['12'] == day20].index[0]
	b = df.loc[df['12'] == day27].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z20=c.loc[:, ('12','13')]
	z20['date']=day20
	z20.rename(columns={ df.columns[12]:"type"}, inplace = True)
	z20.rename(columns={ df.columns[13]:"provider"}, inplace = True)
	D20=z20[['date','type','provider']]
	D20=D20[:-1]
	
	a = df.loc[df['0'] == day21].index[0]
	b = df.loc[df['0'] == day28].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z21=c.loc[:, ('0','1')]
	z21['date']=day21
	z21.rename(columns={ df.columns[0]:"type"}, inplace = True)
	z21.rename(columns={ df.columns[1]:"provider"}, inplace = True)
	D21=z21[['date','type','provider']]
	D21=D21[:-1]
	
	a = df.loc[df['2'] == day22].index[0]
	b = df.loc[df['2'] == day29].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z22=c.loc[:, ('2','3')]
	z22['date']=day22
	z22.rename(columns={ df.columns[2]:"type"}, inplace = True)
	z22.rename(columns={ df.columns[3]:"provider"}, inplace = True)
	D22=z22[['date','type','provider']]
	D22=D22[:-1]
	
	a = df.loc[df['4'] == day23].index[0]
	b = df.loc[df['4'] == day30].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z23=c.loc[:, ('4','5')]
	z23['date']=day23
	z23.rename(columns={ df.columns[4]:"type"}, inplace = True)
	z23.rename(columns={ df.columns[5]:"provider"}, inplace = True)
	D23=z23[['date','type','provider']]
	D23=D23[:-1]
	
	a = df.loc[df['6'] == day24].index[0]
	b = df.loc[df['6'] == day31].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z24=c.loc[:, ('6','7')]
	z24['date']=day24
	z24.rename(columns={ df.columns[6]:"type"}, inplace = True)
	z24.rename(columns={ df.columns[7]:"provider"}, inplace = True)
	D24=z24[['date','type','provider']]
	D24=D24[:-1]
	
	a = df.loc[df['8'] == day25].index[0]
	b = df.loc[df['8'] == day32].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z25=c.loc[:, ('8','9')]
	z25['date']=day25
	z25.rename(columns={ df.columns[8]:"type"}, inplace = True)
	z25.rename(columns={ df.columns[9]:"provider"}, inplace = True)
	D25=z25[['date','type','provider']]
	D25=D25[:-1]
	
	a = df.loc[df['10'] == day26].index[0]
	b = df.loc[df['10'] == day33].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z26=c.loc[:, ('10','11')]
	z26['date']=day26
	z26.rename(columns={ df.columns[10]:"type"}, inplace = True)
	z26.rename(columns={ df.columns[11]:"provider"}, inplace = True)
	D26=z26[['date','type','provider']]
	D26=D26[:-1]
	
	a = df.loc[df['12'] == day27].index[0]
	b = df.loc[df['12'] == day34].index[0]
	c = df.iloc[int(a)+1:int(b)]
	z27=c.loc[:, ('12','13')]
	z27['date']=day27
	z27.rename(columns={ df.columns[12]:"type"}, inplace = True)
	z27.rename(columns={ df.columns[13]:"provider"}, inplace = True)
	D27=z27[['date','type','provider']]
	D27=D27[:-1]
	
	dfx=pd.DataFrame(columns=D0.columns)
	
	dfx=pd.concat([dfx,D0, D1, D2, D3, D4, D5, D6, D7, D8, D9, D10, D11, D12, D13, D14, D15, D16, D17, D18, D19, D20, D21, D22, D23, D24, D25, D26, D27])
	
	dfx['clinic'] = "NYES"
	
	dfx.to_csv('nyes.csv',index=False)
	NYES=dfx.replace("Nyes Rd AM Continuity", "AM - Continuity", regex=True)
	NYES=NYES.replace("Nyes Rd PM Continuity", "PM - Continuity", regex=True)
	NYES.to_csv('nyes.csv',index=False)
	
	#############################################################################################################
	
	NYES['H'] = "H"
	NYEi = NYES[(NYES['type'] == 'AM - Continuity ')]
	NYEi['count'] = NYEi.groupby(['date'])['provider'].cumcount() + 0
	NYEi['class'] = "H" + NYEi['count'].astype(str)
	NYEi = NYEi.loc[:, ('date','type','provider','clinic','class')]
	NYEi.to_csv('1.csv',index=False)
	
	NYES['H'] = "H"
	NYEii = NYES[(NYES['type'] == 'PM - Continuity ')]
	NYEii['count'] = NYEii.groupby(['date'])['provider'].cumcount() + 10
	NYEii['class'] = "H" + NYEii['count'].astype(str)
	NYEii = NYEii.loc[:, ('date','type','provider','clinic','class')]
	NYEii.to_csv('2.csv',index=False)
	
	ETOWN['H'] = "H"
	ETOWNi = ETOWN[(ETOWN['type'] == 'AM - Continuity ')]
	ETOWNi['count'] = ETOWNi.groupby(['date'])['provider'].cumcount() + 0
	ETOWNi['class'] = "H" + ETOWNi['count'].astype(str)
	ETOWNi = ETOWNi.loc[:, ('date','type','provider','clinic','class')]
	ETOWNi.to_csv('3.csv',index=False)
	
	ETOWN['H'] = "H"
	ETOWNii = ETOWN[(ETOWN['type'] == 'PM - Continuity ')]
	ETOWNii['count'] = ETOWNii.groupby(['date'])['provider'].cumcount() + 10
	ETOWNii['class'] = "H" + ETOWNii['count'].astype(str)
	ETOWNii = ETOWNii.loc[:, ('date','type','provider','clinic','class')]
	ETOWNii.to_csv('4.csv',index=False)
	
	hope['class'] = "H"  # Top of Column
	hopeiii = hope[(hope['type'] == 'AM - ACUTES')]
	
	# Group by 'date' and count occurrences
	hopeiii['count'] = hopeiii.groupby(['date'])['provider'].cumcount()
	
	# Reserve H0 and H1 for the first two occurrences, subsequent start from H2
	hopeiii['class'] = hopeiii['count'].apply(
	    lambda count: "H0" if count == 0 else ("H1" if count == 1 else "H" + str(count + 2))
	)
	
	hopeiii = hopeiii.loc[:, ('date', 'type', 'provider', 'clinic', 'class')]
	hopeiii.to_csv('7.csv', index=False)
	
	hope['class'] = "H"  # Top of Column
	hopeiii = hope[(hope['type'] == 'AM - ACUTES ')]
	
	# Group by 'date' and count occurrences
	hopeiii['count'] = hopeiii.groupby(['date'])['provider'].cumcount()
	
	# Reserve H0 and H1 for the first two occurrences, subsequent start from H2
	hopeiii['class'] = hopeiii['count'].apply(
	    lambda count: "H0" if count == 0 else ("H1" if count == 1 else "H" + str(count + 2))
	)
	
	hopeiii = hopeiii.loc[:, ('date', 'type', 'provider', 'clinic', 'class')]
	hopeiii.to_csv('8.csv', index=False)
	
	######STARTS AT H10 (PM - ACUTES)
	hope['class'] = "H"  # Top of Next Column in Hope Drive
	hopeiiiii = hope[(hope['type'] == 'PM - ACUTES ')]
	
	# Group by 'date' and count occurrences
	hopeiiiii['count'] = hopeiiiii.groupby(['date'])['provider'].cumcount()
	
	# Reserve H10 and H11 for the first two occurrences, subsequent start from H12
	hopeiiiii['class'] = hopeiiiii['count'].apply(
	    lambda count: "H10" if count == 0 else ("H11" if count == 1 else "H" + str(count + 12))
	)
	
	hopeiiiii = hopeiiiii.loc[:, ('date', 'type', 'provider', 'clinic', 'class')]
	hopeiiiii.to_csv('9.csv', index=False)
	
	######STARTS AT H2 (AM - Continuity)
	hope['H'] = "H"
	hopei = hope[(hope['type'] == 'AM - Continuity ')]
	
	# Group by 'date' and count occurrences
	hopei['count'] = hopei.groupby(['date'])['provider'].cumcount() + 2  # Start at H2
	hopei['class'] = "H" + hopei['count'].astype(str)
	hopei = hopei.loc[:, ('date', 'type', 'provider', 'clinic', 'class')]
	hopei.to_csv('5.csv', index=False)
	
	######STARTS AT H12 (PM - Continuity)
	hope['H'] = "H"
	hopeii = hope[(hope['type'] == 'PM - Continuity ')]
	
	# Group by 'date' and count occurrences
	hopeii['count'] = hopeii.groupby(['date'])['provider'].cumcount() + 12  # Start at H12
	hopeii['class'] = "H" + hopeii['count'].astype(str)
	hopeii = hopeii.loc[:, ('date', 'type', 'provider', 'clinic', 'class')]
	hopeii.to_csv('6.csv', index=False)
	
	
	
	#############################################################################################################
	#MHS = pd.read_csv('MHS.csv')
	#MHS['H'] = "H"
	#MHSi = MHS[(MHS['type'] == 'AM - Continuity')]
	#MHSi['count'] = MHSi.groupby(['date'])['provider'].cumcount() + 0
	#MHSi['class'] = "H" + MHSi['count'].astype(str)
	#MHSi = MHSi.loc[:, ('date','type','provider','clinic','class')]
	#MHSi.to_csv('10.csv',index=False)
	
	#MHS['H'] = "H"
	#MHSii = MHS[(MHS['type'] == 'PM - Continuity')]
	#MHSii['count'] = MHSii.groupby(['date'])['provider'].cumcount() + 10
	#MHSii['class'] = "H" + MHSii['count'].astype(str)
	#MHSii = MHSii.loc[:, ('date','type','provider','clinic','class')]
	#MHSii.to_csv('11.csv',index=False)
	#############################################################################################################
	
	t1=pd.read_csv('1.csv')
	t2=pd.read_csv('2.csv')
	t3=pd.read_csv('3.csv')
	t4=pd.read_csv('4.csv')
	t5=pd.read_csv('5.csv')
	t6=pd.read_csv('6.csv')
	t7=pd.read_csv('7.csv')
	t8=pd.read_csv('8.csv')
	t9=pd.read_csv('9.csv')
	#t10=pd.read_csv('10.csv')
	#t11=pd.read_csv('11.csv')
	
	final2 = pd.DataFrame(columns=t1.columns)
	final2 = pd.concat([final2,t1,t2,t3,t4,t5,t6,t7,t8,t9]) #t10,t11])
	final2.to_csv('final2.csv',index=False)
	
	#final1 = pd.DataFrame(columns=NYEi.columns)
	#final1 = pd.concat([final1,hopei,hopeii,hopeiii,hopeiiii,ETOWNi,ETOWNii,NYEi,NYEii])
	#final1.to_csv('final1.csv',index=True)
	
	df=pd.read_csv('final2.csv',dtype=str) #MAP to Final2
	
	df['date'] = pd.to_datetime(df['date'])
	df['date'] = df['date'].dt.strftime('%m/%d/%Y')
	
	import csv
	
	dateMAP = xf201[['date','T']]
	
	dateMAP['date'] = pd.to_datetime(dateMAP['date'])
	dateMAP['date'] = dateMAP['date'].dt.strftime('%m/%d/%Y')
	
	dateMAP.to_csv('xxxDATEMAP.csv',index=False)
	
	mydict = {}
	with open('xxxDATEMAP.csv', mode='r')as inp:     #file is the objects you want to map. I want to map the IMP in this file to diagnosis.csv
		reader = csv.reader(inp)
		df1 = {rows[0]:rows[1] for rows in reader} 
	df['datecode'] = df.date.map(df1)               #'type' is the new column in the diagnosis file. 'encounter_id' is the key you are using to MAP 
	
	df['text'] = df['provider'] + " ~ "
	
	df['student'] = ""
	
	df = df.loc[:, ('date','type','provider','student','clinic','text','class','datecode')]
	
	df.to_csv('final.csv',index=False)
	df.to_excel('final.xlsx',index=False)
	
	import openpyxl
	from openpyxl.styles import Alignment
	
	# Load workbooks and worksheets
	wb = openpyxl.load_workbook('final.xlsx')
	ws = wb['Sheet1']
	
	wb1 = openpyxl.load_workbook('OPD.xlsx')
	ws1 = wb1['HOPE_DRIVE']
	
	def generate_mapping(start_value):
	    # Start with the special mappings
	    mapping = {f"H{i}": start_value + i for i in range(0, 20)}
	    
	    return mapping
	
	# Generate the mappings starting from 6 for the first group
	common_mapping_1 = generate_mapping(6)
	# Construct the t_mapping dictionary with the common structure for T0 to T6
	t_mapping_1 = {f"T{i}": common_mapping_1 for i in range(7)}
	
	# Generate the mappings starting from 30 for the second group
	common_mapping_2 = generate_mapping(30)
	# Construct the t_mapping dictionary with the common structure for T7 to T13
	t_mapping_2 = {f"T{i}": common_mapping_2 for i in range(7, 14)}
	
	# Generate the mappings starting from 54 for the second group
	common_mapping_3 = generate_mapping(54)
	# Construct the t_mapping dictionary with the common structure for T14 to T20
	t_mapping_3 = {f"T{i}": common_mapping_3 for i in range(14, 21)}
	
	# Generate the mappings starting from 30 for the second group
	common_mapping_4 = generate_mapping(78)
	# Construct the t_mapping dictionary with the common structure for T21 to T28
	t_mapping_4 = {f"T{i}": common_mapping_4 for i in range(21, 28)}
	
	# Combine both mappings
	combined_t_mapping = {**t_mapping_1, **t_mapping_2, **t_mapping_3, **t_mapping_4}
	
	# Now the `combined_t_mapping` will have the correct structure
	#print(combined_t_mapping)
	
	
	column_mapping = {"T0": 2, "T1": 3, "T2": 4, "T3": 5, "T4": 6, "T5": 7, "T6": 8,
	                  "T7": 2, "T8": 3, "T9": 4, "T10":5, "T11":6, "T12":7, "T13":8,
	                  "T14":2, "T15":3, "T16":4, "T17":5, "T18":6, "T19":7, "T20":8,
	                  "T21":2, "T22":3, "T23":4, "T24":5, "T25":6, "T26":7, "T27":8, 
	                 }
	
	# Iterate through rows in the worksheet
	for row in ws.iter_rows():
	    t_value = row[7].value  # Value in column H (index 7)
	    h_value = row[6].value  # Value in column G (index 6)
	    location = row[4].value  # Value in column E (index 4)
	
	    # Check conditions and apply mapping
	    if location == "HOPE_DRIVE" and t_value in combined_t_mapping and h_value in combined_t_mapping[t_value]:
	        target_row = combined_t_mapping[t_value][h_value]
	        target_column = column_mapping[t_value]
	        ws1.cell(row=target_row, column=target_column).value = row[5].value  # Value in column F (index 5)
	        ws1.cell(row=target_row, column=target_column).alignment = Alignment(horizontal='center')
	
	# Save updated workbook
	wb1.save('OPD.xlsx')
	
	###############################################################################################
	
	import openpyxl
	from openpyxl.styles import Alignment
	
	# Load workbooks and worksheets
	wb = openpyxl.load_workbook('final.xlsx')
	ws = wb['Sheet1']
	
	wb1 = openpyxl.load_workbook('OPD.xlsx')
	ws1 = wb1['ETOWN']
	
	def generate_mapping(start_value):
	    # Create a dictionary with H0 to H19 keys, with values starting from `start_value`
	    mapping = {f"H{i}": start_value + i for i in range(0, 20)}
	    
	    return mapping
	
	# Generate the mappings starting from 6 for the first group
	common_mapping_1 = generate_mapping(6)
	# Construct the t_mapping dictionary with the common structure for T0 to T6
	t_mapping_1 = {f"T{i}": common_mapping_1 for i in range(7)}
	
	# Generate the mappings starting from 30 for the second group
	common_mapping_2 = generate_mapping(30)
	# Construct the t_mapping dictionary with the common structure for T7 to T13
	t_mapping_2 = {f"T{i}": common_mapping_2 for i in range(7, 14)}
	
	# Generate the mappings starting from 54 for the third group
	common_mapping_3 = generate_mapping(54)
	# Construct the t_mapping dictionary with the common structure for T14 to T20
	t_mapping_3 = {f"T{i}": common_mapping_3 for i in range(14, 21)}
	
	# Generate the mappings starting from 78 for the fourth group
	common_mapping_4 = generate_mapping(78)
	# Construct the t_mapping dictionary with the common structure for T21 to T28
	t_mapping_4 = {f"T{i}": common_mapping_4 for i in range(21, 28)}
	
	# Combine both mappings
	combined_t_mapping = {**t_mapping_1, **t_mapping_2, **t_mapping_3, **t_mapping_4}
	
	# Now the `combined_t_mapping` will have the correct structure
	#print(combined_t_mapping)
	
	
	column_mapping = {"T0": 2, "T1": 3, "T2": 4, "T3": 5, "T4": 6, "T5": 7, "T6": 8,
	                  "T7": 2, "T8": 3, "T9": 4, "T10":5, "T11":6, "T12":7, "T13":8,
	                  "T14":2, "T15":3, "T16":4, "T17":5, "T18":6, "T19":7, "T20":8,
	                  "T21":2, "T22":3, "T23":4, "T24":5, "T25":6, "T26":7, "T27":8, 
	                 }
	
	# Iterate through rows in the worksheet
	for row in ws.iter_rows():
	    t_value = row[7].value  # Value in column H (index 7)
	    h_value = row[6].value  # Value in column G (index 6)
	    location = row[4].value  # Value in column E (index 4)
	
	    # Check conditions and apply mapping
	    if location == "ETOWN" and t_value in combined_t_mapping and h_value in combined_t_mapping[t_value]:
	        target_row = combined_t_mapping[t_value][h_value]
	        target_column = column_mapping[t_value]
	        ws1.cell(row=target_row, column=target_column).value = row[5].value  # Value in column F (index 5)
	        ws1.cell(row=target_row, column=target_column).alignment = Alignment(horizontal='center')
	
	# Save updated workbook
	wb1.save('OPD.xlsx')
	
	###############################################################################################
	
	import openpyxl
	from openpyxl.styles import Alignment
	
	# Load workbooks and worksheets
	wb = openpyxl.load_workbook('final.xlsx')
	ws = wb['Sheet1']
	
	wb1 = openpyxl.load_workbook('OPD.xlsx')
	ws1 = wb1['NYES']
	
	def generate_mapping(start_value):
	    # Create a dictionary with H0 to H19 keys, with values starting from `start_value`
	    mapping = {f"H{i}": start_value + i for i in range(0, 20)}
	    
	    return mapping
	
	# Generate the mappings starting from 6 for the first group
	common_mapping_1 = generate_mapping(6)
	# Construct the t_mapping dictionary with the common structure for T0 to T6
	t_mapping_1 = {f"T{i}": common_mapping_1 for i in range(7)}
	
	# Generate the mappings starting from 30 for the second group
	common_mapping_2 = generate_mapping(30)
	# Construct the t_mapping dictionary with the common structure for T7 to T13
	t_mapping_2 = {f"T{i}": common_mapping_2 for i in range(7, 14)}
	
	# Generate the mappings starting from 54 for the third group
	common_mapping_3 = generate_mapping(54)
	# Construct the t_mapping dictionary with the common structure for T14 to T20
	t_mapping_3 = {f"T{i}": common_mapping_3 for i in range(14, 21)}
	
	# Generate the mappings starting from 78 for the fourth group
	common_mapping_4 = generate_mapping(78)
	# Construct the t_mapping dictionary with the common structure for T21 to T28
	t_mapping_4 = {f"T{i}": common_mapping_4 for i in range(21, 28)}
	
	# Combine both mappings
	combined_t_mapping = {**t_mapping_1, **t_mapping_2, **t_mapping_3, **t_mapping_4}
	
	# Now the `combined_t_mapping` will have the correct structure
	#print(combined_t_mapping)
	
	
	column_mapping = {"T0": 2, "T1": 3, "T2": 4, "T3": 5, "T4": 6, "T5": 7, "T6": 8,
	                  "T7": 2, "T8": 3, "T9": 4, "T10":5, "T11":6, "T12":7, "T13":8,
	                  "T14":2, "T15":3, "T16":4, "T17":5, "T18":6, "T19":7, "T20":8,
	                  "T21":2, "T22":3, "T23":4, "T24":5, "T25":6, "T26":7, "T27":8, 
	                 }
	
	# Iterate through rows in the worksheet
	for row in ws.iter_rows():
	    t_value = row[7].value  # Value in column H (index 7)
	    h_value = row[6].value  # Value in column G (index 6)
	    location = row[4].value  # Value in column E (index 4)
	
	    # Check conditions and apply mapping
	    if location == "NYES" and t_value in combined_t_mapping and h_value in combined_t_mapping[t_value]:
	        target_row = combined_t_mapping[t_value][h_value]
	        target_column = column_mapping[t_value]
	        ws1.cell(row=target_row, column=target_column).value = row[5].value  # Value in column F (index 5)
	        ws1.cell(row=target_row, column=target_column).alignment = Alignment(horizontal='center')
	
	# Save updated workbook
	wb1.save('OPD.xlsx')
	###############################################################################################
	# Button to trigger the download
	if st.button('Create OPD'):
	    # Path to the existing 'OPD.xlsx' workbook
	    file_path = 'OPD.xlsx'  # Replace with your file path if it's stored somewhere else
	
	    # Read the workbook into memory
	    with open(file_path, 'rb') as file:
	        file_data = file.read()
	
	    # Provide a download button for the existing OPD.xlsx file
	    st.download_button(
	        label="Download OPD.xlsx",
	        data=file_data,
	        file_name="OPD.xlsx",
	        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
	    )
		
elif st.session_state.page == "Create Student Schedule":
    st.title("Create Student Schedule")
    # Upload the OPD.xlsx file
    uploaded_opd_file = st.file_uploader("Upload OPD.xlsx file", type="xlsx")
    uploaded_book4_file = st.file_uploader("Upload Book4.xlsx file", type="xlsx")
    
    if uploaded_opd_file:
        try:
            # Read the uploaded OPD file into a pandas dataframe
            df_opd = pd.read_excel(uploaded_opd_file)
            st.write("File successfully uploaded and loaded.")
            
            # Store the uploaded file in session state for use later
            st.session_state.uploaded_files['OPD.xlsx'] = uploaded_opd_file
                
        except Exception as e:
            st.error(f"Error reading the uploaded file: {e}")
    else:
        st.write("Please upload the OPD.xlsx file to proceed.")

    if uploaded_book4_file:
        try:
            # Read the uploaded OPD file into a pandas dataframe
            df_opd = pd.read_excel(uploaded_book4_file)
            st.write("File successfully uploaded and loaded.")
            
            # Store the uploaded file in session state for use later
            st.session_state.uploaded_book4_file['Book4.xlsx'] = uploaded_book4_file
                
        except Exception as e:
            st.error(f"Error reading the uploaded file: {e}")
    else:
        st.write("Please upload the Book4.xlsx file to proceed.")

    # Button to go to the next page
    if st.button("Create List"):
        st.session_state.page = "Create List"  # Update the session state to go to the next page
        st.rerun()  # Use st.rerun() instead of st.experimental_rerun() to force rerun and update the page

elif st.session_state.page == "Create List":
    st.title("Create List")

    # Ensure the OPD.xlsx file exists in the session state before proceeding
    if 'OPD.xlsx' in st.session_state.uploaded_files:
        uploaded_opd_file = st.session_state.uploaded_files['OPD.xlsx']
        
        try:
            # Read the OPD file into a dataframe
            df_opd = pd.read_excel(uploaded_opd_file)
            
            # Display the first few rows of the OPD data for verification
            st.dataframe(df_opd.head())
            
            # Save the OPD file again without the index column
            df_opd.to_excel('OPD.xlsx', index=False)
            st.write("OPD.xlsx file has been successfully saved.")
        
        except Exception as e:
            st.error(f"Error processing the OPD file: {e}")
    else:
        st.error("No OPD file found in session state.")


    if 'Book4.xlsx' in st.session_state.uploaded_book4_file:
        uploaded_book4_file = st.session_state.uploaded_book4_file['Book4.xlsx']
        
        try:
            # Read the OPD file into a dataframe
            book4 = pd.read_excel(uploaded_book4_file)
            
            # Display the first few rows of the OPD data for verification
            st.dataframe(book4.head())
            
            # Save the OPD file again without the index column
            book4.to_excel('Book4.xlsx', index=False)
            st.write("Book4.xlsx file has been successfully saved.")
        
        except Exception as e:
            st.error(f"Error processing the Book4 file: {e}")
    else:
        st.error("No Book4 file found in session state.")
	    
    
    # Ensure the "HOPE_DRIVE" sheet exists in the uploaded Excel file
    try:
        df = pd.read_excel(uploaded_opd_file)

        # Extract the value from row 2, column 1
        test_date = df.iloc[2, 1]

        # Ensure that test_date is a valid datetime object
        # If it's a string, convert it into a datetime object using pd.to_datetime
        if isinstance(test_date, str):
            test_date = pd.to_datetime(test_date, errors='coerce')  # Handle invalid date gracefully

        # Check if the date is valid (not NaT)
        if pd.isna(test_date):
            print("Invalid date format in the cell.")
        else:
            # Format the date to mm-dd-yyyy
            formatted_date = test_date.strftime('%m-%d-%Y')

            # Calculate the start date (use test_date directly as it's already a datetime object)
            start_date = test_date

            # Calculate the end date (34 days from the start date)
            end_date = start_date + timedelta(days=34)

            # Generate the date range
            date_range = pd.date_range(start=start_date, end=end_date)

            # Create a DataFrame with the formatted dates
            xf201 = pd.DataFrame({'date': date_range})

            # Convert the date to the desired format (e.g., "November 25, 2024")
            xf201['convert'] = xf201['date'].dt.strftime('%B %-d, %Y')  # This works in Unix-like systems

            # Add additional columns
            xf201['t'] = "T"
            xf201['c'] = xf201.index + 0  # Simple index-based column
            xf201['T'] = xf201['t'].astype(str) + xf201['c'].astype(str)

        # Creating the dateMAP DataFrame
        dateMAP = xf201[['date', 'T']].copy()  # Use .copy() to avoid the SettingWithCopyWarning

        # Convert 'date' column to datetime and then format it
        dateMAP['date'] = pd.to_datetime(dateMAP['date'])
        dateMAP['date'] = dateMAP['date'].dt.strftime('%m/%d/%Y')

        # Save the result to a CSV file
        dateMAP.to_csv('xxxDATEMAP.csv', index=False)
        
        read_file = pd.read_excel(uploaded_opd_file, sheet_name='HOPE_DRIVE')
        read_file.to_csv ('hopedrive.csv', index = False, header=False)
        df=pd.read_csv('hopedrive.csv')
        clinictype=df.iloc[3:23, 0:1]
        a1 = pd.DataFrame(clinictype, columns = ['type'])
        a2 = pd.DataFrame(clinictype, columns = ['type'])
        a3 = pd.DataFrame(clinictype, columns = ['type'])
        a4 = pd.DataFrame(clinictype, columns = ['type'])
        a5 = pd.DataFrame(clinictype, columns = ['type'])
        a6 = pd.DataFrame(clinictype, columns = ['type'])
        a7 = pd.DataFrame(clinictype, columns = ['type'])

        a1['type']=clinictype
        a2['type']=clinictype
        a3['type']=clinictype
        a4['type']=clinictype
        a5['type']=clinictype
        a6['type']=clinictype
        a7['type']=clinictype

        week1day1=a1.replace(to_replace=r'- Continuity', value='', regex=True)
        week1day2=a2.replace(to_replace=r'- Continuity', value='', regex=True)
        week1day3=a3.replace(to_replace=r'- Continuity', value='', regex=True)
        week1day4=a4.replace(to_replace=r'- Continuity', value='', regex=True)
        week1day5=a5.replace(to_replace=r'- Continuity', value='', regex=True)
        week1day6=a6.replace(to_replace=r'- Continuity', value='', regex=True)
        week1day7=a7.replace(to_replace=r'- Continuity', value='', regex=True)
        
        day1=df.iloc[1,1]
        day2=df.iloc[1,2]
        day3=df.iloc[1,3]
        day4=df.iloc[1,4]
        day5=df.iloc[1,5]
        day6=df.iloc[1,6]
        day7=df.iloc[1,7]

        week1day1['date']=day1
        week1day2['date']=day2
        week1day3['date']=day3
        week1day4['date']=day4
        week1day5['date']=day5
        week1day6['date']=day6
        week1day7['date']=day7

        provider1=df.iloc[3:23,1]
        provider2=df.iloc[3:23,2]
        provider3=df.iloc[3:23,3]
        provider4=df.iloc[3:23,4]
        provider5=df.iloc[3:23,5]
        provider6=df.iloc[3:23,6]
        provider7=df.iloc[3:23,7]

        week1day1['provider']=provider1
        week1day2['provider']=provider2
        week1day3['provider']=provider3
        week1day4['provider']=provider4
        week1day5['provider']=provider5
        week1day6['provider']=provider6
        week1day7['provider']=provider7

        week1day1['clinic']="HOPE_DRIVE"
        week1day2['clinic']="HOPE_DRIVE"
        week1day3['clinic']="HOPE_DRIVE"
        week1day4['clinic']="HOPE_DRIVE"
        week1day5['clinic']="HOPE_DRIVE"
        week1day6['clinic']="HOPE_DRIVE"
        week1day7['clinic']="HOPE_DRIVE"

        week1=pd.DataFrame(columns=week1day1.columns)
        week1=pd.concat([week1,week1day1,week1day2,week1day3,week1day4,week1day5,week1day6,week1day7])
        week1.to_csv('week1.csv',index=False)

        clinictype=df.iloc[27:47, 0:1]
        b1 = pd.DataFrame(clinictype, columns = ['type'])
        b2 = pd.DataFrame(clinictype, columns = ['type'])
        b3 = pd.DataFrame(clinictype, columns = ['type'])
        b4 = pd.DataFrame(clinictype, columns = ['type'])
        b5 = pd.DataFrame(clinictype, columns = ['type'])
        b6 = pd.DataFrame(clinictype, columns = ['type'])
        b7 = pd.DataFrame(clinictype, columns = ['type'])

        b1['type']=clinictype
        b2['type']=clinictype
        b3['type']=clinictype
        b4['type']=clinictype
        b5['type']=clinictype
        b6['type']=clinictype
        b7['type']=clinictype

        week2day1=b1.replace(to_replace=r'- Continuity', value='', regex=True)
        week2day2=b2.replace(to_replace=r'- Continuity', value='', regex=True)
        week2day3=b3.replace(to_replace=r'- Continuity', value='', regex=True)
        week2day4=b4.replace(to_replace=r'- Continuity', value='', regex=True)
        week2day5=b5.replace(to_replace=r'- Continuity', value='', regex=True)
        week2day6=b6.replace(to_replace=r'- Continuity', value='', regex=True)
        week2day7=b7.replace(to_replace=r'- Continuity', value='', regex=True)

        day1=df.iloc[25,1]
        day2=df.iloc[25,2]
        day3=df.iloc[25,3]
        day4=df.iloc[25,4]
        day5=df.iloc[25,5]
        day6=df.iloc[25,6]
        day7=df.iloc[25,7]

        week2day1['date']=day1
        week2day2['date']=day2
        week2day3['date']=day3
        week2day4['date']=day4
        week2day5['date']=day5
        week2day6['date']=day6
        week2day7['date']=day7

        provider1=df.iloc[27:47,1]
        provider2=df.iloc[27:47,2]
        provider3=df.iloc[27:47,3]
        provider4=df.iloc[27:47,4]
        provider5=df.iloc[27:47,5]
        provider6=df.iloc[27:47,6]
        provider7=df.iloc[27:47,7]

        week2day1['provider']=provider1
        week2day2['provider']=provider2
        week2day3['provider']=provider3
        week2day4['provider']=provider4
        week2day5['provider']=provider5
        week2day6['provider']=provider6
        week2day7['provider']=provider7

        week2day1['clinic']="HOPE_DRIVE"
        week2day2['clinic']="HOPE_DRIVE"
        week2day3['clinic']="HOPE_DRIVE"
        week2day4['clinic']="HOPE_DRIVE"
        week2day5['clinic']="HOPE_DRIVE"
        week2day6['clinic']="HOPE_DRIVE"
        week2day7['clinic']="HOPE_DRIVE"

        week2=pd.DataFrame(columns=week2day1.columns)
        week2=pd.concat([week2,week2day1,week2day2,week2day3,week2day4,week2day5,week2day6,week2day7])
        week2.to_csv('week2.csv',index=False)

        clinictype=df.iloc[51:71, 0:1]
        c1 = pd.DataFrame(clinictype, columns = ['type'])
        c2 = pd.DataFrame(clinictype, columns = ['type'])
        c3 = pd.DataFrame(clinictype, columns = ['type'])
        c4 = pd.DataFrame(clinictype, columns = ['type'])
        c5 = pd.DataFrame(clinictype, columns = ['type'])
        c6 = pd.DataFrame(clinictype, columns = ['type'])
        c7 = pd.DataFrame(clinictype, columns = ['type'])

        c1['type']=clinictype
        c2['type']=clinictype
        c3['type']=clinictype
        c4['type']=clinictype
        c5['type']=clinictype
        c6['type']=clinictype
        c7['type']=clinictype

        week3day1=c1.replace(to_replace=r'- Continuity', value='', regex=True)
        week3day2=c2.replace(to_replace=r'- Continuity', value='', regex=True)
        week3day3=c3.replace(to_replace=r'- Continuity', value='', regex=True)
        week3day4=c4.replace(to_replace=r'- Continuity', value='', regex=True)
        week3day5=c5.replace(to_replace=r'- Continuity', value='', regex=True)
        week3day6=c6.replace(to_replace=r'- Continuity', value='', regex=True)
        week3day7=c7.replace(to_replace=r'- Continuity', value='', regex=True)

        day1=df.iloc[49,1]
        day2=df.iloc[49,2]
        day3=df.iloc[49,3]
        day4=df.iloc[49,4]
        day5=df.iloc[49,5]
        day6=df.iloc[49,6]
        day7=df.iloc[49,7]

        week3day1['date']=day1
        week3day2['date']=day2
        week3day3['date']=day3
        week3day4['date']=day4
        week3day5['date']=day5
        week3day6['date']=day6
        week3day7['date']=day7

        provider1=df.iloc[51:71,1]
        provider2=df.iloc[51:71,2]
        provider3=df.iloc[51:71,3]
        provider4=df.iloc[51:71,4]
        provider5=df.iloc[51:71,5]
        provider6=df.iloc[51:71,6]
        provider7=df.iloc[51:71,7]

        week3day1['provider']=provider1
        week3day2['provider']=provider2
        week3day3['provider']=provider3
        week3day4['provider']=provider4
        week3day5['provider']=provider5
        week3day6['provider']=provider6
        week3day7['provider']=provider7

        week3day1['clinic']="HOPE_DRIVE"
        week3day2['clinic']="HOPE_DRIVE"
        week3day3['clinic']="HOPE_DRIVE"
        week3day4['clinic']="HOPE_DRIVE"
        week3day5['clinic']="HOPE_DRIVE"
        week3day6['clinic']="HOPE_DRIVE"
        week3day7['clinic']="HOPE_DRIVE"

        week3=pd.DataFrame(columns=week3day1.columns)
        week3=pd.concat([week3,week3day1,week3day2,week3day3,week3day4,week3day5,week3day6,week3day7])
        week3.to_csv('week3.csv',index=False)

        clinictype=df.iloc[75:95, 0:1]
        d1 = pd.DataFrame(clinictype, columns = ['type'])
        d2 = pd.DataFrame(clinictype, columns = ['type'])
        d3 = pd.DataFrame(clinictype, columns = ['type'])
        d4 = pd.DataFrame(clinictype, columns = ['type'])
        d5 = pd.DataFrame(clinictype, columns = ['type'])
        d6 = pd.DataFrame(clinictype, columns = ['type'])
        d7 = pd.DataFrame(clinictype, columns = ['type'])

        d1['type']=clinictype
        d2['type']=clinictype
        d3['type']=clinictype
        d4['type']=clinictype
        d5['type']=clinictype
        d6['type']=clinictype
        d7['type']=clinictype

        week4day1=d1.replace(to_replace=r'- Continuity', value='', regex=True)
        week4day2=d2.replace(to_replace=r'- Continuity', value='', regex=True)
        week4day3=d3.replace(to_replace=r'- Continuity', value='', regex=True)
        week4day4=d4.replace(to_replace=r'- Continuity', value='', regex=True)
        week4day5=d5.replace(to_replace=r'- Continuity', value='', regex=True)
        week4day6=d6.replace(to_replace=r'- Continuity', value='', regex=True)
        week4day7=d7.replace(to_replace=r'- Continuity', value='', regex=True)


        day1=df.iloc[73,1]
        day2=df.iloc[73,2]
        day3=df.iloc[73,3]
        day4=df.iloc[73,4]
        day5=df.iloc[73,5]
        day6=df.iloc[73,6]
        day7=df.iloc[73,7]

        week4day1['date']=day1
        week4day2['date']=day2
        week4day3['date']=day3
        week4day4['date']=day4
        week4day5['date']=day5
        week4day6['date']=day6
        week4day7['date']=day7

        provider1=df.iloc[75:95,1]
        provider2=df.iloc[75:95,2]
        provider3=df.iloc[75:95,3]
        provider4=df.iloc[75:95,4]
        provider5=df.iloc[75:95,5]
        provider6=df.iloc[75:95,6]
        provider7=df.iloc[75:95,7]

        week4day1['provider']=provider1
        week4day2['provider']=provider2
        week4day3['provider']=provider3
        week4day4['provider']=provider4
        week4day5['provider']=provider5
        week4day6['provider']=provider6
        week4day7['provider']=provider7

        week4day1['clinic']="HOPE_DRIVE"
        week4day2['clinic']="HOPE_DRIVE"
        week4day3['clinic']="HOPE_DRIVE"
        week4day4['clinic']="HOPE_DRIVE"
        week4day5['clinic']="HOPE_DRIVE"
        week4day6['clinic']="HOPE_DRIVE"
        week4day7['clinic']="HOPE_DRIVE"

        week4=pd.DataFrame(columns=week4day1.columns)
        week4=pd.concat([week4,week4day1,week4day2,week4day3,week4day4,week4day5,week4day6,week4day7])
        week4.to_csv('week4.csv',index=False)

        hope=pd.DataFrame(columns=week1.columns)
        hope=pd.concat([hope,week1,week2,week3,week4])

        hope.to_csv('hope.csv',index=False)

        # Handle AM Continuity for hopei
        hope['H'] = "H"
        hopei = hope[hope['type'] == 'AM '].copy()  # Ensure we're working with a copy of the slice
        hopei['count'] = hopei.groupby(['date'])['provider'].cumcount() + 2  # Starts at H2 for AM
        hopei['class'] = "H" + hopei['count'].astype(str)
        hopei = hopei.loc[:, ('date', 'type', 'provider', 'clinic', 'class')]
        hopei.to_csv('5.csv', index=False)

        # Handle PM Continuity for hopeii
        hopeii = hope[hope['type'] == 'PM '].copy()  # Ensure we're working with a copy of the slice
        hopeii['count'] = hopeii.groupby(['date'])['provider'].cumcount() + 12  # Starts at H12 for PM
        hopeii['class'] = "H" + hopeii['count'].astype(str)
        hopeii = hopeii.loc[:, ('date', 'type', 'provider', 'clinic', 'class')]
        hopeii.to_csv('6.csv', index=False)

        # Handle AM - ACUTES for hopeiii
        hopeiii = hope[hope['type'] == 'AM - ACUTES'].copy()  # Ensure we're working with a copy of the slice
        hopeiii['count'] = hopeiii.groupby(['date'])['provider'].cumcount()  # Starts at H0 for AM-ACUTES
        hopeiii['class'] = "H" + hopeiii['count'].astype(str)
        hopeiii = hopeiii.loc[:, ('date', 'type', 'provider', 'clinic', 'class')]
        hopeiii.to_csv('7.csv', index=False)

        # Handle PM - ACUTES for hopeiiii
        hopeiiii = hope[hope['type'] == 'PM - ACUTES'].copy()  # Ensure we're working with a copy of the slice
        hopeiiii['count'] = hopeiiii.groupby(['date'])['provider'].cumcount() + 10  # Starts at H10 for PM-ACUTES
        hopeiiii['class'] = "H" + hopeiiii['count'].astype(str)
        hopeiiii = hopeiiii.loc[:, ('date', 'type', 'provider', 'clinic', 'class')]
        hopeiiii.to_csv('8.csv', index=False)

        # Combine all the data into one DataFrame
        hopes = pd.DataFrame(columns=hopei.columns)
        hopes = pd.concat([hopei, hopeii, hopeiii, hopeiiii])

        # Save the combined DataFrame to CSV
        hopes.to_csv('hopes.csv', index=False)


        ####################################NYES#############################################################################
        import pandas as pd
        read_file = pd.read_excel (uploaded_opd_file, sheet_name='NYES')
        read_file.to_csv ('nyesroad.csv', index = False, header=False)
        df=pd.read_csv('nyesroad.csv')

        clinictype=df.iloc[3:23, 0:1]
        a1 = pd.DataFrame(clinictype, columns = ['type'])
        a2 = pd.DataFrame(clinictype, columns = ['type'])
        a3 = pd.DataFrame(clinictype, columns = ['type'])
        a4 = pd.DataFrame(clinictype, columns = ['type'])
        a5 = pd.DataFrame(clinictype, columns = ['type'])
        a6 = pd.DataFrame(clinictype, columns = ['type'])
        a7 = pd.DataFrame(clinictype, columns = ['type'])

        a1['type']=clinictype
        a2['type']=clinictype
        a3['type']=clinictype
        a4['type']=clinictype
        a5['type']=clinictype
        a6['type']=clinictype
        a7['type']=clinictype

        week1day1=a1.replace(to_replace=r'- Continuity', value='', regex=True)
        week1day2=a2.replace(to_replace=r'- Continuity', value='', regex=True)
        week1day3=a3.replace(to_replace=r'- Continuity', value='', regex=True)
        week1day4=a4.replace(to_replace=r'- Continuity', value='', regex=True)
        week1day5=a5.replace(to_replace=r'- Continuity', value='', regex=True)
        week1day6=a6.replace(to_replace=r'- Continuity', value='', regex=True)
        week1day7=a7.replace(to_replace=r'- Continuity', value='', regex=True)

        day1=df.iloc[1,1]
        day2=df.iloc[1,2]
        day3=df.iloc[1,3]
        day4=df.iloc[1,4]
        day5=df.iloc[1,5]
        day6=df.iloc[1,6]
        day7=df.iloc[1,7]

        week1day1['date']=day1
        week1day2['date']=day2
        week1day3['date']=day3
        week1day4['date']=day4
        week1day5['date']=day5
        week1day6['date']=day6
        week1day7['date']=day7

        provider1=df.iloc[3:23,1]
        provider2=df.iloc[3:23,2]
        provider3=df.iloc[3:23,3]
        provider4=df.iloc[3:23,4]
        provider5=df.iloc[3:23,5]
        provider6=df.iloc[3:23,6]
        provider7=df.iloc[3:23,7]

        week1day1['provider']=provider1
        week1day2['provider']=provider2
        week1day3['provider']=provider3
        week1day4['provider']=provider4
        week1day5['provider']=provider5
        week1day6['provider']=provider6
        week1day7['provider']=provider7

        week1day1['clinic']="NYES"
        week1day2['clinic']="NYES"
        week1day3['clinic']="NYES"
        week1day4['clinic']="NYES"
        week1day5['clinic']="NYES"
        week1day6['clinic']="NYES"
        week1day7['clinic']="NYES"

        week1=pd.DataFrame(columns=week1day1.columns)
        week1=pd.concat([week1,week1day1,week1day2,week1day3,week1day4,week1day5,week1day6,week1day7])
        week1.to_csv('week1.csv',index=False)

        clinictype=df.iloc[27:47, 0:1]
        b1 = pd.DataFrame(clinictype, columns = ['type'])
        b2 = pd.DataFrame(clinictype, columns = ['type'])
        b3 = pd.DataFrame(clinictype, columns = ['type'])
        b4 = pd.DataFrame(clinictype, columns = ['type'])
        b5 = pd.DataFrame(clinictype, columns = ['type'])
        b6 = pd.DataFrame(clinictype, columns = ['type'])
        b7 = pd.DataFrame(clinictype, columns = ['type'])

        b1['type']=clinictype
        b2['type']=clinictype
        b3['type']=clinictype
        b4['type']=clinictype
        b5['type']=clinictype
        b6['type']=clinictype
        b7['type']=clinictype

        week2day1=b1.replace(to_replace=r'- Continuity', value='', regex=True)
        week2day2=b2.replace(to_replace=r'- Continuity', value='', regex=True)
        week2day3=b3.replace(to_replace=r'- Continuity', value='', regex=True)
        week2day4=b4.replace(to_replace=r'- Continuity', value='', regex=True)
        week2day5=b5.replace(to_replace=r'- Continuity', value='', regex=True)
        week2day6=b6.replace(to_replace=r'- Continuity', value='', regex=True)
        week2day7=b7.replace(to_replace=r'- Continuity', value='', regex=True)

        day1=df.iloc[25,1]
        day2=df.iloc[25,2]
        day3=df.iloc[25,3]
        day4=df.iloc[25,4]
        day5=df.iloc[25,5]
        day6=df.iloc[25,6]
        day7=df.iloc[25,7]

        week2day1['date']=day1
        week2day2['date']=day2
        week2day3['date']=day3
        week2day4['date']=day4
        week2day5['date']=day5
        week2day6['date']=day6
        week2day7['date']=day7

        provider1=df.iloc[27:47,1]
        provider2=df.iloc[27:47,2]
        provider3=df.iloc[27:47,3]
        provider4=df.iloc[27:47,4]
        provider5=df.iloc[27:47,5]
        provider6=df.iloc[27:47,6]
        provider7=df.iloc[27:47,7]

        week2day1['provider']=provider1
        week2day2['provider']=provider2
        week2day3['provider']=provider3
        week2day4['provider']=provider4
        week2day5['provider']=provider5
        week2day6['provider']=provider6
        week2day7['provider']=provider7

        week2day1['clinic']="NYES"
        week2day2['clinic']="NYES"
        week2day3['clinic']="NYES"
        week2day4['clinic']="NYES"
        week2day5['clinic']="NYES"
        week2day6['clinic']="NYES"
        week2day7['clinic']="NYES"

        week2=pd.DataFrame(columns=week2day1.columns)
        week2=pd.concat([week2,week2day1,week2day2,week2day3,week2day4,week2day5,week2day6,week2day7])
        week2.to_csv('week2.csv',index=False)

        clinictype=df.iloc[51:71, 0:1]
        c1 = pd.DataFrame(clinictype, columns = ['type'])
        c2 = pd.DataFrame(clinictype, columns = ['type'])
        c3 = pd.DataFrame(clinictype, columns = ['type'])
        c4 = pd.DataFrame(clinictype, columns = ['type'])
        c5 = pd.DataFrame(clinictype, columns = ['type'])
        c6 = pd.DataFrame(clinictype, columns = ['type'])
        c7 = pd.DataFrame(clinictype, columns = ['type'])

        c1['type']=clinictype
        c2['type']=clinictype
        c3['type']=clinictype
        c4['type']=clinictype
        c5['type']=clinictype
        c6['type']=clinictype
        c7['type']=clinictype

        week3day1=c1.replace(to_replace=r'- Continuity', value='', regex=True)
        week3day2=c2.replace(to_replace=r'- Continuity', value='', regex=True)
        week3day3=c3.replace(to_replace=r'- Continuity', value='', regex=True)
        week3day4=c4.replace(to_replace=r'- Continuity', value='', regex=True)
        week3day5=c5.replace(to_replace=r'- Continuity', value='', regex=True)
        week3day6=c6.replace(to_replace=r'- Continuity', value='', regex=True)
        week3day7=c7.replace(to_replace=r'- Continuity', value='', regex=True)

        day1=df.iloc[49,1]
        day2=df.iloc[49,2]
        day3=df.iloc[49,3]
        day4=df.iloc[49,4]
        day5=df.iloc[49,5]
        day6=df.iloc[49,6]
        day7=df.iloc[49,7]

        week3day1['date']=day1
        week3day2['date']=day2
        week3day3['date']=day3
        week3day4['date']=day4
        week3day5['date']=day5
        week3day6['date']=day6
        week3day7['date']=day7

        provider1=df.iloc[51:71,1]
        provider2=df.iloc[51:71,2]
        provider3=df.iloc[51:71,3]
        provider4=df.iloc[51:71,4]
        provider5=df.iloc[51:71,5]
        provider6=df.iloc[51:71,6]
        provider7=df.iloc[51:71,7]

        week3day1['provider']=provider1
        week3day2['provider']=provider2
        week3day3['provider']=provider3
        week3day4['provider']=provider4
        week3day5['provider']=provider5
        week3day6['provider']=provider6
        week3day7['provider']=provider7

        week3day1['clinic']="NYES"
        week3day2['clinic']="NYES"
        week3day3['clinic']="NYES"
        week3day4['clinic']="NYES"
        week3day5['clinic']="NYES"
        week3day6['clinic']="NYES"
        week3day7['clinic']="NYES"

        week3=pd.DataFrame(columns=week3day1.columns)
        week3=pd.concat([week3,week3day1,week3day2,week3day3,week3day4,week3day5,week3day6,week3day7])
        week3.to_csv('week3.csv',index=False)

        clinictype=df.iloc[75:95, 0:1]
        d1 = pd.DataFrame(clinictype, columns = ['type'])
        d2 = pd.DataFrame(clinictype, columns = ['type'])
        d3 = pd.DataFrame(clinictype, columns = ['type'])
        d4 = pd.DataFrame(clinictype, columns = ['type'])
        d5 = pd.DataFrame(clinictype, columns = ['type'])
        d6 = pd.DataFrame(clinictype, columns = ['type'])
        d7 = pd.DataFrame(clinictype, columns = ['type'])

        d1['type']=clinictype
        d2['type']=clinictype
        d3['type']=clinictype
        d4['type']=clinictype
        d5['type']=clinictype
        d6['type']=clinictype
        d7['type']=clinictype

        week4day1=d1.replace(to_replace=r'- Continuity', value='', regex=True)
        week4day2=d2.replace(to_replace=r'- Continuity', value='', regex=True)
        week4day3=d3.replace(to_replace=r'- Continuity', value='', regex=True)
        week4day4=d4.replace(to_replace=r'- Continuity', value='', regex=True)
        week4day5=d5.replace(to_replace=r'- Continuity', value='', regex=True)
        week4day6=d6.replace(to_replace=r'- Continuity', value='', regex=True)
        week4day7=d7.replace(to_replace=r'- Continuity', value='', regex=True)


        day1=df.iloc[73,1]
        day2=df.iloc[73,2]
        day3=df.iloc[73,3]
        day4=df.iloc[73,4]
        day5=df.iloc[73,5]
        day6=df.iloc[73,6]
        day7=df.iloc[73,7]

        week4day1['date']=day1
        week4day2['date']=day2
        week4day3['date']=day3
        week4day4['date']=day4
        week4day5['date']=day5
        week4day6['date']=day6
        week4day7['date']=day7

        provider1=df.iloc[75:95,1]
        provider2=df.iloc[75:95,2]
        provider3=df.iloc[75:95,3]
        provider4=df.iloc[75:95,4]
        provider5=df.iloc[75:95,5]
        provider6=df.iloc[75:95,6]
        provider7=df.iloc[75:95,7]

        week4day1['provider']=provider1
        week4day2['provider']=provider2
        week4day3['provider']=provider3
        week4day4['provider']=provider4
        week4day5['provider']=provider5
        week4day6['provider']=provider6
        week4day7['provider']=provider7

        week4day1['clinic']="NYES"
        week4day2['clinic']="NYES"
        week4day3['clinic']="NYES"
        week4day4['clinic']="NYES"
        week4day5['clinic']="NYES"
        week4day6['clinic']="NYES"
        week4day7['clinic']="NYES"

        week4=pd.DataFrame(columns=week4day1.columns)
        week4=pd.concat([week4,week4day1,week4day2,week4day3,week4day4,week4day5,week4day6,week4day7])
        week4.to_csv('week4.csv',index=False)

        hope=pd.DataFrame(columns=week1.columns)
        hope=pd.concat([hope,week1,week2,week3,week4])
        hope.to_csv('nyes.csv',index=False)

        # Handle AM Continuity for NYE (First set)
        hope['H'] = "H"
        NYEi = hope[hope['type'] == 'AM '].copy()  # Ensure we're working with a copy
        NYEi.loc[:, 'count'] = NYEi.groupby(['date'])['provider'].cumcount() + 0  # Starts at H0 for AM
        NYEi.loc[:, 'class'] = "H" + NYEi['count'].astype(str)
        NYEi = NYEi.loc[:, ('date', 'type', 'provider', 'clinic', 'class')]
        NYEi.to_csv('1.csv', index=False)

        # Handle PM Continuity for NYE (Second set)
        hope['H'] = "H"
        NYEii = hope[hope['type'] == 'PM '].copy()  # Ensure we're working with a copy
        NYEii.loc[:, 'count'] = NYEii.groupby(['date'])['provider'].cumcount() + 10  # Starts at H10 for PM
        NYEii.loc[:, 'class'] = "H" + NYEii['count'].astype(str)
        NYEii = NYEii.loc[:, ('date', 'type', 'provider', 'clinic', 'class')]
        NYEii.to_csv('2.csv', index=False)

        # Combine AM and PM DataFrames
        nyess = pd.DataFrame(columns=NYEi.columns)
        nyess = pd.concat([NYEi, NYEii])
        nyess.to_csv('nyess.csv', index=False)

        ##############################ETOWN##############################################################################################
        import pandas as pd
        read_file = pd.read_excel (uploaded_opd_file, sheet_name='ETOWN')
        read_file.to_csv ('etownroad.csv', index = False, header=False)
        df=pd.read_csv('etownroad.csv')

        clinictype=df.iloc[3:23, 0:1]
        a1 = pd.DataFrame(clinictype, columns = ['type'])
        a2 = pd.DataFrame(clinictype, columns = ['type'])
        a3 = pd.DataFrame(clinictype, columns = ['type'])
        a4 = pd.DataFrame(clinictype, columns = ['type'])
        a5 = pd.DataFrame(clinictype, columns = ['type'])
        a6 = pd.DataFrame(clinictype, columns = ['type'])
        a7 = pd.DataFrame(clinictype, columns = ['type'])

        a1['type']=clinictype
        a2['type']=clinictype
        a3['type']=clinictype
        a4['type']=clinictype
        a5['type']=clinictype
        a6['type']=clinictype
        a7['type']=clinictype

        week1day1=a1.replace(to_replace=r'- Continuity', value='', regex=True)
        week1day2=a2.replace(to_replace=r'- Continuity', value='', regex=True)
        week1day3=a3.replace(to_replace=r'- Continuity', value='', regex=True)
        week1day4=a4.replace(to_replace=r'- Continuity', value='', regex=True)
        week1day5=a5.replace(to_replace=r'- Continuity', value='', regex=True)
        week1day6=a6.replace(to_replace=r'- Continuity', value='', regex=True)
        week1day7=a7.replace(to_replace=r'- Continuity', value='', regex=True)

        day1=df.iloc[1,1]
        day2=df.iloc[1,2]
        day3=df.iloc[1,3]
        day4=df.iloc[1,4]
        day5=df.iloc[1,5]
        day6=df.iloc[1,6]
        day7=df.iloc[1,7]

        week1day1['date']=day1
        week1day2['date']=day2
        week1day3['date']=day3
        week1day4['date']=day4
        week1day5['date']=day5
        week1day6['date']=day6
        week1day7['date']=day7

        provider1=df.iloc[3:23,1]
        provider2=df.iloc[3:23,2]
        provider3=df.iloc[3:23,3]
        provider4=df.iloc[3:23,4]
        provider5=df.iloc[3:23,5]
        provider6=df.iloc[3:23,6]
        provider7=df.iloc[3:23,7]

        week1day1['provider']=provider1
        week1day2['provider']=provider2
        week1day3['provider']=provider3
        week1day4['provider']=provider4
        week1day5['provider']=provider5
        week1day6['provider']=provider6
        week1day7['provider']=provider7

        week1day1['clinic']="ETOWN"
        week1day2['clinic']="ETOWN"
        week1day3['clinic']="ETOWN"
        week1day4['clinic']="ETOWN"
        week1day5['clinic']="ETOWN"
        week1day6['clinic']="ETOWN"
        week1day7['clinic']="ETOWN"

        week1=pd.DataFrame(columns=week1day1.columns)
        week1=pd.concat([week1,week1day1,week1day2,week1day3,week1day4,week1day5,week1day6,week1day7])
        week1.to_csv('week1.csv',index=False)

        clinictype=df.iloc[27:47, 0:1]
        b1 = pd.DataFrame(clinictype, columns = ['type'])
        b2 = pd.DataFrame(clinictype, columns = ['type'])
        b3 = pd.DataFrame(clinictype, columns = ['type'])
        b4 = pd.DataFrame(clinictype, columns = ['type'])
        b5 = pd.DataFrame(clinictype, columns = ['type'])
        b6 = pd.DataFrame(clinictype, columns = ['type'])
        b7 = pd.DataFrame(clinictype, columns = ['type'])

        b1['type']=clinictype
        b2['type']=clinictype
        b3['type']=clinictype
        b4['type']=clinictype
        b5['type']=clinictype
        b6['type']=clinictype
        b7['type']=clinictype

        week2day1=b1.replace(to_replace=r'- Continuity', value='', regex=True)
        week2day2=b2.replace(to_replace=r'- Continuity', value='', regex=True)
        week2day3=b3.replace(to_replace=r'- Continuity', value='', regex=True)
        week2day4=b4.replace(to_replace=r'- Continuity', value='', regex=True)
        week2day5=b5.replace(to_replace=r'- Continuity', value='', regex=True)
        week2day6=b6.replace(to_replace=r'- Continuity', value='', regex=True)
        week2day7=b7.replace(to_replace=r'- Continuity', value='', regex=True)

        day1=df.iloc[25,1]
        day2=df.iloc[25,2]
        day3=df.iloc[25,3]
        day4=df.iloc[25,4]
        day5=df.iloc[25,5]
        day6=df.iloc[25,6]
        day7=df.iloc[25,7]

        week2day1['date']=day1
        week2day2['date']=day2
        week2day3['date']=day3
        week2day4['date']=day4
        week2day5['date']=day5
        week2day6['date']=day6
        week2day7['date']=day7

        provider1=df.iloc[27:47,1]
        provider2=df.iloc[27:47,2]
        provider3=df.iloc[27:47,3]
        provider4=df.iloc[27:47,4]
        provider5=df.iloc[27:47,5]
        provider6=df.iloc[27:47,6]
        provider7=df.iloc[27:47,7]

        week2day1['provider']=provider1
        week2day2['provider']=provider2
        week2day3['provider']=provider3
        week2day4['provider']=provider4
        week2day5['provider']=provider5
        week2day6['provider']=provider6
        week2day7['provider']=provider7

        week2day1['clinic']="ETOWN"
        week2day2['clinic']="ETOWN"
        week2day3['clinic']="ETOWN"
        week2day4['clinic']="ETOWN"
        week2day5['clinic']="ETOWN"
        week2day6['clinic']="ETOWN"
        week2day7['clinic']="ETOWN"

        week2=pd.DataFrame(columns=week2day1.columns)
        week2=pd.concat([week2,week2day1,week2day2,week2day3,week2day4,week2day5,week2day6,week2day7])
        week2.to_csv('week2.csv',index=False)

        clinictype=df.iloc[51:71, 0:1]
        c1 = pd.DataFrame(clinictype, columns = ['type'])
        c2 = pd.DataFrame(clinictype, columns = ['type'])
        c3 = pd.DataFrame(clinictype, columns = ['type'])
        c4 = pd.DataFrame(clinictype, columns = ['type'])
        c5 = pd.DataFrame(clinictype, columns = ['type'])
        c6 = pd.DataFrame(clinictype, columns = ['type'])
        c7 = pd.DataFrame(clinictype, columns = ['type'])

        c1['type']=clinictype
        c2['type']=clinictype
        c3['type']=clinictype
        c4['type']=clinictype
        c5['type']=clinictype
        c6['type']=clinictype
        c7['type']=clinictype

        week3day1=c1.replace(to_replace=r'- Continuity', value='', regex=True)
        week3day2=c2.replace(to_replace=r'- Continuity', value='', regex=True)
        week3day3=c3.replace(to_replace=r'- Continuity', value='', regex=True)
        week3day4=c4.replace(to_replace=r'- Continuity', value='', regex=True)
        week3day5=c5.replace(to_replace=r'- Continuity', value='', regex=True)
        week3day6=c6.replace(to_replace=r'- Continuity', value='', regex=True)
        week3day7=c7.replace(to_replace=r'- Continuity', value='', regex=True)

        day1=df.iloc[49,1]
        day2=df.iloc[49,2]
        day3=df.iloc[49,3]
        day4=df.iloc[49,4]
        day5=df.iloc[49,5]
        day6=df.iloc[49,6]
        day7=df.iloc[49,7]

        week3day1['date']=day1
        week3day2['date']=day2
        week3day3['date']=day3
        week3day4['date']=day4
        week3day5['date']=day5
        week3day6['date']=day6
        week3day7['date']=day7

        provider1=df.iloc[51:71,1]
        provider2=df.iloc[51:71,2]
        provider3=df.iloc[51:71,3]
        provider4=df.iloc[51:71,4]
        provider5=df.iloc[51:71,5]
        provider6=df.iloc[51:71,6]
        provider7=df.iloc[51:71,7]

        week3day1['provider']=provider1
        week3day2['provider']=provider2
        week3day3['provider']=provider3
        week3day4['provider']=provider4
        week3day5['provider']=provider5
        week3day6['provider']=provider6
        week3day7['provider']=provider7

        week3day1['clinic']="ETOWN"
        week3day2['clinic']="ETOWN"
        week3day3['clinic']="ETOWN"
        week3day4['clinic']="ETOWN"
        week3day5['clinic']="ETOWN"
        week3day6['clinic']="ETOWN"
        week3day7['clinic']="ETOWN"

        week3=pd.DataFrame(columns=week3day1.columns)
        week3=pd.concat([week3,week3day1,week3day2,week3day3,week3day4,week3day5,week3day6,week3day7])
        week3.to_csv('week3.csv',index=False)

        clinictype=df.iloc[75:95, 0:1]
        d1 = pd.DataFrame(clinictype, columns = ['type'])
        d2 = pd.DataFrame(clinictype, columns = ['type'])
        d3 = pd.DataFrame(clinictype, columns = ['type'])
        d4 = pd.DataFrame(clinictype, columns = ['type'])
        d5 = pd.DataFrame(clinictype, columns = ['type'])
        d6 = pd.DataFrame(clinictype, columns = ['type'])
        d7 = pd.DataFrame(clinictype, columns = ['type'])

        d1['type']=clinictype
        d2['type']=clinictype
        d3['type']=clinictype
        d4['type']=clinictype
        d5['type']=clinictype
        d6['type']=clinictype
        d7['type']=clinictype

        week4day1=d1.replace(to_replace=r'- Continuity', value='', regex=True)
        week4day2=d2.replace(to_replace=r'- Continuity', value='', regex=True)
        week4day3=d3.replace(to_replace=r'- Continuity', value='', regex=True)
        week4day4=d4.replace(to_replace=r'- Continuity', value='', regex=True)
        week4day5=d5.replace(to_replace=r'- Continuity', value='', regex=True)
        week4day6=d6.replace(to_replace=r'- Continuity', value='', regex=True)
        week4day7=d7.replace(to_replace=r'- Continuity', value='', regex=True)


        day1=df.iloc[73,1]
        day2=df.iloc[73,2]
        day3=df.iloc[73,3]
        day4=df.iloc[73,4]
        day5=df.iloc[73,5]
        day6=df.iloc[73,6]
        day7=df.iloc[73,7]

        week4day1['date']=day1
        week4day2['date']=day2
        week4day3['date']=day3
        week4day4['date']=day4
        week4day5['date']=day5
        week4day6['date']=day6
        week4day7['date']=day7

        provider1=df.iloc[75:95,1]
        provider2=df.iloc[75:95,2]
        provider3=df.iloc[75:95,3]
        provider4=df.iloc[75:95,4]
        provider5=df.iloc[75:95,5]
        provider6=df.iloc[75:95,6]
        provider7=df.iloc[75:95,7]

        week4day1['provider']=provider1
        week4day2['provider']=provider2
        week4day3['provider']=provider3
        week4day4['provider']=provider4
        week4day5['provider']=provider5
        week4day6['provider']=provider6
        week4day7['provider']=provider7

        week4day1['clinic']="ETOWN"
        week4day2['clinic']="ETOWN"
        week4day3['clinic']="ETOWN"
        week4day4['clinic']="ETOWN"
        week4day5['clinic']="ETOWN"
        week4day6['clinic']="ETOWN"
        week4day7['clinic']="ETOWN"

        week4=pd.DataFrame(columns=week4day1.columns)
        week4=pd.concat([week4,week4day1,week4day2,week4day3,week4day4,week4day5,week4day6,week4day7])
        week4.to_csv('week4.csv',index=False)

        hope=pd.DataFrame(columns=week1.columns)
        hope=pd.concat([hope,week1,week2,week3,week4])
        hope.to_csv('etown.csv',index=False)

        # Handle AM Continuity for ETOWN (First set)
        hope['H'] = "H"
        ETOWNi = hope[hope['type'] == 'AM '].copy()  # Ensure we're working with a copy
        ETOWNi.loc[:, 'count'] = ETOWNi.groupby(['date'])['provider'].cumcount() + 0  # Starts at H0 for AM
        ETOWNi.loc[:, 'class'] = "H" + ETOWNi['count'].astype(str)
        ETOWNi = ETOWNi.loc[:, ('date', 'type', 'provider', 'clinic', 'class')]
        ETOWNi.to_csv('3.csv', index=False)

        # Handle PM Continuity for ETOWN (Second set)
        hope['H'] = "H"
        ETOWNii = hope[hope['type'] == 'PM '].copy()  # Ensure we're working with a copy
        ETOWNii.loc[:, 'count'] = ETOWNii.groupby(['date'])['provider'].cumcount() + 10  # Starts at H10 for PM
        ETOWNii.loc[:, 'class'] = "H" + ETOWNii['count'].astype(str)
        ETOWNii = ETOWNii.loc[:, ('date', 'type', 'provider', 'clinic', 'class')]
        ETOWNii.to_csv('4.csv', index=False)

        # Combine AM and PM DataFrames for ETOWN
        etowns = pd.DataFrame(columns=ETOWNi.columns)
        etowns = pd.concat([ETOWNi, ETOWNii])
        etowns.to_csv('etowns.csv', index=False)

        ##############################EXTRA##############################################################################################
        import pandas as pd
        read_file = pd.read_excel (uploaded_opd_file, sheet_name='EXTRA')
        read_file.to_csv ('extra.csv', index = False, header=False)
        df=pd.read_csv('extra.csv')

        clinictype=df.iloc[3:23, 0:1]
        a1 = pd.DataFrame(clinictype, columns = ['type'])
        a2 = pd.DataFrame(clinictype, columns = ['type'])
        a3 = pd.DataFrame(clinictype, columns = ['type'])
        a4 = pd.DataFrame(clinictype, columns = ['type'])
        a5 = pd.DataFrame(clinictype, columns = ['type'])
        a6 = pd.DataFrame(clinictype, columns = ['type'])
        a7 = pd.DataFrame(clinictype, columns = ['type'])

        a1['type']=clinictype
        a2['type']=clinictype
        a3['type']=clinictype
        a4['type']=clinictype
        a5['type']=clinictype
        a6['type']=clinictype
        a7['type']=clinictype

        week1day1=a1.replace(to_replace=r'- Continuity', value='', regex=True)
        week1day2=a2.replace(to_replace=r'- Continuity', value='', regex=True)
        week1day3=a3.replace(to_replace=r'- Continuity', value='', regex=True)
        week1day4=a4.replace(to_replace=r'- Continuity', value='', regex=True)
        week1day5=a5.replace(to_replace=r'- Continuity', value='', regex=True)
        week1day6=a6.replace(to_replace=r'- Continuity', value='', regex=True)
        week1day7=a7.replace(to_replace=r'- Continuity', value='', regex=True)

        day1=df.iloc[1,1]
        day2=df.iloc[1,2]
        day3=df.iloc[1,3]
        day4=df.iloc[1,4]
        day5=df.iloc[1,5]
        day6=df.iloc[1,6]
        day7=df.iloc[1,7]

        week1day1['date']=day1
        week1day2['date']=day2
        week1day3['date']=day3
        week1day4['date']=day4
        week1day5['date']=day5
        week1day6['date']=day6
        week1day7['date']=day7

        provider1=df.iloc[3:23,1]
        provider2=df.iloc[3:23,2]
        provider3=df.iloc[3:23,3]
        provider4=df.iloc[3:23,4]
        provider5=df.iloc[3:23,5]
        provider6=df.iloc[3:23,6]
        provider7=df.iloc[3:23,7]

        week1day1['provider']=provider1
        week1day2['provider']=provider2
        week1day3['provider']=provider3
        week1day4['provider']=provider4
        week1day5['provider']=provider5
        week1day6['provider']=provider6
        week1day7['provider']=provider7

        week1day1['clinic']="EXTRA"
        week1day2['clinic']="EXTRA"
        week1day3['clinic']="EXTRA"
        week1day4['clinic']="EXTRA"
        week1day5['clinic']="EXTRA"
        week1day6['clinic']="EXTRA"
        week1day7['clinic']="EXTRA"

        week1=pd.DataFrame(columns=week1day1.columns)
        week1=pd.concat([week1,week1day1,week1day2,week1day3,week1day4,week1day5,week1day6,week1day7])
        week1.to_csv('week1.csv',index=False)

        clinictype=df.iloc[27:47, 0:1]
        b1 = pd.DataFrame(clinictype, columns = ['type'])
        b2 = pd.DataFrame(clinictype, columns = ['type'])
        b3 = pd.DataFrame(clinictype, columns = ['type'])
        b4 = pd.DataFrame(clinictype, columns = ['type'])
        b5 = pd.DataFrame(clinictype, columns = ['type'])
        b6 = pd.DataFrame(clinictype, columns = ['type'])
        b7 = pd.DataFrame(clinictype, columns = ['type'])

        b1['type']=clinictype
        b2['type']=clinictype
        b3['type']=clinictype
        b4['type']=clinictype
        b5['type']=clinictype
        b6['type']=clinictype
        b7['type']=clinictype

        week2day1=b1.replace(to_replace=r'- Continuity', value='', regex=True)
        week2day2=b2.replace(to_replace=r'- Continuity', value='', regex=True)
        week2day3=b3.replace(to_replace=r'- Continuity', value='', regex=True)
        week2day4=b4.replace(to_replace=r'- Continuity', value='', regex=True)
        week2day5=b5.replace(to_replace=r'- Continuity', value='', regex=True)
        week2day6=b6.replace(to_replace=r'- Continuity', value='', regex=True)
        week2day7=b7.replace(to_replace=r'- Continuity', value='', regex=True)

        day1=df.iloc[25,1]
        day2=df.iloc[25,2]
        day3=df.iloc[25,3]
        day4=df.iloc[25,4]
        day5=df.iloc[25,5]
        day6=df.iloc[25,6]
        day7=df.iloc[25,7]

        week2day1['date']=day1
        week2day2['date']=day2
        week2day3['date']=day3
        week2day4['date']=day4
        week2day5['date']=day5
        week2day6['date']=day6
        week2day7['date']=day7

        provider1=df.iloc[27:47,1]
        provider2=df.iloc[27:47,2]
        provider3=df.iloc[27:47,3]
        provider4=df.iloc[27:47,4]
        provider5=df.iloc[27:47,5]
        provider6=df.iloc[27:47,6]
        provider7=df.iloc[27:47,7]

        week2day1['provider']=provider1
        week2day2['provider']=provider2
        week2day3['provider']=provider3
        week2day4['provider']=provider4
        week2day5['provider']=provider5
        week2day6['provider']=provider6
        week2day7['provider']=provider7

        week2day1['clinic']="EXTRA"
        week2day2['clinic']="EXTRA"
        week2day3['clinic']="EXTRA"
        week2day4['clinic']="EXTRA"
        week2day5['clinic']="EXTRA"
        week2day6['clinic']="EXTRA"
        week2day7['clinic']="EXTRA"

        week2=pd.DataFrame(columns=week2day1.columns)
        week2=pd.concat([week2,week2day1,week2day2,week2day3,week2day4,week2day5,week2day6,week2day7])
        week2.to_csv('week2.csv',index=False)

        clinictype=df.iloc[51:71, 0:1]
        c1 = pd.DataFrame(clinictype, columns = ['type'])
        c2 = pd.DataFrame(clinictype, columns = ['type'])
        c3 = pd.DataFrame(clinictype, columns = ['type'])
        c4 = pd.DataFrame(clinictype, columns = ['type'])
        c5 = pd.DataFrame(clinictype, columns = ['type'])
        c6 = pd.DataFrame(clinictype, columns = ['type'])
        c7 = pd.DataFrame(clinictype, columns = ['type'])

        c1['type']=clinictype
        c2['type']=clinictype
        c3['type']=clinictype
        c4['type']=clinictype
        c5['type']=clinictype
        c6['type']=clinictype
        c7['type']=clinictype

        week3day1=c1.replace(to_replace=r'- Continuity', value='', regex=True)
        week3day2=c2.replace(to_replace=r'- Continuity', value='', regex=True)
        week3day3=c3.replace(to_replace=r'- Continuity', value='', regex=True)
        week3day4=c4.replace(to_replace=r'- Continuity', value='', regex=True)
        week3day5=c5.replace(to_replace=r'- Continuity', value='', regex=True)
        week3day6=c6.replace(to_replace=r'- Continuity', value='', regex=True)
        week3day7=c7.replace(to_replace=r'- Continuity', value='', regex=True)

        day1=df.iloc[49,1]
        day2=df.iloc[49,2]
        day3=df.iloc[49,3]
        day4=df.iloc[49,4]
        day5=df.iloc[49,5]
        day6=df.iloc[49,6]
        day7=df.iloc[49,7]

        week3day1['date']=day1
        week3day2['date']=day2
        week3day3['date']=day3
        week3day4['date']=day4
        week3day5['date']=day5
        week3day6['date']=day6
        week3day7['date']=day7

        provider1=df.iloc[51:71,1]
        provider2=df.iloc[51:71,2]
        provider3=df.iloc[51:71,3]
        provider4=df.iloc[51:71,4]
        provider5=df.iloc[51:71,5]
        provider6=df.iloc[51:71,6]
        provider7=df.iloc[51:71,7]

        week3day1['provider']=provider1
        week3day2['provider']=provider2
        week3day3['provider']=provider3
        week3day4['provider']=provider4
        week3day5['provider']=provider5
        week3day6['provider']=provider6
        week3day7['provider']=provider7

        week3day1['clinic']="EXTRA"
        week3day2['clinic']="EXTRA"
        week3day3['clinic']="EXTRA"
        week3day4['clinic']="EXTRA"
        week3day5['clinic']="EXTRA"
        week3day6['clinic']="EXTRA"
        week3day7['clinic']="EXTRA"

        week3=pd.DataFrame(columns=week3day1.columns)
        week3=pd.concat([week3,week3day1,week3day2,week3day3,week3day4,week3day5,week3day6,week3day7])
        week3.to_csv('week3.csv',index=False)

        clinictype=df.iloc[75:95, 0:1]
        d1 = pd.DataFrame(clinictype, columns = ['type'])
        d2 = pd.DataFrame(clinictype, columns = ['type'])
        d3 = pd.DataFrame(clinictype, columns = ['type'])
        d4 = pd.DataFrame(clinictype, columns = ['type'])
        d5 = pd.DataFrame(clinictype, columns = ['type'])
        d6 = pd.DataFrame(clinictype, columns = ['type'])
        d7 = pd.DataFrame(clinictype, columns = ['type'])

        d1['type']=clinictype
        d2['type']=clinictype
        d3['type']=clinictype
        d4['type']=clinictype
        d5['type']=clinictype
        d6['type']=clinictype
        d7['type']=clinictype

        week4day1=d1.replace(to_replace=r'- Continuity', value='', regex=True)
        week4day2=d2.replace(to_replace=r'- Continuity', value='', regex=True)
        week4day3=d3.replace(to_replace=r'- Continuity', value='', regex=True)
        week4day4=d4.replace(to_replace=r'- Continuity', value='', regex=True)
        week4day5=d5.replace(to_replace=r'- Continuity', value='', regex=True)
        week4day6=d6.replace(to_replace=r'- Continuity', value='', regex=True)
        week4day7=d7.replace(to_replace=r'- Continuity', value='', regex=True)


        day1=df.iloc[73,1]
        day2=df.iloc[73,2]
        day3=df.iloc[73,3]
        day4=df.iloc[73,4]
        day5=df.iloc[73,5]
        day6=df.iloc[73,6]
        day7=df.iloc[73,7]

        week4day1['date']=day1
        week4day2['date']=day2
        week4day3['date']=day3
        week4day4['date']=day4
        week4day5['date']=day5
        week4day6['date']=day6
        week4day7['date']=day7

        provider1=df.iloc[75:95,1]
        provider2=df.iloc[75:95,2]
        provider3=df.iloc[75:95,3]
        provider4=df.iloc[75:95,4]
        provider5=df.iloc[75:95,5]
        provider6=df.iloc[75:95,6]
        provider7=df.iloc[75:95,7]

        week4day1['provider']=provider1
        week4day2['provider']=provider2
        week4day3['provider']=provider3
        week4day4['provider']=provider4
        week4day5['provider']=provider5
        week4day6['provider']=provider6
        week4day7['provider']=provider7

        week4day1['clinic']="EXTRA"
        week4day2['clinic']="EXTRA"
        week4day3['clinic']="EXTRA"
        week4day4['clinic']="EXTRA"
        week4day5['clinic']="EXTRA"
        week4day6['clinic']="EXTRA"
        week4day7['clinic']="EXTRA"

        week4=pd.DataFrame(columns=week4day1.columns)
        week4=pd.concat([week4,week4day1,week4day2,week4day3,week4day4,week4day5,week4day6,week4day7])
        week4.to_csv('week4.csv',index=False)

        hope=pd.DataFrame(columns=week1.columns)
        hope=pd.concat([hope,week1,week2,week3,week4])
        hope.to_csv('extra.csv',index=False)

        hope['H'] = "H"
        extrai = hope[hope['type'] == 'AM '].copy()  # Make sure we're working with a copy
        extrai.loc[:, 'count'] = extrai.groupby(['date'])['provider'].cumcount() + 0  # Starts at H0 for AM
        extrai.loc[:, 'class'] = "H" + extrai['count'].astype(str)
        extrai = extrai.loc[:, ('date', 'type', 'provider', 'clinic', 'class')]
        extrai.to_csv('3.csv', index=False)

        # Handle PM Continuity for EXTRAI
        hope['H'] = "H"
        extraii = hope[hope['type'] == 'PM '].copy()  # Make sure we're working with a copy
        extraii.loc[:, 'count'] = extraii.groupby(['date'])['provider'].cumcount() + 10  # Starts at H10 for PM
        extraii.loc[:, 'class'] = "H" + extraii['count'].astype(str)
        extraii = extraii.loc[:, ('date', 'type', 'provider', 'clinic', 'class')]
        extraii.to_csv('4.csv', index=False)

        # Combine AM and PM DataFrames for EXTRAI
        extras = pd.DataFrame(columns=extrai.columns)
        extras = pd.concat([extrai, extraii])
        extras.to_csv('extras.csv', index=False)
        ##############################MHS##############################################################################################
        import pandas as pd
        read_file = pd.read_excel (uploaded_opd_file, sheet_name='MHS')
        read_file.to_csv ('mhss.csv', index = False, header=False)
        df=pd.read_csv('mhss.csv')

        clinictype=df.iloc[3:23, 0:1]
        a1 = pd.DataFrame(clinictype, columns = ['type'])
        a2 = pd.DataFrame(clinictype, columns = ['type'])
        a3 = pd.DataFrame(clinictype, columns = ['type'])
        a4 = pd.DataFrame(clinictype, columns = ['type'])
        a5 = pd.DataFrame(clinictype, columns = ['type'])
        a6 = pd.DataFrame(clinictype, columns = ['type'])
        a7 = pd.DataFrame(clinictype, columns = ['type'])

        a1['type']=clinictype
        a2['type']=clinictype
        a3['type']=clinictype
        a4['type']=clinictype
        a5['type']=clinictype
        a6['type']=clinictype
        a7['type']=clinictype

        week1day1=a1.replace(to_replace=r'- Continuity', value='', regex=True)
        week1day2=a2.replace(to_replace=r'- Continuity', value='', regex=True)
        week1day3=a3.replace(to_replace=r'- Continuity', value='', regex=True)
        week1day4=a4.replace(to_replace=r'- Continuity', value='', regex=True)
        week1day5=a5.replace(to_replace=r'- Continuity', value='', regex=True)
        week1day6=a6.replace(to_replace=r'- Continuity', value='', regex=True)
        week1day7=a7.replace(to_replace=r'- Continuity', value='', regex=True)

        day1=df.iloc[1,1]
        day2=df.iloc[1,2]
        day3=df.iloc[1,3]
        day4=df.iloc[1,4]
        day5=df.iloc[1,5]
        day6=df.iloc[1,6]
        day7=df.iloc[1,7]

        week1day1['date']=day1
        week1day2['date']=day2
        week1day3['date']=day3
        week1day4['date']=day4
        week1day5['date']=day5
        week1day6['date']=day6
        week1day7['date']=day7

        provider1=df.iloc[3:23,1]
        provider2=df.iloc[3:23,2]
        provider3=df.iloc[3:23,3]
        provider4=df.iloc[3:23,4]
        provider5=df.iloc[3:23,5]
        provider6=df.iloc[3:23,6]
        provider7=df.iloc[3:23,7]

        week1day1['provider']=provider1
        week1day2['provider']=provider2
        week1day3['provider']=provider3
        week1day4['provider']=provider4
        week1day5['provider']=provider5
        week1day6['provider']=provider6
        week1day7['provider']=provider7

        week1day1['clinic']="AAC"
        week1day2['clinic']="AAC"
        week1day3['clinic']="AAC"
        week1day4['clinic']="AAC"
        week1day5['clinic']="AAC"
        week1day6['clinic']="AAC"
        week1day7['clinic']="AAC"

        week1=pd.DataFrame(columns=week1day1.columns)
        week1=pd.concat([week1,week1day1,week1day2,week1day3,week1day4,week1day5,week1day6,week1day7])
        week1.to_csv('week1.csv',index=False)

        clinictype=df.iloc[27:47, 0:1]
        b1 = pd.DataFrame(clinictype, columns = ['type'])
        b2 = pd.DataFrame(clinictype, columns = ['type'])
        b3 = pd.DataFrame(clinictype, columns = ['type'])
        b4 = pd.DataFrame(clinictype, columns = ['type'])
        b5 = pd.DataFrame(clinictype, columns = ['type'])
        b6 = pd.DataFrame(clinictype, columns = ['type'])
        b7 = pd.DataFrame(clinictype, columns = ['type'])

        b1['type']=clinictype
        b2['type']=clinictype
        b3['type']=clinictype
        b4['type']=clinictype
        b5['type']=clinictype
        b6['type']=clinictype
        b7['type']=clinictype

        week2day1=b1.replace(to_replace=r'- Continuity', value='', regex=True)
        week2day2=b2.replace(to_replace=r'- Continuity', value='', regex=True)
        week2day3=b3.replace(to_replace=r'- Continuity', value='', regex=True)
        week2day4=b4.replace(to_replace=r'- Continuity', value='', regex=True)
        week2day5=b5.replace(to_replace=r'- Continuity', value='', regex=True)
        week2day6=b6.replace(to_replace=r'- Continuity', value='', regex=True)
        week2day7=b7.replace(to_replace=r'- Continuity', value='', regex=True)

        day1=df.iloc[25,1]
        day2=df.iloc[25,2]
        day3=df.iloc[25,3]
        day4=df.iloc[25,4]
        day5=df.iloc[25,5]
        day6=df.iloc[25,6]
        day7=df.iloc[25,7]

        week2day1['date']=day1
        week2day2['date']=day2
        week2day3['date']=day3
        week2day4['date']=day4
        week2day5['date']=day5
        week2day6['date']=day6
        week2day7['date']=day7

        provider1=df.iloc[27:47,1]
        provider2=df.iloc[27:47,2]
        provider3=df.iloc[27:47,3]
        provider4=df.iloc[27:47,4]
        provider5=df.iloc[27:47,5]
        provider6=df.iloc[27:47,6]
        provider7=df.iloc[27:47,7]

        week2day1['provider']=provider1
        week2day2['provider']=provider2
        week2day3['provider']=provider3
        week2day4['provider']=provider4
        week2day5['provider']=provider5
        week2day6['provider']=provider6
        week2day7['provider']=provider7

        week2day1['clinic']="AAC"
        week2day2['clinic']="AAC"
        week2day3['clinic']="AAC"
        week2day4['clinic']="AAC"
        week2day5['clinic']="AAC"
        week2day6['clinic']="AAC"
        week2day7['clinic']="AAC"

        week2=pd.DataFrame(columns=week2day1.columns)
        week2=pd.concat([week2,week2day1,week2day2,week2day3,week2day4,week2day5,week2day6,week2day7])
        week2.to_csv('week2.csv',index=False)

        clinictype=df.iloc[51:71, 0:1]
        c1 = pd.DataFrame(clinictype, columns = ['type'])
        c2 = pd.DataFrame(clinictype, columns = ['type'])
        c3 = pd.DataFrame(clinictype, columns = ['type'])
        c4 = pd.DataFrame(clinictype, columns = ['type'])
        c5 = pd.DataFrame(clinictype, columns = ['type'])
        c6 = pd.DataFrame(clinictype, columns = ['type'])
        c7 = pd.DataFrame(clinictype, columns = ['type'])

        c1['type']=clinictype
        c2['type']=clinictype
        c3['type']=clinictype
        c4['type']=clinictype
        c5['type']=clinictype
        c6['type']=clinictype
        c7['type']=clinictype

        week3day1=c1.replace(to_replace=r'- Continuity', value='', regex=True)
        week3day2=c2.replace(to_replace=r'- Continuity', value='', regex=True)
        week3day3=c3.replace(to_replace=r'- Continuity', value='', regex=True)
        week3day4=c4.replace(to_replace=r'- Continuity', value='', regex=True)
        week3day5=c5.replace(to_replace=r'- Continuity', value='', regex=True)
        week3day6=c6.replace(to_replace=r'- Continuity', value='', regex=True)
        week3day7=c7.replace(to_replace=r'- Continuity', value='', regex=True)

        day1=df.iloc[49,1]
        day2=df.iloc[49,2]
        day3=df.iloc[49,3]
        day4=df.iloc[49,4]
        day5=df.iloc[49,5]
        day6=df.iloc[49,6]
        day7=df.iloc[49,7]

        week3day1['date']=day1
        week3day2['date']=day2
        week3day3['date']=day3
        week3day4['date']=day4
        week3day5['date']=day5
        week3day6['date']=day6
        week3day7['date']=day7

        provider1=df.iloc[51:71,1]
        provider2=df.iloc[51:71,2]
        provider3=df.iloc[51:71,3]
        provider4=df.iloc[51:71,4]
        provider5=df.iloc[51:71,5]
        provider6=df.iloc[51:71,6]
        provider7=df.iloc[51:71,7]

        week3day1['provider']=provider1
        week3day2['provider']=provider2
        week3day3['provider']=provider3
        week3day4['provider']=provider4
        week3day5['provider']=provider5
        week3day6['provider']=provider6
        week3day7['provider']=provider7

        week3day1['clinic']="AAC"
        week3day2['clinic']="AAC"
        week3day3['clinic']="AAC"
        week3day4['clinic']="AAC"
        week3day5['clinic']="AAC"
        week3day6['clinic']="AAC"
        week3day7['clinic']="AAC"

        week3=pd.DataFrame(columns=week3day1.columns)
        week3=pd.concat([week3,week3day1,week3day2,week3day3,week3day4,week3day5,week3day6,week3day7])
        week3.to_csv('week3.csv',index=False)

        clinictype=df.iloc[75:95, 0:1]
        d1 = pd.DataFrame(clinictype, columns = ['type'])
        d2 = pd.DataFrame(clinictype, columns = ['type'])
        d3 = pd.DataFrame(clinictype, columns = ['type'])
        d4 = pd.DataFrame(clinictype, columns = ['type'])
        d5 = pd.DataFrame(clinictype, columns = ['type'])
        d6 = pd.DataFrame(clinictype, columns = ['type'])
        d7 = pd.DataFrame(clinictype, columns = ['type'])

        d1['type']=clinictype
        d2['type']=clinictype
        d3['type']=clinictype
        d4['type']=clinictype
        d5['type']=clinictype
        d6['type']=clinictype
        d7['type']=clinictype

        week4day1=d1.replace(to_replace=r'- Continuity', value='', regex=True)
        week4day2=d2.replace(to_replace=r'- Continuity', value='', regex=True)
        week4day3=d3.replace(to_replace=r'- Continuity', value='', regex=True)
        week4day4=d4.replace(to_replace=r'- Continuity', value='', regex=True)
        week4day5=d5.replace(to_replace=r'- Continuity', value='', regex=True)
        week4day6=d6.replace(to_replace=r'- Continuity', value='', regex=True)
        week4day7=d7.replace(to_replace=r'- Continuity', value='', regex=True)


        day1=df.iloc[73,1]
        day2=df.iloc[73,2]
        day3=df.iloc[73,3]
        day4=df.iloc[73,4]
        day5=df.iloc[73,5]
        day6=df.iloc[73,6]
        day7=df.iloc[73,7]

        week4day1['date']=day1
        week4day2['date']=day2
        week4day3['date']=day3
        week4day4['date']=day4
        week4day5['date']=day5
        week4day6['date']=day6
        week4day7['date']=day7

        provider1=df.iloc[75:95,1]
        provider2=df.iloc[75:95,2]
        provider3=df.iloc[75:95,3]
        provider4=df.iloc[75:95,4]
        provider5=df.iloc[75:95,5]
        provider6=df.iloc[75:95,6]
        provider7=df.iloc[75:95,7]

        week4day1['provider']=provider1
        week4day2['provider']=provider2
        week4day3['provider']=provider3
        week4day4['provider']=provider4
        week4day5['provider']=provider5
        week4day6['provider']=provider6
        week4day7['provider']=provider7

        week4day1['clinic']="AAC"
        week4day2['clinic']="AAC"
        week4day3['clinic']="AAC"
        week4day4['clinic']="AAC"
        week4day5['clinic']="AAC"
        week4day6['clinic']="AAC"
        week4day7['clinic']="AAC"

        week4=pd.DataFrame(columns=week4day1.columns)
        week4=pd.concat([week4,week4day1,week4day2,week4day3,week4day4,week4day5,week4day6,week4day7])
        week4.to_csv('week4.csv',index=False)

        hope=pd.DataFrame(columns=week1.columns)
        hope=pd.concat([hope,week1,week2,week3,week4])
        hope.to_csv('mhss.csv',index=False)

        hope['H'] = "H"
        extrai = hope[hope['type'] == 'AM'].copy()  # Ensure we're working with a copy
        extrai.loc[:, 'count'] = extrai.groupby(['date'])['provider'].cumcount() + 0  # Starts at H0 for AM
        extrai.loc[:, 'class'] = "H" + extrai['count'].astype(str)
        extrai = extrai.loc[:, ('date', 'type', 'provider', 'clinic', 'class')]
        extrai.to_csv('3.csv', index=False)

        # Handle PM Continuity for EXTRAI
        hope['H'] = "H"
        extraii = hope[hope['type'] == 'PM'].copy()  # Ensure we're working with a copy
        extraii.loc[:, 'count'] = extraii.groupby(['date'])['provider'].cumcount() + 10  # Starts at H10 for PM
        extraii.loc[:, 'class'] = "H" + extraii['count'].astype(str)
        extraii = extraii.loc[:, ('date', 'type', 'provider', 'clinic', 'class')]
        extraii.to_csv('4.csv', index=False)

        extras=pd.DataFrame(columns=extrai.columns)
        extras=pd.concat([extrai,extraii])
        extras.to_csv('mhss.csv',index=False)
        
        df1=pd.read_csv('etowns.csv')
        df2=pd.read_csv('hopes.csv')
        df3=pd.read_csv('nyess.csv')
        df4=pd.read_csv('extras.csv')
        df5=pd.read_csv('mhss.csv')

        dfx=pd.DataFrame(columns=df1.columns)
        dfx=pd.concat([dfx,df1,df2,df3,df4,df5])
        dfx['providers']=dfx['provider'].str.split('~').str[0]
        dfx['student']=dfx['provider'].str.split('~').str[1]
        dfx1=dfx[['date','type','providers','student','clinic','provider','class']]

        dfx1['date'] = pd.to_datetime(dfx1['date'])
        dfx1['date'] = dfx1['date'].dt.strftime('%m/%d/%Y')

        mydict = {}
        with open('xxxDATEMAP.csv', mode='r')as inp:     #file is the objects you want to map. I want to map the IMP in this file to diagnosis.csv
            reader = csv.reader(inp)
            df1 = {rows[0]:rows[1] for rows in reader} 
        dfx1['datecode'] = dfx1.date.map(df1)               #'type' is the new column in the diagnosis file. 'encounter_id' is the key you are using to MAP 

        dfx1.to_excel('STUDENTLIST.xlsx',index=False)
        dfx1.to_csv('STUDENTLIST.csv',index=False)

        dfx1['type'] = dfx1['type'].str.lstrip()
        dfx1['type'] = dfx1['type'].str.rstrip()

        dfx1['providers'] = dfx1['providers'].str.lstrip()
        dfx1['providers'] = dfx1['providers'].str.rstrip()

        dfx1['student'] = dfx1['student'].str.lstrip()
        dfx1['student'] = dfx1['student'].str.rstrip()

        dfs = [["AM","S1"],["PM","S2"],["AM - ACUTES","S11"],["PM - ACUTES","S12"]]
        dftype = pd.DataFrame(dfs, columns = ['type', 'datecode2'])
        dftype.to_csv('dftype.csv',index=False)

        mydict = {}
        with open('dftype.csv', mode='r')as inp:     #file is the objects you want to map. I want to map the IMP in this file to diagnosis.csv
            reader = csv.reader(inp)
            df1 = {rows[0]:rows[1] for rows in reader} 
        dfx1['datecode2'] = dfx1.type.map(df1)               #'type' is the new column in the diagnosis file. 'encounter_id' is the key you are using to MAP 

        dfx1.to_excel('PALIST.xlsx',index=False)

        dfx1.to_csv('PALIST.csv',index=False)

        dfx1 = pd.read_csv('PALIST.csv')

        new_row = pd.DataFrame({'date':0, 'type':0, 'providers':0,
                                'student':0, 'clinic':0, 'provider':0,
                                'class':0, 'datecode':0, 'datecode2':0},
                                                                    index =[0])
        # simply concatenate both dataframes
        df = pd.concat([new_row, dfx1]).reset_index(drop = True)

        df.to_csv('PALIST.csv',index=False)
        
        st.dataframe(df.head())
        
        df = pd.read_excel('Book4.xlsx')

        # Keep the first NaN in the first row as it is for the first column
        df.iloc[0, 0] = np.nan  # Ensure the first cell is NaN (or leave it as it is)

        # Initialize a counter for BLANK replacements
        blank_counter = 1

        # Replace NaN values in the first column with BLANK1, BLANK2, etc. starting from the second row
        for i in range(1, len(df)):
            if pd.isna(df.iloc[i, 0]):
                df.iloc[i, 0] = f'BLANK{blank_counter}'
                blank_counter += 1

        x = (df.loc[0, 'Week 1'])

        x = x.strftime("%m/%d/%Y")

        test_date = datetime.datetime.strptime(x, "%m/%d/%Y")

        # initializing K
        K = 28

        res = []

        for day in range(K):
            date = (test_date + datetime.timedelta(days = day)).strftime("%-m/%-d/%Y")
            res.append(date)

        res

        dates = pd.DataFrame(res, columns =['dates'])

        dates['x'] = "y"

        dates['i'] = dates.index+1

        dates['i2']= dates.index+0

        dates['T'] = "T" + dates['i2'].astype(str)

        dates['x'] = dates['x'].astype(str)+dates['i'].astype(str)

        dates['x'] = dates['x'].astype(str) + "=" + "'"+dates['dates'].astype(str) + "'"

        dateT = dates[['dates','T']]

        dates = dates[['x']]

        dates.to_csv('dates.csv',index=False)

        dateT.to_csv('datesT.csv',index=False)

        import numpy as np

        datesdf = pd.read_csv('dates.csv')

        dates = datesdf['x'].astype(str)

        numpy_array=dates.to_numpy()
        np.savetxt("dates.py",numpy_array, fmt="%s")

        exec(open('dates.py').read())

        #################################STUDENT NAME and CLINICAL EXPERIENCES INPUT#################################
        column = "Student Name:"
        L=df[(column)].unique().tolist() #Must Use Unique
        xf200 = pd.DataFrame({'col':L})
        xf200['i'] = xf200.index

        column = "Week 1"
        L=df[(column)].tolist() #No Unique Required 
        xf201 = pd.DataFrame({'col':L})
        xf201['i'] = xf201.index

        column = "Week 2"
        L=df[(column)].tolist()
        xf202 = pd.DataFrame({'col':L})
        xf202['i'] = xf202.index

        column = "Week 3"
        L=df[(column)].tolist()
        xf203 = pd.DataFrame({'col':L})
        xf203['i'] = xf203.index

        column = "Week 4"
        L=df[(column)].tolist()
        xf204 = pd.DataFrame({'col':L})
        xf204['i'] = xf204.index

        ######################################################################################################################
        # Create a workbook and add the main worksheet
        # Create a workbook and add the main worksheet
        workbook = xlsxwriter.Workbook('Main_Schedule_MS.xlsx')

        # Define the formats
        format1 = workbook.add_format({
            'font_size': 14,
            'bold': 1,
            'align': 'center',
            'valign': 'vcenter',
            'color': 'black',
            'text_wrap': True,
            'bg_color': '#FEFFCC',
            'border': 1
        })

        format2 = workbook.add_format({
            'font_size': 10,
            'bold': 1,
            'align': 'center',
            'valign': 'vcenter',
            'color': 'yellow',
            'bg_color': 'black',
            'border': 1,
            'text_wrap': True
        })

        format3 = workbook.add_format({
            'font_size':12,
            'bold': 1,
            'align': 'center',
            'valign': 'vcenter',
            'color':'black',
            'bg_color':'#FFC7CE',
            'border':1
        })

        format4 = workbook.add_format({'num_format':'mm/dd/yyyy','font_size':12,'bold': 1,'align': 'center','valign': 'vcenter','color':'black','bg_color':'#F4F6F7','border':1})
        format5 = workbook.add_format({'font_size':12,'bold': 1,'align': 'center','valign': 'vcenter','color':'black','bg_color':'#F4F6F7','border':1})
        format6 = workbook.add_format({'bg_color':'black','border':1})
        format7 = workbook.add_format({'font_size':12,'bold': 1,'align': 'center','valign': 'vcenter','color':'black','bg_color':'#90EE90','border':1})
        format8 = workbook.add_format({'font_size':12,'bold': 1,'align': 'center','valign':'vcenter','color':'black','bg_color':'#89CFF0','border':1})
        # Initialize a list to store worksheets
        worksheets = []

        # Add a worksheet for each unique name in the DataFrame
        for name in xf200['col']:
            # Ensure that the name is a valid string (i.e., no NaN or floats)
            if isinstance(name, str):  # Check if it's a string
                name = name[:31]  # Truncate the name if it's longer than 31 characters
            else:
                continue  # Skip invalid entries (like NaN or numbers)

            # Ensure the name is not an empty string
            if len(name) > 0:
                worksheet = workbook.add_worksheet(name)
                worksheets.append(worksheet)  # Store the worksheet in the list

                # Merge range for the student name column
                worksheet.merge_range('A1:A2', 'Student Name:', format1)
                worksheet.merge_range('B1:B2', name, format1)

                # Merge range for the note
                note = '*Note** Asynchronous time is for coursework only. During this time period, we expect students to do coursework, be available for any additional educational activities, and any extra clinical time that may be available. If the student is not available during this time period and has not made an absence request, the student will be cited for unprofessionalism and will risk failing the course.'
                worksheet.merge_range('C1:H2', note, format2)

                # Set column widths (you can set them all at once or in a loop)
                worksheet.set_column('A:A', 20)
                worksheet.set_column('B:B', 47)
                worksheet.set_column('C:H', 47)

                # Set row height for header row
                worksheet.set_row(0, 37.25)

                days_of_week = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']

                # Start row positions for each week
                start_rows = [3, 11, 19, 27]

                # Loop through each week's starting row and write the days
                for start_row in start_rows:
                    for col, day in enumerate(days_of_week, start=1):  
                        worksheet.write(chr(65 + col) + str(start_row), day, format3)

                # List of weeks
                weeks = ['Week 1', 'Week 2', 'Week 3', 'Week 4']

                for week_index, week in enumerate(weeks):
                    worksheet.write(f'A{4 + (week_index * 8)}', week, format3)

                for week_index in range(4):  # For each week (4 weeks)
                    row = 6 + (week_index * 8)  # Calculate row for AM (6, 14, 22, 30)
                    worksheet.write(f'A{row}', 'AM', format3)

                for week_index in range(4):  # For each week (4 weeks)
                    row = 7 + (week_index * 8)  # Calculate row for PM (7, 15, 23, 31)
                    worksheet.write(f'A{row}', 'PM', format3)

                column_idx = 1  # Starting at column B (Excel columns are 1-indexed)
                date_idx = 0   # Start from the first date in `res`

                # Loop through the weeks and the corresponding dates
                for week_row in start_rows:
                    for col in range(7):  # Loop through Monday to Sunday (7 days)
                        if date_idx < len(res):
                            date = res[date_idx]
                            worksheet.write(week_row, column_idx + col, date, format4)
                            date_idx += 1  # Move to the next date

                worksheet.set_zoom(70)
                worksheet.merge_range('A10:H10','',format6)
                worksheet.merge_range('A18:H18','',format6)
                worksheet.merge_range('A26:H26','',format6)
                worksheet.merge_range('A34:H34','',format6)

                worksheet.write('A9',' ', format7)
                worksheet.write('B9',' ', format7)
                worksheet.write('C9',' ', format7)
                worksheet.write('D9',' ', format7)
                worksheet.write('E9',' ', format7)
                worksheet.write('F9',' ', format7)
                worksheet.write('G9',' ', format7)
                worksheet.write('H9',' ', format7)

                worksheet.write('A17',' ', format7)
                worksheet.write('B17',' ', format7)
                worksheet.write('C17',' ', format7)
                worksheet.write('D17',' ', format7)
                worksheet.write('E17',' ', format7)
                worksheet.write('F17',' ', format7)
                worksheet.write('G17',' ', format7)
                worksheet.write('H17',' ', format7)

                worksheet.write('A25',' ', format7)
                worksheet.write('B25',' ', format7)
                worksheet.write('C25',' ', format7)
                worksheet.write('D25',' ', format7)
                worksheet.write('E25',' ', format7)
                worksheet.write('F25',' ', format7)
                worksheet.write('G25',' ', format7)
                worksheet.write('H25',' ', format7)

                worksheet.write('A33',' ', format7)
                worksheet.write('B33',' ', format7)
                worksheet.write('C33',' ', format7)
                worksheet.write('D33',' ', format7)
                worksheet.write('E33',' ', format7)
                worksheet.write('F33',' ', format7)
                worksheet.write('G33',' ', format7)
                worksheet.write('H33',' ', format7)

                # Writing to row 8
                worksheet.write('A8', 'ASSIGNMENT DUE:', format8)
                worksheet.write('B8', ' ', format8)
                worksheet.write('C8', ' ', format8)
                worksheet.write('D8', 'Quiz 1 Due', format8)
                worksheet.write('E8', ' ', format8)
                worksheet.write('F8', 'Ask for Feedback!', format8)
                worksheet.write('G8', ' ', format8)
                worksheet.write('H8', 'Quiz 2 Due', format8)

                # Writing to row 16
                worksheet.write('A16', 'ASSIGNMENT DUE:', format8)
                worksheet.write('B16', ' ', format8)
                worksheet.write('C16', ' ', format8)
                worksheet.write('D16', ' ', format8)
                worksheet.write('E16', ' ', format8)
                worksheet.write('F16', 'Ask for Feedback!', format8)
                worksheet.write('G16', ' ', format8)
                worksheet.write('H16', 'Quiz 3 Due', format8)

                # Writing to row 24
                worksheet.write('A24', 'ASSIGNMENT DUE:', format8)
                worksheet.write('B24', ' ', format8)
                worksheet.write('C24', ' ', format8)
                worksheet.write('D24', ' ', format8)
                worksheet.write('E24', ' ', format8)
                worksheet.write('F24', 'Ask for Feedback!', format8)
                worksheet.write('G24', ' ', format8)
                worksheet.write('H24', 'Quiz 4 Due', format8)

                # Writing to row 32
                worksheet.write('A32', 'ASSIGNMENT DUE:', format8)
                worksheet.write('B32', ' ', format8)
                worksheet.write('C32', ' ', format8)
                worksheet.write('D32', ' ', format8)
                worksheet.write('E32', ' ', format8)
                worksheet.write('F32', 'Write Up, Developmental Assessment of Pediatric Patient, Clinical Encounters are Due!', format8)
                worksheet.write('G32', ' ', format8)
                worksheet.write('H32', ' ', format8)




        # Now, we'll write the location data
        locations = xf201['col'].tolist()[1:]  # Assuming 'col' is a column in xf201 dataframe

        # Ensure locations match the number of worksheets
        if len(locations) != len(worksheets):
            raise ValueError(f"Number of locations ({len(locations)}) does not match number of worksheets ({len(worksheets)})")

        # Iterate over the names and locations and write the data
        for i, (name, location) in enumerate(zip(xf200['col'], locations)):
            worksheet = worksheets[i]  # Get the corresponding worksheet for this index

            # Write the location data to rows 6 and 7 for each worksheet
            for row in range(6, 8):  # Rows 6 and 7
                worksheet.write(f'B{row}', location, format5)
                worksheet.write(f'C{row}', location, format5)
                worksheet.write(f'D{row}', location, format5)
                worksheet.write(f'E{row}', location, format5)
                worksheet.write(f'F{row}', location, format5)
                worksheet.write(f'G{row}', "OFF", format5)
                worksheet.write(f'H{row}', "OFF", format5)

        # Now, we'll write the location data
        locations = xf202['col'].tolist()[1:]  # Assuming 'col' is a column in xf201 dataframe

        # Ensure locations match the number of worksheets
        if len(locations) != len(worksheets):
            raise ValueError(f"Number of locations ({len(locations)}) does not match number of worksheets ({len(worksheets)})")

        # Iterate over the names and locations and write the data
        for i, (name, location) in enumerate(zip(xf200['col'], locations)):
            worksheet = worksheets[i]  # Get the corresponding worksheet for this index

            for row in range(14, 16): 
                worksheet.write(f'B{row}', location, format5)
                worksheet.write(f'C{row}', location, format5)
                worksheet.write(f'D{row}', location, format5)
                worksheet.write(f'E{row}', location, format5)
                worksheet.write(f'F{row}', location, format5)
                worksheet.write(f'G{row}', "OFF", format5)
                worksheet.write(f'H{row}', "OFF", format5)

        # Now, we'll write the location data
        locations = xf203['col'].tolist()[1:]  # Assuming 'col' is a column in xf201 dataframe

        # Ensure locations match the number of worksheets
        if len(locations) != len(worksheets):
            raise ValueError(f"Number of locations ({len(locations)}) does not match number of worksheets ({len(worksheets)})")

        # Iterate over the names and locations and write the data
        for i, (name, location) in enumerate(zip(xf200['col'], locations)):
            worksheet = worksheets[i]  # Get the corresponding worksheet for this index

            for row in range(22, 24): 
                worksheet.write(f'B{row}', location, format5)
                worksheet.write(f'C{row}', location, format5)
                worksheet.write(f'D{row}', location, format5)
                worksheet.write(f'E{row}', location, format5)
                worksheet.write(f'F{row}', location, format5)
                worksheet.write(f'G{row}', "OFF", format5)
                worksheet.write(f'H{row}', "OFF", format5)

        # Now, we'll write the location data
        locations = xf204['col'].tolist()[1:]  # Assuming 'col' is a column in xf201 dataframe

        # Ensure locations match the number of worksheets
        if len(locations) != len(worksheets):
            raise ValueError(f"Number of locations ({len(locations)}) does not match number of worksheets ({len(worksheets)})")

        # Iterate over the names and locations and write the data
        for i, (name, location) in enumerate(zip(xf200['col'], locations)):
            worksheet = worksheets[i]  # Get the corresponding worksheet for this index

            for row in range(30, 32): 
                worksheet.write(f'B{row}', location, format5)
                worksheet.write(f'C{row}', location, format5)
                worksheet.write(f'D{row}', location, format5)
                worksheet.write(f'E{row}', location, format5)
                worksheet.write(f'F{row}', location, format5)
                worksheet.write(f'G{row}', "OFF", format5)
                worksheet.write(f'H{row}', "OFF", format5)


        # Close the workbook
        workbook.close()
        
        df = pd.read_csv('datesT.csv')
        dx = df

        df["dates"] = pd.to_datetime(df["dates"])
        df['dates'] = df['dates'].dt.strftime('%m/%d/%Y')

        df.to_csv('datesT.csv',index=False)

        df=pd.read_csv('PALIST.csv',dtype=str)

        df['text'] = df['providers'] + " - " + "[" + df['clinic'] + "]"
        df = df[['datecode','type','student','text','date','clinic']]


        mydict = {}
        with open('datesT.csv', mode='r')as inp:     #file is the objects you want to map. I want to map the IMP in this file to diagnosis.csv
            reader = csv.reader(inp)
            df1 = {rows[0]:rows[1] for rows in reader} 
        df['datecode'] = df.date.map(df1)

        df = df[~df['student'].isnull()]

        df = df.loc[df['student'] != "0"]

        df.to_excel('Source1.xlsx', index=False)

        import openpyxl
        import numpy as np 

        column = "student"
        L=df[(column)].tolist() #Must Use Unique
        xf200 = pd.DataFrame({'col':L})
        xf200['i'] = xf200.index
        xf200['blank'] = ''
        
        wb = openpyxl.load_workbook('Source1.xlsx')
        ws = wb['Sheet1']
        wb1 = openpyxl.load_workbook('Main_Schedule_MS.xlsx')

        # Assuming xf200 is a pandas DataFrame loaded from somewhere, for example:
        # xf200 = pd.read_csv('your_dataframe.csv')  # Or load your DataFrame as needed

        # Iterate over the names in the xf200 DataFrame
        for name in xf200['col']:  # Assuming 'names' is the column with the names you want to look up
            # Check if the sheet with the name exists in the workbook
            if name in wb1.sheetnames:
                ws1 = wb1[name]  # Access the sheet directly since it exists
            else:
                print(f"Sheet for {name} not found, skipping.")
                continue  # Skip to the next name if the sheet does not exist

            for row in ws.iter_rows():
                if row[2].value == name:  # Check if the name matches (assuming name is in column 3)
                    # Handle T0 to T6
                    if row[0].value == "T0":
                        if row[1].value == "AM":
                            ws1.cell(row=6, column=2).value = row[3].value  # AM, T0
                        elif row[1].value == "PM":
                            ws1.cell(row=7, column=2).value = row[3].value  # PM, T0
                    elif row[0].value == "T1":
                        if row[1].value == "AM":
                            ws1.cell(row=6, column=3).value = row[3].value  # AM, T1
                        elif row[1].value == "PM":
                            ws1.cell(row=7, column=3).value = row[3].value  # PM, T1
                    elif row[0].value == "T2":
                        if row[1].value == "AM":
                            ws1.cell(row=6, column=4).value = row[3].value  # AM, T2
                        elif row[1].value == "PM":
                            ws1.cell(row=7, column=4).value = row[3].value  # PM, T2
                    elif row[0].value == "T3":
                        if row[1].value == "AM":
                            ws1.cell(row=6, column=5).value = row[3].value  # AM, T3
                        elif row[1].value == "PM":
                            ws1.cell(row=7, column=5).value = row[3].value  # PM, T3
                    elif row[0].value == "T4":
                        if row[1].value == "AM":
                            ws1.cell(row=6, column=6).value = row[3].value  # AM, T4
                        elif row[1].value == "PM":
                            ws1.cell(row=7, column=6).value = row[3].value  # PM, T4
                    elif row[0].value == "T5":
                        if row[1].value == "AM":
                            ws1.cell(row=6, column=7).value = row[3].value  # AM, T5
                        elif row[1].value == "PM":
                            ws1.cell(row=7, column=7).value = row[3].value  # PM, T5
                    elif row[0].value == "T6":
                        if row[1].value == "AM":
                            ws1.cell(row=6, column=8).value = row[3].value  # AM, T6
                        elif row[1].value == "PM":
                            ws1.cell(row=7, column=8).value = row[3].value  # PM, T6

                    # Handle T7 to T13
                    elif row[0].value == "T7":
                        if row[1].value == "AM":
                            ws1.cell(row=14, column=2).value = row[3].value  # AM, T7
                        elif row[1].value == "PM":
                            ws1.cell(row=15, column=2).value = row[3].value  # PM, T7
                    elif row[0].value == "T8":
                        if row[1].value == "AM":
                            ws1.cell(row=14, column=3).value = row[3].value  # AM, T8
                        elif row[1].value == "PM":
                            ws1.cell(row=15, column=3).value = row[3].value  # PM, T8
                    elif row[0].value == "T9":
                        if row[1].value == "AM":
                            ws1.cell(row=14, column=4).value = row[3].value  # AM, T9
                        elif row[1].value == "PM":
                            ws1.cell(row=15, column=4).value = row[3].value  # PM, T9
                    elif row[0].value == "T10":
                        if row[1].value == "AM":
                            ws1.cell(row=14, column=5).value = row[3].value  # AM, T10
                        elif row[1].value == "PM":
                            ws1.cell(row=15, column=5).value = row[3].value  # PM, T10
                    elif row[0].value == "T11":
                        if row[1].value == "AM":
                            ws1.cell(row=14, column=6).value = row[3].value  # AM, T11
                        elif row[1].value == "PM":
                            ws1.cell(row=15, column=6).value = row[3].value  # PM, T11
                    elif row[0].value == "T12":
                        if row[1].value == "AM":
                            ws1.cell(row=14, column=7).value = row[3].value  # AM, T12
                        elif row[1].value == "PM":
                            ws1.cell(row=15, column=7).value = row[3].value  # PM, T12
                    elif row[0].value == "T13":
                        if row[1].value == "AM":
                            ws1.cell(row=14, column=8).value = row[3].value  # AM, T13
                        elif row[1].value == "PM":
                            ws1.cell(row=15, column=8).value = row[3].value  # PM, T13

                    # Handle T14 to T20
                    elif row[0].value == "T14":
                        if row[1].value == "AM":
                            ws1.cell(row=22, column=2).value = row[3].value  # AM, T14
                        elif row[1].value == "PM":
                            ws1.cell(row=23, column=2).value = row[3].value  # PM, T14
                    elif row[0].value == "T15":
                        if row[1].value == "AM":
                            ws1.cell(row=22, column=3).value = row[3].value  # AM, T15
                        elif row[1].value == "PM":
                            ws1.cell(row=23, column=3).value = row[3].value  # PM, T15
                    elif row[0].value == "T16":
                        if row[1].value == "AM":
                            ws1.cell(row=22, column=4).value = row[3].value  # AM, T16
                        elif row[1].value == "PM":
                            ws1.cell(row=23, column=4).value = row[3].value  # PM, T16
                    elif row[0].value == "T17":
                        if row[1].value == "AM":
                            ws1.cell(row=22, column=5).value = row[3].value  # AM, T17
                        elif row[1].value == "PM":
                            ws1.cell(row=23, column=5).value = row[3].value  # PM, T17
                    elif row[0].value == "T18":
                        if row[1].value == "AM":
                            ws1.cell(row=22, column=6).value = row[3].value  # AM, T18
                        elif row[1].value == "PM":
                            ws1.cell(row=23, column=6).value = row[3].value  # PM, T18
                    elif row[0].value == "T19":
                        if row[1].value == "AM":
                            ws1.cell(row=22, column=7).value = row[3].value  # AM, T19
                        elif row[1].value == "PM":
                            ws1.cell(row=23, column=7).value = row[3].value  # PM, T19
                    elif row[0].value == "T20":
                        if row[1].value == "AM":
                            ws1.cell(row=22, column=8).value = row[3].value  # AM, T20
                        elif row[1].value == "PM":
                            ws1.cell(row=23, column=8).value = row[3].value  # PM, T20

                    # Handle T21 to T27
                    elif row[0].value == "T21":
                        if row[1].value == "AM":
                            ws1.cell(row=30, column=2).value = row[3].value  # AM, T21
                        elif row[1].value == "PM":
                            ws1.cell(row=31, column=2).value = row[3].value  # PM, T21
                    elif row[0].value == "T22":
                        if row[1].value == "AM":
                            ws1.cell(row=30, column=3).value = row[3].value  # AM, T22
                        elif row[1].value == "PM":
                            ws1.cell(row=31, column=3).value = row[3].value  # PM, T22
                    elif row[0].value == "T23":
                        if row[1].value == "AM":
                            ws1.cell(row=30, column=4).value = row[3].value  # AM, T23
                        elif row[1].value == "PM":
                            ws1.cell(row=31, column=4).value = row[3].value  # PM, T23
                    elif row[0].value == "T24":
                        if row[1].value == "AM":
                            ws1.cell(row=30, column=5).value = row[3].value  # AM, T24
                        elif row[1].value == "PM":
                            ws1.cell(row=31, column=5).value = row[3].value  # PM, T24
                    elif row[0].value == "T25":
                        if row[1].value == "AM":
                            ws1.cell(row=30, column=6).value = row[3].value  # AM, T25
                        elif row[1].value == "PM":
                            ws1.cell(row=31, column=6).value = row[3].value  # PM, T25
                    elif row[0].value == "T26":
                        if row[1].value == "AM":
                            ws1.cell(row=30, column=7).value = row[3].value  # AM, T26
                        elif row[1].value == "PM":
                            ws1.cell(row=31, column=7).value = row[3].value  # PM, T26
                    elif row[0].value == "T27":
                        if row[1].value == "AM":
                            ws1.cell(row=30, column=8).value = row[3].value  # AM, T27
                        elif row[1].value == "PM":
                            ws1.cell(row=31, column=8).value = row[3].value  # PM, T27

            for row in ws.iter_rows():
                if row[2].value == name:  # Check if the name matches (assuming name is in column 3)

                    # Handle T0 to T6
                    if row[0].value == "T0":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=6, column=2).value = row[3].value  # AM - ACUTES, T0
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=7, column=2).value = row[3].value  # PM - ACUTES, T0
                    elif row[0].value == "T1":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=6, column=3).value = row[3].value  # AM - ACUTES, T1
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=7, column=3).value = row[3].value  # PM - ACUTES, T1
                    elif row[0].value == "T2":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=6, column=4).value = row[3].value  # AM - ACUTES, T2
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=7, column=4).value = row[3].value  # PM - ACUTES, T2
                    elif row[0].value == "T3":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=6, column=5).value = row[3].value  # AM - ACUTES, T3
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=7, column=5).value = row[3].value  # PM - ACUTES, T3
                    elif row[0].value == "T4":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=6, column=6).value = row[3].value  # AM - ACUTES, T4
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=7, column=6).value = row[3].value  # PM - ACUTES, T4
                    elif row[0].value == "T5":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=6, column=7).value = row[3].value  # AM - ACUTES, T5
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=7, column=7).value = row[3].value  # PM - ACUTES, T5
                    elif row[0].value == "T6":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=6, column=8).value = row[3].value  # AM - ACUTES, T6
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=7, column=8).value = row[3].value  # PM - ACUTES, T6

                    # Handle T7 to T13
                    elif row[0].value == "T7":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=14, column=2).value = row[3].value  # AM - ACUTES, T7
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=15, column=2).value = row[3].value  # PM - ACUTES, T7
                    elif row[0].value == "T8":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=14, column=3).value = row[3].value  # AM - ACUTES, T8
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=15, column=3).value = row[3].value  # PM - ACUTES, T8
                    elif row[0].value == "T9":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=14, column=4).value = row[3].value  # AM - ACUTES, T9
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=15, column=4).value = row[3].value  # PM - ACUTES, T9
                    elif row[0].value == "T10":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=14, column=5).value = row[3].value  # AM - ACUTES, T10
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=15, column=5).value = row[3].value  # PM - ACUTES, T10
                    elif row[0].value == "T11":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=14, column=6).value = row[3].value  # AM - ACUTES, T11
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=15, column=6).value = row[3].value  # PM - ACUTES, T11
                    elif row[0].value == "T12":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=14, column=7).value = row[3].value  # AM - ACUTES, T12
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=15, column=7).value = row[3].value  # PM - ACUTES, T12
                    elif row[0].value == "T13":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=14, column=8).value = row[3].value  # AM - ACUTES, T13
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=15, column=8).value = row[3].value  # PM - ACUTES, T13

                    # Handle T14 to T20
                    elif row[0].value == "T14":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=22, column=2).value = row[3].value  # AM - ACUTES, T14
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=23, column=2).value = row[3].value  # PM - ACUTES, T14
                    elif row[0].value == "T15":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=22, column=3).value = row[3].value  # AM - ACUTES, T15
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=23, column=3).value = row[3].value  # PM - ACUTES, T15
                    elif row[0].value == "T16":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=22, column=4).value = row[3].value  # AM - ACUTES, T16
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=23, column=4).value = row[3].value  # PM - ACUTES, T16
                    elif row[0].value == "T17":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=22, column=5).value = row[3].value  # AM - ACUTES, T17
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=23, column=5).value = row[3].value  # PM - ACUTES, T17
                    elif row[0].value == "T18":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=22, column=6).value = row[3].value  # AM - ACUTES, T18
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=23, column=6).value = row[3].value  # PM - ACUTES, T18
                    elif row[0].value == "T19":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=22, column=7).value = row[3].value  # AM - ACUTES, T19
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=23, column=7).value = row[3].value  # PM - ACUTES, T19
                    elif row[0].value == "T20":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=22, column=8).value = row[3].value  # AM - ACUTES, T20
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=23, column=8).value = row[3].value  # PM - ACUTES, T20

                    # Handle T21 to T27
                    elif row[0].value == "T21":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=30, column=2).value = row[3].value  # AM - ACUTES, T21
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=31, column=2).value = row[3].value  # PM - ACUTES, T21
                    elif row[0].value == "T22":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=30, column=3).value = row[3].value  # AM - ACUTES, T22
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=31, column=3).value = row[3].value  # PM - ACUTES, T22
                    elif row[0].value == "T23":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=30, column=4).value = row[3].value  # AM - ACUTES, T23
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=31, column=4).value = row[3].value  # PM - ACUTES, T23
                    elif row[0].value == "T24":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=30, column=5).value = row[3].value  # AM - ACUTES, T24
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=31, column=5).value = row[3].value  # PM - ACUTES, T24
                    elif row[0].value == "T25":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=30, column=6).value = row[3].value  # AM - ACUTES, T25
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=31, column=6).value = row[3].value  # PM - ACUTES, T25
                    elif row[0].value == "T26":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=30, column=7).value = row[3].value  # AM - ACUTES, T26
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=31, column=7).value = row[3].value  # PM - ACUTES, T26
                    elif row[0].value == "T27":
                        if row[1].value == "AM - ACUTES":
                            ws1.cell(row=30, column=8).value = row[3].value  # AM - ACUTES, T27
                        elif row[1].value == "PM - ACUTES":
                            ws1.cell(row=31, column=8).value = row[3].value  # PM - ACUTES, T27

        # Save the modified workbook
        #wb1.save('Main_Schedule_MS.xlsx')
        
        
        try:
            # Save the workbook to the file system
            wb1.save('Main_Schedule_MS.xlsx')

            # Function to save the workbook to a BytesIO object
            def save_to_bytes(wb):
                # Create a BytesIO object to hold the Excel file data in memory
                output = BytesIO()
                wb.save(output)
                output.seek(0)  # Rewind the file pointer to the start
                return output

            # Prepare the workbook for download
            wb_bytes = save_to_bytes(wb1)

            # Create a download button in Streamlit
            st.download_button(
                label="Download Modified Schedule",
                data=wb_bytes,
                file_name="Main_Schedule_MS.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as e:
        st.error(f"Error processing the HOPE_DRIVE sheet: {e}")
